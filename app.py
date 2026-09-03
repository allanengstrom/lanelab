import os
import sys

# Determinism: pin Python's hash seed BEFORE any string-keyed set/dict iteration
# happens elsewhere in the codebase. PYTHONHASHSEED must be set at interpreter
# startup, so if the env var is missing we re-exec ourselves with it set.
if os.environ.get("PYTHONHASHSEED") != "0":
    os.environ["PYTHONHASHSEED"] = "0"
    os.execv(sys.executable, [sys.executable] + sys.argv)

import json
import re
import traceback
from collections import defaultdict

import numpy as np
# Seed numpy's GLOBAL rng once at startup. strategy_robust uses a local seeded
# rng for its CRN bank, but other simulations (relay MC, simulate_match,
# simulate_relay_trials) use np.random.normal without explicit seeding. Pinning
# the global state here makes those reproducible across process restarts.
np.random.seed(42)

from flask import Flask, jsonify, render_template, render_template_string, request, Response

from Results_scraper import find_team_meets, parse_meet_results
from team_names import nvsl_teams
from Optimizer import (
    build_profiles, self_optimal,
    lineup_to_entries,
    save_lineup, mc_event, parse_event, race_points,
    simulate_our_total, score_stats, TARGET_SCORE,
    normalize_name, AGE_GROUP_ORDER,
)
# Pre-import to avoid lazy-load cost on first PDF upload request
try:
    from parse_ladder import parse_pdf as _parse_ladder_pdf
except Exception:
    _parse_ladder_pdf = None

# Coach lineup predictor (replaces self_optimal as opp's "best independent" guess).
# Loads model lazily. Falls back to self_optimal if the .pkl is missing.
try:
    import coach_predictor as _coach_predictor
    _COACH_PREDICTOR_AVAILABLE = True
except Exception as _e:
    _coach_predictor = None
    _COACH_PREDICTOR_AVAILABLE = False
    print(f"[startup] coach_predictor unavailable: {_e}  — falling back to self_optimal", flush=True)


def _predict_opp_lineup_or_fallback(opp_team, year, week, opp_profiles, events):
    """Return the opp's predicted lineup using the trained coach_predictor when
    available; falls back to self_optimal otherwise.

    Caller must supply (year, week); when either is None we fall back too — this
    keeps legacy code paths that don't have that context safe."""
    if (_COACH_PREDICTOR_AVAILABLE and _coach_predictor is not None
            and opp_team and year and week):
        try:
            return _coach_predictor.predict_opp_lineup(
                opp_team, year, week, opp_profiles, events
            )
        except Exception as _e:
            print(f"[coach_predictor] predict failed for {opp_team} W{week}: {_e}", flush=True)
    return self_optimal(opp_profiles, events)


_8U_DIV_PCTS = None
def _load_8u_div_pcts():
    """Per-division 8U fill percentiles (built by build_8u_div_fill.py from 2024
    structure). { '<div>': [p_fast, p_med, p_slow] }. Cached; {} if missing."""
    global _8U_DIV_PCTS
    if _8U_DIV_PCTS is None:
        try:
            with open(os.path.join(BASE_DIR, "w1_8u_div_pcts.json")) as f:
                _8U_DIV_PCTS = json.load(f)
        except Exception:
            _8U_DIV_PCTS = {}
    return _8U_DIV_PCTS


_8U_DIV_AVG = None
def _load_8u_div_avg():
    """Per-division 8U fill tiers from ACTUAL Week-1 averages (build_8u_div_avg.py,
    2021-2024 W1 data): { '<div>': { 'fill_n': N, 'Boys': {stroke:[t1,t2,t3]},
    'Girls': {...} } }. t1/t2/t3 = div-avg #1/#2/#3 with a real falloff, grounded in
    W1 reality (rookie-heavy) rather than percentiles into the seasoned season pool.
    Cached; {} if missing."""
    global _8U_DIV_AVG
    if _8U_DIV_AVG is None:
        try:
            with open(os.path.join(BASE_DIR, "w1_8u_div_avg.json")) as f:
                _8U_DIV_AVG = json.load(f)
        except Exception:
            _8U_DIV_AVG = {}
    return _8U_DIV_AVG


def _hybrid_fill_opp_8u(opp_team, year, opp_lineup, opp_profiles, events):
    """Hybrid 8U opponent field.

    At W1 the v5 predictor under-fields the opponent's 8U (often 1-2 swimmers in
    the technical strokes), and league-avg imputation fills gaps with sandbagged-
    slow times — so our ladder-deep 8U sweeps unrealistically. This keeps the
    opponent's REAL (non-imputed) 8U swimmers and fields the division's actual
    Week-1-average 8U depth (div-avg #1/#2/#3 from build_8u_div_avg.py), then takes
    the fastest `fill_n` of {real ∪ div-avg tiers}. Older bands are untouched.

    Returns (new_lineup, augmented_profiles). Inputs are not mutated — profiles
    are shallow-copied and only synthetic keys are added, so _cache is safe.
    """
    IMPUTE = {"league_avg_active", "league_avg_returner", "prior_year_z"}
    try:
        import w1_predictor as _W1
        div = _W1.get_team_division(opp_team, year)
    except Exception as e:
        print(f"[hybrid8u] skipped ({e})", flush=True)
        return opp_lineup, opp_profiles

    # Per-division 8U fill tiers from ACTUAL Week-1 averages (build_8u_div_avg.py):
    # da[gen][stroke] = [t1, t2, t3] = div-avg #1/#2/#3 (real W1 falloff); fill_n =
    # typical W1 8U turnout so we don't over-fill a thin division. Grounded in W1
    # reality (rookie-heavy) instead of percentiles into the seasoned season pool —
    # which over-rated the top divisions (e.g. D7 #1 19.1 -> 22.0, the real average).
    da = _load_8u_div_avg().get(str(div)) or {}
    # Fill the full 3-lane heat for display. The DQ model (Optimizer._finish_rate)
    # thins the technical strokes IN SCORING (a 3rd fly swimmer finishes ~64% of the
    # time), so expected finishers ≈ the old per-stroke fill_n while the heat shows
    # full. Symmetric — our own 8U technical swimmers get the same finish rate.
    fill_target = 3

    new_lineup = dict(opp_lineup)
    new_prof   = dict(opp_profiles)
    n_filled = 0
    for ev in events:
        if not ev.startswith("8U"):
            continue
        try:
            age, gen, stroke_full = parse_event(ev)
        except Exception:
            continue
        stroke = stroke_full.split("-")[-1]
        # Opponent's REAL (non-imputed) swimmers eligible here, fastest first.
        real = []
        for nm, pr in opp_profiles.items():
            if pr.get("home_age_group") != age or pr.get("gender") != gen:
                continue
            st = pr.get("strokes", {}).get(stroke_full)
            if st and st.get("mean") and st.get("source") not in IMPUTE:
                real.append((st["mean"], nm))
        real.sort()
        # ALWAYS materialize the three division-typical tiers (#1/#2/#3 = fast/med/slow)
        # for this 8U event, and model the opponent's field as the FASTEST `fill_target`
        # of {their real non-imputed swimmers} ∪ {the division tiers}. A division-typical
        # swimmer REPLACES a slower real/imputed roster swimmer (per the operator: real 8U data
        # is sparse/unreliable, and a real team fields ~division-typical depth — faster
        # real swimmers are still kept). Applies to BOTH individual events (the lineup
        # below) and relays (which draw the synthetic tiers from the profile). Validated
        # SHB vs Chesterbrook W1: opp 8U relay model ~77s vs an actual 1:15 (75s).
        div_pool = []   # (time, name) for the synthetic division tiers
        tiers = (da.get(gen) or {}).get(stroke) or []
        for i, t in enumerate(tiers):
            if not t or t <= 0:
                continue
            sname = f"div-avg {gen} {stroke} #{i+1}"
            new_prof[sname] = {"home_age_group": age, "gender": gen,
                               "strokes": {stroke_full: {"mean": float(t),
                                                         "std":  float(t) * 0.05,
                                                         "source": "div_fill"}}}
            div_pool.append((float(t), sname))
            n_filled += 1
        # Field = fastest `fill_target` of real ∪ division tiers.
        pool = sorted(list(real) + div_pool)
        field = [nm for _, nm in pool[:fill_target]]
        if not field:
            continue   # no real swimmers and no division data — leave as-is
        new_lineup[ev] = {"swimmers": field,
                          "expected_points": opp_lineup.get(ev, {}).get("expected_points", 0)}
    if n_filled:
        print(f"[hybrid8u] {opp_team} (div {div}): filled {n_filled} synthetic 8U slot(s)", flush=True)
    return new_lineup, new_prof
from Relay import (
    optimize_age_relays, optimize_mixed_relay,
    monte_carlo_relay, AGE_RELAY_DEFS, MIXED_SLOTS,
)

app = Flask(__name__)
app.secret_key = "nvsl-local"
BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def _ordinal(p):
    """Return ordinal suffix for integer p (handles 0 correctly)."""
    if p is None:
        return ""
    return {1: "st", 2: "nd", 3: "rd"}.get(
        p % 10 if p % 100 not in (11, 12, 13) else 0, "th"
    )


def _normalize_lineup(lineup):
    """
    Normalize swimmer names in a lineup dict to match build_profiles() keys.
    parse_meet_results() returns raw HTML names (may include middle initials);
    build_profiles() strips them via normalize_name(). This ensures lookups
    in opp_profiles succeed during simulation.
    """
    return {
        el: {
            "swimmers": [normalize_name(s) for s in data.get("swimmers", [])],
            "expected_points": data.get("expected_points", 0),
        }
        for el, data in lineup.items()
    }

# Update these dates each season
# NVSL: 5 regular-meet Saturdays starting the 3rd Saturday of June
SEASON_WEEKS = {
    2026: [
        {"week": 1, "date": "20260620", "label": "Week 1 · Jun 20"},
        {"week": 2, "date": "20260627", "label": "Week 2 · Jun 27"},
        {"week": 3, "date": "20260704", "label": "Week 3 · Jul 4"},
        {"week": 4, "date": "20260711", "label": "Week 4 · Jul 11"},
        {"week": 5, "date": "20260718", "label": "Week 5 · Jul 18"},
    ],
    2025: [
        {"week": 1, "date": "20250614", "label": "Week 1 · Jun 14"},
        {"week": 2, "date": "20250621", "label": "Week 2 · Jun 21"},
        {"week": 3, "date": "20250628", "label": "Week 3 · Jun 28"},
        {"week": 4, "date": "20250705", "label": "Week 4 · Jul 5"},
        {"week": 5, "date": "20250712", "label": "Week 5 · Jul 12"},
    ],
    2024: [
        {"week": 1, "date": "20240615", "label": "Week 1 · Jun 15"},
        {"week": 2, "date": "20240622", "label": "Week 2 · Jun 22"},
        {"week": 3, "date": "20240629", "label": "Week 3 · Jun 29"},
        {"week": 4, "date": "20240706", "label": "Week 4 · Jul 6"},
        {"week": 5, "date": "20240713", "label": "Week 5 · Jul 13"},
    ],
}

# In-memory cache for last run — fine for a single-user local tool
_cache = {}

# Disk snapshot of the last optimize so the /lineup page survives a server
# restart (the in-memory _cache is otherwise wiped on every restart).
_LINEUP_SNAPSHOT_PATH = os.path.join(BASE_DIR, "lineup_snapshot.json")

def _eligible_by_event(profiles, event_labels):
    """Map each event label → list of our swimmers who can be entered in it,
    ordered by their time in that stroke (fastest first), with eligible swimmers
    who have NO real time in the stroke listed afterwards (alphabetically).

    Eligibility is by gender + age (current group or one legal swim-up, with the
    8U->9-10 50m block); a missing stroke time no longer excludes a swimmer. This
    lets the Check editor field anyone age/gender-legal, keeps the useful (timed)
    swimmers on top, and — because every entered swimmer stays a valid option —
    stops a no-time swimmer's dropdown from silently collapsing onto another
    swimmer. Per-swimmer guarded so one malformed profile can't blank an event."""
    from Optimizer import AGE_GROUP_ORDER
    out = {}
    if not profiles or not event_labels:
        return out
    for label in event_labels:
        try:
            age, gen, stroke = parse_event(label)
            event_idx = AGE_GROUP_ORDER.index(age)
        except Exception:
            continue
        timed, untimed = [], []
        for n, p in profiles.items():
            try:
                if p.get("gender") != gen:
                    continue
                home = p.get("home_age_group")
                if home not in AGE_GROUP_ORDER:
                    continue
                # current age group or exactly one legal swim-up
                if not (0 <= event_idx - AGE_GROUP_ORDER.index(home) <= 1):
                    continue
                # 8U swimmers don't swim up to a 50m event (distance doubles)
                if home == "8U" and age == "9-10" and stroke.startswith("50-"):
                    continue
                sd = (p.get("strokes") or {}).get(stroke) or {}
                mean = sd.get("mean")
                # imputed/league-avg isn't a real time → rank it with the no-timers
                if mean is not None and not str(sd.get("source", "")).startswith("league_avg"):
                    timed.append((mean, n))
                else:
                    untimed.append(n)
            except Exception:
                continue
        timed.sort(key=lambda t: t[0])
        out[label] = [n for _, n in timed] + sorted(untimed)
    return out

def _no_time_by_event(profiles, lineup_data):
    """Map each event label → our entered swimmers with NO real seed time in that
    stroke (imputed/league-avg/prior-year, or absent). Precomputed at snapshot
    time because raw profiles aren't persisted — without it the /check page, after
    a snapshot rehydrate, has no profiles to read and flags EVERY swimmer."""
    out = {}
    yp = profiles or {}
    if not yp:
        return out
    yp_norm = {normalize_name(k): v for k, v in yp.items()}
    for ev in (lineup_data or {}).get("events", []) or []:
        label = ev.get("event", "")
        try:
            stroke = parse_event(label)[2]
        except Exception:
            stroke = None
        ours = [ln["swimmer"] for ln in (ev.get("lanes_predicted") or [])
                if ln.get("team") == "us" and ln.get("swimmer")]
        nt = []
        if stroke:
            for s in ours:
                prof = yp.get(s) or yp_norm.get(normalize_name(s)) or {}
                sd = (prof.get("strokes") or {}).get(stroke) or {}
                if sd.get("mean") is None or sd.get("source"):
                    nt.append(s)
        out[label] = nt
    return out


def _save_lineup_snapshot():
    """Persist the just-computed lineup so /lineup can rehydrate after a restart."""
    try:
        snap = {
            "lineup_data":       _cache.get("lineup_data"),
            "relay_data":        _cache.get("relay_data"),
            "config":            _cache.get("config"),
            # Precomputed so the Check editor's dropdowns work after a restart
            # (raw profiles aren't kept in the snapshot).
            "eligible_by_event": _eligible_by_event(_cache.get("your_profiles"),
                                                     _cache.get("events")),
            # Same reason: the Check page's "no real time" flags are profile-derived,
            # so precompute them here while profiles are still in memory.
            "no_time_by_event":  _no_time_by_event(_cache.get("your_profiles"),
                                                   _cache.get("lineup_data")),
        }
        with open(_LINEUP_SNAPSHOT_PATH, "w") as f:
            json.dump(snap, f)
    except Exception as e:
        print(f"[snapshot] save failed: {e}", flush=True)

def _load_lineup_snapshot():
    """Rehydrate _cache from the on-disk snapshot. Returns True if a lineup was
    restored. Only fills keys that are currently empty so it never clobbers a
    fresh in-memory run."""
    try:
        if not os.path.exists(_LINEUP_SNAPSHOT_PATH):
            return False
        with open(_LINEUP_SNAPSHOT_PATH) as f:
            snap = json.load(f) or {}
        if not snap.get("lineup_data"):
            return False
        _cache["lineup_data"] = snap["lineup_data"]
        if snap.get("relay_data") is not None and "relay_data" not in _cache:
            _cache["relay_data"] = snap["relay_data"]
        if snap.get("config") is not None and "config" not in _cache:
            _cache["config"] = snap["config"]
        if snap.get("eligible_by_event") and "eligible_by_event" not in _cache:
            _cache["eligible_by_event"] = snap["eligible_by_event"]
        if snap.get("no_time_by_event") and "no_time_by_event" not in _cache:
            _cache["no_time_by_event"] = snap["no_time_by_event"]
        return True
    except Exception as e:
        print(f"[snapshot] load failed: {e}", flush=True)
        return False


# ── Shared HTML pieces ─────────────────────────────────────────────────────────

def page(active, title, body):
    """Renders a page using the shared base.html via templates/_body.html.
    The `body` arg is pre-built HTML; `active` is the nav-highlight key."""
    return render_template("_body.html",
                           page_title=title,
                           active_page=active,
                           body=body)


def _all_team_names():
    """Teams for the picker, taken from the divisions file so the list always
    matches whatever dataset is installed (real league or the demo league).
    Falls back to the static NVSL code table if the file is missing."""
    try:
        d = json.load(open(os.path.join(BASE_DIR, "nvsl_divisions_by_year.json")))
        for y in sorted((k for k in d if k.isdigit()), reverse=True):
            teams = sorted({t for ts in d[y].values() for t in ts})
            if teams:
                return teams
    except Exception:
        pass
    return sorted(nvsl_teams.keys())


def team_options(selected=""):
    # Empty placeholder first so the combobox shows "Select..." until the user picks one
    opts = f'<option value=""{" selected" if not selected else ""}></option>'
    opts += "".join(
        f'<option value="{n}"{" selected" if n == selected else ""}>{n}</option>'
        for n in _all_team_names()
    )
    return opts


def week_options(year, selected_week=None):
    weeks = SEASON_WEEKS.get(year, [])
    opts = '<option value="">select week</option>'
    for w in weeks:
        sel = "selected" if w["week"] == selected_week else ""
        opts += f'<option value="{w["week"]}" {sel}>{w["label"]}</option>'
    return opts


# Official NVSL pool codes (as used on mynvsl.com result sheets), mined from
# nvsl_meet_history.json by majority vote per team. E.g. Sleepy Hollow B & R
# is "SHB", Hamlet is "HSC" — not derivable from the names.
try:
    with open(os.path.join(BASE_DIR, "team_codes.json")) as _f:
        _TEAM_CODES = json.load(_f)
except Exception as _e:
    print(f"[startup] team_codes.json unavailable: {_e} — falling back to initials", flush=True)
    _TEAM_CODES = {}

def team_code(name):
    """Scoreboard code for a team name, used everywhere a matchup is displayed.
    Prefers the official NVSL pool code; falls back to initials for names not
    in the mapping. Full names belong in tooltips only."""
    name = (name or "").strip()
    if name in _TEAM_CODES:
        return _TEAM_CODES[name]
    words = [w for w in re.split(r"[\s\-&]+", name) if w]
    if not words:
        return ""
    if len(words) == 1:
        return words[0][:3].upper()
    return "".join(w[0].upper() for w in words)

app.jinja_env.filters["team_code"] = team_code


# Make helper functions available in all Jinja templates (used by setup.html, etc.)
@app.context_processor
def _inject_template_helpers():
    # has_lineup → show the "Clear lineup" button only when there's actually
    # something to clear (in-memory result or a saved snapshot on disk).
    has_lineup = bool(_cache.get("lineup_data")) or os.path.exists(_LINEUP_SNAPSHOT_PATH)
    # nav_meet → matchup chip in the nav (matchup + week of the current lineup).
    # In-memory only: not worth parsing the snapshot JSON on every request.
    ld = _cache.get("lineup_data") or {}
    nav_meet = ({"you": ld.get("your_team"), "opp": ld["opp_team"],
                 "you_label": team_code(ld.get("your_team")),
                 "opp_label": team_code(ld["opp_team"]),
                 "week": ld.get("week_num")}
                if ld.get("opp_team") else None)
    return dict(team_options=team_options, week_options=week_options,
                has_lineup=has_lineup, nav_meet=nav_meet)


# ── Error pages (themed, instead of Flask's bare white defaults) ──────────────

@app.errorhandler(404)
def _not_found(e):
    if request.path.startswith("/api/"):
        return jsonify({"error": "not found", "path": request.path}), 404
    return render_template("error.html", code=404,
                           heading="Page not found",
                           message="That page doesn't exist — maybe a typo in the URL?"), 404


@app.errorhandler(500)
def _server_error(e):
    if request.path.startswith("/api/"):
        return jsonify({"error": "internal server error"}), 500
    return render_template("error.html", code=500,
                           heading="Something went wrong",
                           message="The server hit an unexpected error. Check the terminal logs for the traceback."), 500


# ── Home ───────────────────────────────────────────────────────────────────────

@app.route("/favicon.ico")
def favicon_ico():
    return app.send_static_file("favicon.png")


@app.route("/", methods=["GET"])
def home():
    """Landing page — description + 'Start a Lineup' CTA."""
    return render_template("home.html", active_page="home")


@app.route("/how-it-works", methods=["GET"])
def how_it_works():
    """Deep-dive explainer for the optimizer pipeline (linked from home)."""
    return render_template("how_it_works.html", active_page="home")


@app.route("/setup", methods=["GET"])
def setup():
    """The 3-step setup flow (formerly the home page)."""
    cfg = _cache.get("config", {})
    return render_template("setup.html",
                           active_page="setup",   # nav highlights "Build"
                           cfg=cfg,
                           year=cfg.get("year", 2026),
                           week=cfg.get("week_num", None))




def _fetch_team_prev_lineup(team_name, week_num, year):
    """
    Find this team's most recent meet before week_num (against anyone) and return
    their normalized lineup as a (lineup_dict, meet_id, week_label) tuple.
    Used for both our team and the opponent.
    """
    weeks      = SEASON_WEEKS.get(year, [])
    prev_weeks = [w for w in weeks if w["week"] < week_num]
    if not prev_weeks:
        raise ValueError(f"No weeks before Week {week_num}.")

    meets = find_team_meets(team_name, prev_weeks, year)
    if not meets:
        raise ValueError(f"No meets found for '{team_name}' before Week {week_num}.")

    prior = max(meets, key=lambda m: m["week"])
    # parse_meet_results extracts swimmers labeled as opp_team_name's team code.
    # For our own team's lineup, treat team_name as "opp_team_name" arg — same logic.
    lineup, code = parse_meet_results(prior["meet_id"], team_name)
    lineup = _normalize_lineup(lineup)
    return lineup, prior["meet_id"], prior["label"]


def _fetch_opp_prev_lineup(opp_team, week_num, year):
    """
    Find the opponent's most recent meet before week_num (against anyone)
    and return their normalized lineup.
    Returns (lineup, meet_id, week_label) or raises on failure.
    """
    weeks      = SEASON_WEEKS.get(year, [])
    prev_weeks = [w for w in weeks if w["week"] < week_num]
    if not prev_weeks:
        raise ValueError(f"No weeks before Week {week_num} — nothing to fetch.")

    opp_meets = find_team_meets(opp_team, prev_weeks, year)
    if not opp_meets:
        raise ValueError(
            f"No meets found for '{opp_team}' before Week {week_num}. "
            "Check the team name or try a later week."
        )

    prior = max(opp_meets, key=lambda m: m["week"])
    opp_lineup, opp_code = parse_meet_results(prior["meet_id"], opp_team)
    opp_lineup = _normalize_lineup(opp_lineup)
    save_lineup(opp_lineup, os.path.join(BASE_DIR, "opp_lineup.json"))
    _cache["opp_actual"]  = opp_lineup
    _cache["opp_meet_id"] = prior["meet_id"]
    return opp_lineup, prior["meet_id"], prior["label"], opp_code


def _safe_team_path(name):
    """Filesystem-safe slug for a team name."""
    return re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")


def _ladder_path(team_name):
    d = os.path.join(BASE_DIR, "time_trials")
    os.makedirs(d, exist_ok=True)
    return os.path.join(d, f"{_safe_team_path(team_name)}.json")


def _aliases_path(team_name):
    d = os.path.join(BASE_DIR, "time_trials")
    os.makedirs(d, exist_ok=True)
    return os.path.join(d, f"{_safe_team_path(team_name)}_aliases.json")


def _reconcile_ladder_to_history(team_name, entries):
    """Rewrite SwimTopia ladder PREFERRED names to the REGISTERED names used in
    meet history. The ladder PDF uses what kids go by ('Nick', 'Fonsi'); meet
    results use the registered name ('Nicholas S Ferrante', 'Alfonso F Marquez').
    The static nickname map catches common ones; this catches the rest by matching
    on (last name, gender), disambiguating siblings by longest-common first-name
    substring (>=3 chars). Without it a swimmer splits into two identities — one
    with ladder data, one stuck in the impute list."""
    if not entries:
        return entries
    from collections import defaultdict
    history = _load_history() or {}
    idx, hist_norms = defaultdict(dict), set()   # (last,gender) -> {norm: raw}
    for y in sorted((k for k in history if k.isdigit()), reverse=True)[:3]:
        for ms in (history[y] or {}).values():
            for m in ms.values():
                for side in ("team_a", "team_b"):
                    if (m.get(side) or {}).get("name") != team_name:
                        continue
                    for ev, d in (m[side].get("lineup") or {}).items():
                        try:
                            _ag, gender, _sf = parse_event(ev)
                        except Exception:
                            continue
                        for sw in (d.get("swimmers") or []):
                            raw = sw.get("name")
                            if not raw:
                                continue
                            norm = normalize_name(raw)
                            hist_norms.add(norm)
                            idx[(norm.split()[-1].lower(), gender)].setdefault(norm, raw)

    def lcs(a, b):
        a, b = a.lower(), b.lower()
        best = 0
        for i in range(len(a)):
            for j in range(i + 3, len(a) + 1):
                if a[i:j] in b:
                    best = max(best, j - i)
        return best

    n = 0
    for e in entries:
        norm = normalize_name(e.get("Name", ""))
        if not norm or norm in hist_norms:
            continue                       # already matches a registered name
        cands = idx.get((norm.split()[-1].lower(), e.get("Gender")))
        if not cands:
            continue
        first = norm.split()[0]
        scored = sorted(((lcs(first, hn.split()[0]), raw)
                         for hn, raw in cands.items()), reverse=True)
        if scored and scored[0][0] >= 3 and (len(scored) == 1 or scored[0][0] > scored[1][0]):
            e["Name"] = scored[0][1]       # rewrite to the registered name
            n += 1
    if n:
        print(f"[ladder] {team_name}: reconciled {n} preferred name(s) to registered names", flush=True)
    return entries


def _team_hist_name_index(team_name):
    """Build a name index from this team's last-3-years meet history:
      idx        : (last_name, gender) -> {norm_name: registered_raw_name}
      hist_norms : set of every normalized name that raced
      ages       : norm_name -> set of age-group prefixes ('8U', '15-18', ...)
      obs        : norm_name -> set of (year:int, age_group) observations
    Shared by the auto-reconciler's logic and the name-flag review."""
    from collections import defaultdict
    history = _load_history() or {}
    idx, hist_norms = defaultdict(dict), set()
    ages, obs = defaultdict(set), defaultdict(set)
    for y in sorted((k for k in history if k.isdigit()), reverse=True)[:3]:
        yi = int(y)
        for ms in (history[y] or {}).values():
            for m in ms.values():
                for side in ("team_a", "team_b"):
                    if (m.get(side) or {}).get("name") != team_name:
                        continue
                    for ev, d in (m[side].get("lineup") or {}).items():
                        try:
                            ag, gender, _sf = parse_event(ev)
                        except Exception:
                            continue
                        for sw in (d.get("swimmers") or []):
                            raw = sw.get("name")
                            if not raw:
                                continue
                            norm = normalize_name(raw)
                            hist_norms.add(norm)
                            idx[(norm.split()[-1].lower(), gender)].setdefault(norm, raw)
                            ages[norm].add(ag)
                            obs[norm].add((yi, ag))
    return idx, hist_norms, ages, obs


_ACTUAL_AGE_CACHE = {}
def _team_actual_ages(team_name):
    """norm_name -> set of (year, nvsl_age) for this team, from every
    leaders_cache*.json on disk (one file per season). NVSL ages are exact season
    ages, so they tell siblings apart precisely AND catch aged-out swimmers
    (projected age >= 19) that the age-GROUP check can't see inside the wide
    15-18 band — e.g. an 18-year-old last year can't be on this year's ladder."""
    if team_name in _ACTUAL_AGE_CACHE:
        return _ACTUAL_AGE_CACHE[team_name]
    import glob
    from collections import defaultdict
    out = defaultdict(set)
    for path in glob.glob(os.path.join(BASE_DIR, "leaders_cache*.json")):
        try:
            with open(path) as f:
                d = json.load(f)
        except Exception:
            continue
        yr = d.get("metadata", {}).get("year")
        if not isinstance(yr, int):
            continue
        team_data = (d.get("teams") or {}).get(team_name) or {}
        for entries in team_data.values():
            for e in entries:
                nm, age = e.get("name"), e.get("age")
                if nm and age is not None:
                    out[normalize_name(nm)].add((yr, age))
    out = dict(out)
    _ACTUAL_AGE_CACHE[team_name] = out
    return out


def _aged_out_norm(team_name, year):
    """Normalized names whose projected NVSL age this season is >= 19 — aged out of
    the league (max competing age is 18), so they CANNOT be on the team and must not
    be fielded. Uses the most recent exact age from results (leaders_cache), projected
    forward one year per season. Definitive — no user review needed (e.g. Sasha Reyes &
    Colin Bramble: 18 in 2025 -> 19 in 2026)."""
    out = set()
    try:
        y = int(year)
    except Exception:
        return out
    for nm, obs in _team_actual_ages(team_name).items():
        if not obs:
            continue
        ly, la = max(obs, key=lambda t: t[0])     # most recent (year, age)
        if la + (y - ly) >= 19:
            out.add(nm)
    return out


_AGE_RANK  = {"8U": 0, "9-10": 1, "11-12": 2, "13-14": 3, "15-18": 4}
_AGE_RANGE = {"8U": (4, 8), "9-10": (9, 10), "11-12": (11, 12),
              "13-14": (13, 14), "15-18": (15, 18)}

def _age_to_band(age):
    """Map an NVSL season age to its age-group band."""
    if age <= 8:  return "8U"
    if age <= 10: return "9-10"
    if age <= 12: return "11-12"
    if age <= 14: return "13-14"
    return "15-18"

def _age_says_different(ladder_age, hist_obs, ladder_year, actual_ages=None, slack=1):
    """True when the ladder swimmer can't be the same person as the history
    swimmer, on age grounds. Kids only age UP (≈1 group/season), so:
      (1) you can't be in a YOUNGER group now than the oldest you ever swam,
      (2) you can't be OLDER than your last race plus the seasons since, and
      (3) if exact NVSL ages are known (from leaders cache), the history swimmer's
          projected age this season must be ≤18 (else aged out of the league) and
          must fit the ladder swimmer's age group.
    hist_obs is a set of (year, age_group); actual_ages a set of (year, nvsl_age).
    Conservative: returns False ('can't rule out — still ask') on missing/unusual
    data, so a real match is never silently hidden."""
    lr     = _AGE_RANK.get(ladder_age)
    lrange = _AGE_RANGE.get(ladder_age)
    if lr is None or lrange is None:
        return False
    # (3) Exact-age test (preferred — real season ages, monotonic year-to-year).
    if actual_ages:
        yr, age = max(actual_ages)              # most recent season on record
        proj = age + (ladder_year - yr)         # their NVSL age this season
        if proj >= 19:
            return True                         # aged out — can't be on the ladder
        if proj < lrange[0] - slack or proj > lrange[1] + slack:
            return True                         # real age doesn't fit this group
    # (1)+(2) Age-GROUP fallback (covers swimmers with no leaders-cache age).
    if hist_obs:
        ranks = [_AGE_RANK[g] for (_y, g) in hist_obs if g in _AGE_RANK]
        if ranks:
            if lr < max(ranks):
                return True
            caps = [_AGE_RANGE[g][1] + (ladder_year - y) + slack
                    for (y, g) in hist_obs if g in _AGE_RANGE]
            if caps and lrange[0] > min(caps):
                return True
    return False


def _flag_name_matches(team_name, year):
    """Ladder swimmers that did NOT auto-merge to a registered name but DO have a
    plausible same-last-name + same-gender history candidate. These are exactly
    the ambiguous cases the auto-reconciler leaves alone (siblings vs. nicknames):
    surface them so the user can confirm 'same person?' once, permanently.

    Skips any pair that's already resolved — saved as an alias ('same person'),
    in the team's confirmed-distinct list, or AGE-INCOMPATIBLE (a kid can't be the
    same person if they'd have to be younger now, or impossibly older). Returns
    (flags, n_auto_resolved): flags sorted most-likely-same first, and a count of
    pairs the age rule resolved silently."""
    entries = _load_ladder_for_team(team_name)   # already alias-applied + reconciled
    if not entries:
        return [], 0
    idx, hist_norms, ages, obs = _team_hist_name_index(team_name)
    actual   = _team_actual_ages(team_name)
    aliases  = _load_aliases(team_name)
    distinct = _load_name_distinct(team_name)

    def lcs(a, b):
        a, b = a.lower(), b.lower()
        best = 0
        for i in range(len(a)):
            for j in range(i + 2, len(a) + 1):
                if a[i:j] in b:
                    best = max(best, j - i)
        return best

    def fmt_ages(s):
        return ", ".join(sorted(s, key=lambda a: _AGE_RANK.get(a, 9))) if s else "?"

    seen, out, n_auto = set(), [], 0
    for e in entries:
        norm = normalize_name(e.get("Name", ""))
        if not norm or norm in hist_norms or norm in seen or norm in aliases:
            continue                       # already a registered name / already merged
        seen.add(norm)
        cands = idx.get((norm.split()[-1].lower(), e.get("Gender")))
        if not cands:
            continue                       # no plausible match -> genuinely a new swimmer
        first    = norm.split()[0]
        lad_age  = e.get("AgeGroup", "?")
        # Rank candidates by first-name overlap, then drop any the age rule rules
        # out as impossible. The best SURVIVING candidate (if any) gets flagged.
        scored = sorted(((lcs(first, hn.split()[0]), hn, raw)
                         for hn, raw in cands.items()), reverse=True)
        viable = [(ov, hn, raw) for (ov, hn, raw) in scored
                  if not _age_says_different(lad_age, obs.get(hn), year, actual.get(hn))]
        if not viable:
            n_auto += 1                    # every candidate is age-impossible -> auto-different
            continue
        ov, hist_norm, hist_raw = viable[0]
        if f"{norm}||{hist_norm}" in distinct:
            continue                       # user already said "different people"
        same_group = lad_age in {g for (_y, g) in obs.get(hist_norm, ())}
        if ov >= 2:
            hint = "names share letters — could be a nickname"
        elif same_group:
            hint = f"both swim {lad_age}, but the names look unrelated"
        else:
            hint = "names look unrelated — likely different swimmers"
        out.append({
            "ladder_name": e.get("Name"),
            "ladder_norm": norm,
            "ladder_age":  lad_age,
            "hist_name":   hist_raw,
            "hist_norm":   hist_norm,
            "hist_age":    fmt_ages(ages.get(hist_norm)),
            "overlap":     ov,
            "hint":        hint,
        })
    # Most-likely-same first (shared letters), then alphabetical.
    out.sort(key=lambda x: (-x["overlap"], x["ladder_name"]))
    return out, n_auto


_LADDER_CACHE = {}
def _load_ladder_for_team(team_name):
    """Load ladder entries (parsed PDF) for a team, with names reconciled to the
    registered forms used in meet history. Returns [] if no file. Cached by mtime."""
    path = _ladder_path(team_name)
    if not os.path.exists(path):
        return []
    try:
        ck = (path, os.path.getmtime(path))
        if ck in _LADDER_CACHE:
            return _LADDER_CACHE[ck]
        with open(path) as f:
            entries = json.load(f).get("entries", [])
        # User-confirmed "same person" merges (saved during setup review) apply
        # first, then the automatic last-name reconciler catches the rest.
        entries = _apply_aliases(entries, _load_aliases(team_name))
        entries = _reconcile_ladder_to_history(team_name, entries)
        _LADDER_CACHE[ck] = entries
        return entries
    except Exception:
        return []


def _build_dated_results(team_results_undated, default_date="2099-01-01"):
    """Convert {event: {swimmer: [times]}} to dated format for recency weighting.
    Scraped times have no real date; we use a future date so they're treated as
    'most recent' (highest weight). Ladder entries (which have real dates) will be
    weighted relative to this — older ladder entries (e.g., time trials) get less
    weight, which is appropriate."""
    return {
        event: {
            swimmer: [(default_date, t) for t in times]
            for swimmer, times in swimmers.items()
        }
        for event, swimmers in team_results_undated.items()
    }


def _merge_ladder_into_dated(dated_results, ladder_entries, tolerance=0.01, max_date=None):
    """Append ladder entries to dated_results, deduping times within tolerance.
    If max_date is provided (ISO YYYY-MM-DD), entries with date >= max_date
    are skipped — prevents future-data leakage when an uploaded ladder PDF
    includes races from after the meet we're predicting.
    Returns (added, skipped, future_filtered) counts."""
    added = 0
    skipped = 0
    future = 0
    for e in ladder_entries:
        event_label = f"{e['AgeGroup']} {e['Gender']} {e['Distance']}-{e['Stroke']}"
        name = normalize_name(e['Name'])
        t    = e['Time']
        date = e['Date']
        if max_date and date and date >= max_date:
            future += 1
            continue
        ev_dict = dated_results.setdefault(event_label, {})
        existing = ev_dict.setdefault(name, [])
        if any(abs(et - t) < tolerance for _, et in existing):
            skipped += 1
            continue
        existing.append((date, t))
        added += 1
    return added, skipped, future


def _load_aliases(team_name):
    """Load alias map for a team. Returns {} if no file.
    Strips keys starting with '_' (treated as comments / metadata)."""
    path = _aliases_path(team_name)
    if not os.path.exists(path):
        return {}
    try:
        with open(path) as f:
            raw = json.load(f)
        return {k: v for k, v in raw.items() if not k.startswith("_")}
    except Exception:
        return {}


def _apply_aliases(entries, aliases):
    """Rewrite entry names according to the alias map (and normalize middle inits)."""
    if not aliases:
        # Still normalize middle initials at minimum
        for e in entries:
            e["Name"] = normalize_name(e["Name"])
        return entries
    for e in entries:
        canonical = aliases.get(e["Name"], e["Name"])
        e["Name"] = normalize_name(canonical)
    return entries


def _load_name_distinct(team_name):
    """Set of 'ladder_norm||hist_norm' pairs the user confirmed are DIFFERENT
    people (siblings, etc.). Stored under the '_distinct' key of the alias file
    so it never collides with the alias map (_load_aliases strips '_' keys)."""
    path = _aliases_path(team_name)
    if not os.path.exists(path):
        return set()
    try:
        with open(path) as f:
            raw = json.load(f)
        return set(raw.get("_distinct", []) or [])
    except Exception:
        return set()


def _save_name_decisions(team_name, decisions):
    """Persist 'same person' / 'different' answers to the team alias file.
    decisions: list of {ladder_norm, hist_norm, hist_name, same: bool}.
      same=True  -> alias ladder_norm -> registered hist_name (merge identities)
      same=False -> record 'ladder_norm||hist_norm' as confirmed-distinct
    Returns (n_merged, n_distinct_total). Busts the ladder cache so the next
    load re-applies the new aliases."""
    path = _aliases_path(team_name)
    raw = {}
    if os.path.exists(path):
        try:
            with open(path) as f:
                raw = json.load(f)
        except Exception:
            raw = {}
    distinct = set(raw.get("_distinct", []) or [])
    merged = 0
    for d in (decisions or []):
        ln = (d.get("ladder_norm") or "").strip()
        hn = (d.get("hist_norm") or "").strip()
        if not ln or not hn:
            continue
        if d.get("same"):
            raw[ln] = d.get("hist_name") or hn     # alias to the registered name
            distinct.discard(f"{ln}||{hn}")
            merged += 1
        else:
            distinct.add(f"{ln}||{hn}")
            raw.pop(ln, None)                       # undo any prior merge
    raw["_distinct"] = sorted(distinct)
    raw.setdefault("_comment",
                   "Name reconciliation. Top-level keys map a ladder name to the "
                   "registered name (same person). _distinct lists confirmed "
                   "different-people pairs ('ladder_norm||hist_norm').")
    tmp = path + ".tmp"
    with open(tmp, "w") as f:
        json.dump(raw, f, indent=2)
    os.replace(tmp, path)
    _LADDER_CACHE.clear()        # force aliases to re-apply on next ladder load
    return merged, len(distinct)


_HISTORY_CACHE = {}  # year_str -> data dict; loaded once per process


# ── Imputation (prior-year-z-score for missing stroke profiles) ──────────────

def _roster_excludes_path(team_name):
    d = os.path.join(BASE_DIR, "roster_excludes")
    os.makedirs(d, exist_ok=True)
    return os.path.join(d, f"{_safe_team_path(team_name)}.json")


def _meet_prefs_path(team_name, year, week):
    d = os.path.join(BASE_DIR, "meet_prefs")
    os.makedirs(d, exist_ok=True)
    return os.path.join(d, f"{_safe_team_path(team_name)}_{year}_w{week}.json")


def _load_meet_prefs(team_name, year, week):
    """Per-(team, year, week) saved setup — absences + the imputation toggle. Scoped
    per week so a meet's choices stick without bleeding across weeks (week 1's
    absences aren't week 2's). Roster excludes ('left the team') stay per-team."""
    try:
        with open(_meet_prefs_path(team_name, year, week)) as f:
            return json.load(f) or {}
    except Exception:
        return {}


def _save_meet_prefs(team_name, year, week, **fields):
    """Merge-save the given fields (skips None) for this team+week."""
    try:
        cur = _load_meet_prefs(team_name, year, week)
        cur.update({k: v for k, v in fields.items() if v is not None})
        path = _meet_prefs_path(team_name, year, week)
        tmp = path + ".tmp"
        with open(tmp, "w") as f:
            json.dump(cur, f, indent=2)
        os.replace(tmp, path)
    except Exception as e:
        print(f"[meet_prefs] save failed for {team_name} {year} W{week}: {e}", flush=True)


def _load_roster_excludes(team_name):
    """Load the set of normalized swimmer names excluded from imputation for this team."""
    path = _roster_excludes_path(team_name)
    if not os.path.exists(path):
        return set()
    try:
        with open(path) as f:
            data = json.load(f) or {}
        return {normalize_name(n) for n in data.get("excludes", [])}
    except Exception:
        return set()


def _save_roster_excludes(team_name, excludes):
    """Persist exclusion list for this team."""
    path = _roster_excludes_path(team_name)
    payload = {
        "team":      team_name,
        "excludes":  sorted({normalize_name(n) for n in excludes}),
        "saved_at":  __import__("datetime").datetime.now().isoformat(timespec="seconds"),
    }
    tmp = path + ".tmp"
    with open(tmp, "w") as f:
        json.dump(payload, f, indent=2)
    os.replace(tmp, path)


# ── Manually-entered times ──────────────────────────────────────────────────────
# For a real swim the data doesn't have — most often a DQ'd race, whose time is legal
# but gets dropped because the swim was disqualified. The coach enters it on the Build
# page and it's merged in exactly like an uploaded-ladder time, so the optimizer can
# field the swimmer on their true speed instead of imputing them.

def _manual_times_path(team_name):
    d = os.path.join(BASE_DIR, "manual_times")
    os.makedirs(d, exist_ok=True)
    return os.path.join(d, f"{_safe_team_path(team_name)}.json")


def _load_manual_times(team_name):
    """Return the list of {name, event, time, time_sec} entries for this team."""
    path = _manual_times_path(team_name)
    if not os.path.exists(path):
        return []
    try:
        with open(path) as f:
            return (json.load(f) or {}).get("times", [])
    except Exception:
        return []


def _save_manual_times(team_name, times):
    path = _manual_times_path(team_name)
    payload = {"team": team_name, "times": times,
               "saved_at": __import__("datetime").datetime.now().isoformat(timespec="seconds")}
    tmp = path + ".tmp"
    with open(tmp, "w") as f:
        json.dump(payload, f, indent=2)
    os.replace(tmp, path)


def _parse_time_str(s):
    """'33.5' or '1:05.21' -> seconds (float); None if unparseable."""
    s = str(s or "").strip()
    if not s:
        return None
    try:
        if ":" in s:
            mins, rest = s.split(":", 1)
            return float(mins) * 60.0 + float(rest)
        return float(s)
    except ValueError:
        return None


def _manual_times_as_ladder(team_name):
    """Manual times formatted as _merge_ladder_into_dated() entries so they merge in
    on the SAME 'real time' path as an uploaded ladder."""
    out = []
    for t in _load_manual_times(team_name):
        ev = t.get("event", "")
        parts = ev.split()
        if len(parts) < 3 or "-" not in parts[-1]:
            continue
        dist, stroke = parts[-1].split("-", 1)
        sec = t.get("time_sec")
        if sec is None:
            sec = _parse_time_str(t.get("time"))
        if sec is None:
            continue
        # Date is used for recency weighting + year parsing; a manual time is a
        # current-season swim, so stamp today when none was given.
        date = t.get("date") or __import__("datetime").date.today().isoformat()
        out.append({"AgeGroup": parts[0], "Gender": parts[1], "Distance": dist,
                    "Stroke": stroke, "Name": t.get("name", ""),
                    "Time": float(sec), "Date": date})
    return out


_LEAGUE_BASELINES_CACHE = {}  # year -> {event: {mean, std}}
_TEAM_YEAR_TIMES_CACHE  = {}  # (team_name, year) -> {(swimmer_norm, event): [times]}


def _team_year_times_index(team_name, year):
    """Lazy-built per-team-per-year index: (swimmer_norm, event) -> [time_sec, ...].
    Replaces full-history scans in _lookup_prior_year_times / _team_swimmers_*.
    First call per (team, year) is O(meets); subsequent calls are O(1)."""
    key = (team_name, year)
    if key in _TEAM_YEAR_TIMES_CACHE:
        return _TEAM_YEAR_TIMES_CACHE[key]
    history = _load_history() or {}
    idx = defaultdict(list)
    for wl, meets in (history.get(str(year)) or {}).items():
        for m in meets.values():
            for side in ("team_a", "team_b"):
                if m[side].get("name") != team_name:
                    continue
                for ev, d in m[side].get("lineup", {}).items():
                    for s in d.get("swimmers", []):
                        t = s.get("time_sec")
                        if t and 5 < t < 600:
                            idx[(normalize_name(s["name"]), ev)].append(t)
    _TEAM_YEAR_TIMES_CACHE[key] = idx
    return idx


def _build_league_baselines(year):
    """Compute league-wide per-event mean/std from all teams' meet history."""
    if year in _LEAGUE_BASELINES_CACHE:
        return _LEAGUE_BASELINES_CACHE[year]
    import numpy as np
    history = _load_history()
    times_by_event = defaultdict(list)
    for wl, meets in (history.get(str(year)) or {}).items():
        for m in meets.values():
            for side in ("team_a", "team_b"):
                for ev, d in m[side].get("lineup", {}).items():
                    for s in d.get("swimmers", []):
                        t = s.get("time_sec")
                        if t and 5 < t < 600:
                            times_by_event[ev].append(t)
    baselines = {ev: {"mean": float(np.mean(ts)), "std": float(np.std(ts))}
                 for ev, ts in times_by_event.items() if ts}
    _LEAGUE_BASELINES_CACHE[year] = baselines
    return baselines


def _all_league_baselines():
    """Return all years' baselines (used for z-score lookup)."""
    history = _load_history()
    return {int(y): _build_league_baselines(int(y)) for y in history.keys()}


def _lookup_prior_year_times(team_name, swimmer_norm_name, event, current_year, max_back=5):
    """Find the swimmer's prior times for this STROKE to project a current estimate — but
    only from swims that still represent their CURRENT group:
      • the SAME age group (any of the last few seasons), or
      • the band ONE below, but only from the immediately prior season — i.e. they aged
        up into this group just last year.
    A swim two-plus bands down, or a lone stale swim from several seasons ago, is not a
    fair basis for a current projection (e.g. a single 13-14 fly from three years back),
    so it's ignored and the swimmer falls through to no prior-year estimate. The caller
    z-scores each returned time against THAT band/year's baseline so the projection is
    field-relative (handles growth and the 25m->50m distance jump)."""
    from Optimizer import parse_event, AGE_GROUP_ORDER
    try:
        target_age, gender, stroke = parse_event(event)
        stroke_type = stroke.split("-")[-1]
        target_idx = AGE_GROUP_ORDER.index(target_age)
    except Exception:
        return []
    for year in range(current_year - 1, current_year - 1 - max_back, -1):
        idx = _team_year_times_index(team_name, year)
        hits = []
        for (sw, ev), times in idx.items():
            if sw != swimmer_norm_name:
                continue
            try:
                a2, g2, st2 = parse_event(ev)
                b2 = AGE_GROUP_ORDER.index(a2)
            except Exception:
                continue
            if g2 != gender or st2.split("-")[-1] != stroke_type:
                continue
            same_band = (b2 == target_idx)
            aged_up_last_year = (b2 == target_idx - 1 and year == current_year - 1)
            if not (same_band or aged_up_last_year):
                continue   # stale / too far down to represent current form
            for t in times:
                hits.append({"year": year, "time": t, "event": ev})
        if hits:
            return hits   # most-recent qualifying season — closest to current form
    return []


# Extra slowness (in league-baseline std units) added to league-average FILL entries
# (swimmers with no real data). Bigger = fills place lower and bank fewer of the phantom
# points that inflate thin / weak-division teams (impute_forfeit_test: weak-div bias
# +10.9 with full-strength fills vs -1.4 with none). Default 0 = unchanged; the
# calibrated value is set after the sweep. Overridable in-process for the sweep.
_FILL_SLOWNESS_BONUS = float(os.environ.get("FILL_SLOWNESS", "0.0"))


def _impute_swimmer_event(team_name, swimmer_norm_name, event, current_year,
                           baselines_all_years, is_active_2025=False, team_z=0.0):
    """
    Return {"mean", "std", "n_times", "best", "source"} or None for one swimmer/event.

    For swimmers with prior-year data for THIS event: use z-score → accurate.
    For swimmers without that data, apply a slowness penalty (in std deviations
    of the event's league baseline) — they're a fill-the-slot option, not a primary pick:

      is_active_2025 = True  (raced any 2025 event):  +1.0 std slower than league avg
      is_active_2025 = False (only prior years):      +1.5 std slower than league avg

    team_z (TEAM_COND_IMPUTE): the team's median observed z. Anchors the fallback
    to the TEAM's level — unknown kids are slow *for their team*, not league-mean
    slow. League-mean anchoring (team_z=0) compresses every opponent toward the
    middle: elite opponents' fill entries come out too slow (we over-predict
    ourselves +17 vs elite) and weak opponents' too fast (−21 vs weak) — the
    opponent-strength staircase measured on the 2026-06-10 production baseline.
    """
    import numpy as np
    cur_base = baselines_all_years.get(current_year, {}).get(event, {})
    if not cur_base.get("mean") or not cur_base.get("std", 0):
        return None
    prior = _lookup_prior_year_times(team_name, swimmer_norm_name, event, current_year)
    if prior:
        z_scores = []
        for entry in prior:
            # z against the band the time was ACTUALLY swum in (may be a younger band than
            # the current event, since the swimmer aged up), then projected to cur_base below.
            grp = baselines_all_years.get(entry["year"], {}).get(entry.get("event", event), {})
            if grp.get("std", 0) > 0:
                z = (entry["time"] - grp["mean"]) / grp["std"]
                z_scores.append(z)
        if z_scores:
            avg_z = float(np.mean(z_scores))
            pred_time = cur_base["mean"] + avg_z * cur_base["std"]
            return {"mean":     pred_time,
                    "std":      max(cur_base["std"] * 0.3, 0.2),
                    "n_times":  len(z_scores),
                    "best":     pred_time,
                    "source":   "prior_year_z"}

    # Fallback: team-anchored (or league-avg when team_z=0) tiered slowness penalty
    penalty_std = (1.0 if is_active_2025 else 1.5) + _FILL_SLOWNESS_BONUS
    pred_time = cur_base["mean"] + (team_z + penalty_std) * cur_base["std"]
    return {"mean":     pred_time,
            "std":      cur_base["std"] * 0.5,
            "n_times":  0,
            "best":     pred_time,
            "source":   "league_avg_active" if is_active_2025 else "league_avg_returner"}


def _team_swimmers_2025(team_name, year):
    """Set of normalized swimmer names who raced for this team in any current-year meet.

    When a SwimTopia roster is available for the team, treat the roster as
    authoritative for "currently active" — useful pre-season (no race data yet)
    and for kids who race once-a-season."""
    import roster as _roster
    roster_set = _roster.roster_swimmer_set(team_name, year)
    if roster_set is not None:
        return roster_set
    idx = _team_year_times_index(team_name, year)
    return {sw for (sw, _ev) in idx.keys()}


def _team_swimmers_all_prior(team_name, year, max_back=1):
    """Set of normalized swimmer names who raced for this team in the immediately prior year."""
    out = set()
    for y in range(year - 1, year - 1 - max_back, -1):
        idx = _team_year_times_index(team_name, y)
        out |= {sw for (sw, _ev) in idx.keys()}
    return out


def _leaders_swimmers_with_max_age(team_name, year):
    """{sw_norm: max_age_seen_in_current_year_leaders} for this team in current year.
    Returns {} if leaders cache is missing or doesn't have this team / year."""
    cache = _load_leaders_cache()
    if not cache:
        return {}
    # leaders_cache.json has metadata.year — only use if it matches
    if cache.get("metadata", {}).get("year") != year:
        return {}
    team_data = cache.get("teams", {}).get(team_name)
    if not team_data:
        return {}
    out = {}
    for ev, entries in team_data.items():
        for e in entries:
            nm = e.get("name")
            age = e.get("age")
            if not nm:
                continue
            sw_norm = normalize_name(nm)
            if age is not None:
                cur = out.get(sw_norm)
                if cur is None or age > cur:
                    out[sw_norm] = age
            else:
                out.setdefault(sw_norm, None)
    return out


def _first_year_in_age_group(team_name, sw_norm, age_group_prefix):
    """Earliest year (across all available history) this swimmer raced for this
    team in any event starting with `age_group_prefix` (e.g. '15-18'). Returns
    None if no such race exists. Used for the aged-out heuristic."""
    history = _load_history() or {}
    earliest = None
    for y_str, weeks in history.items():
        try: y = int(y_str)
        except ValueError: continue
        for wl, meets in weeks.items():
            for m in meets.values():
                for side in ("team_a", "team_b"):
                    if m[side].get("name") != team_name:
                        continue
                    for ev, d in m[side].get("lineup", {}).items():
                        if not ev.startswith(age_group_prefix):
                            continue
                        for s in d.get("swimmers", []):
                            if normalize_name(s["name"]) == sw_norm:
                                if earliest is None or y < earliest:
                                    earliest = y
                                break
    return earliest


def _swimmers_needing_review(team_name, year):
    """
    Swimmers who would be imputed from prior years but have NO current-year data.
    Returns list of dicts: {name, last_raced_year, age_group, n_prior_races}

    If a SwimTopia roster is available for this team, the review list is
    restricted to roster members — automatically drops aged-out kids and
    quitters whose race history would otherwise show up here.
    """
    import roster as _roster
    history     = _load_history()
    # Use race-only "active" set (not the roster-augmented one) — for the review,
    # we want to know who has RACE DATA this year, not just who's on the roster.
    # Augment with leaders_cache (catches swimmers who raced in meets we didn't
    # scrape — B-meets, time trials, etc.) and gives us authoritative ages.
    idx         = _team_year_times_index(team_name, year)
    cur_set     = {sw for (sw, _ev) in idx.keys()}
    leaders_ages = _leaders_swimmers_with_max_age(team_name, year)
    cur_set |= set(leaders_ages.keys())
    # An uploaded ladder IS current-year data — its swimmers have raced this season
    # (time trials / B-meets), so they don't belong in the "would be imputed" list.
    ladder_active = {normalize_name(e["Name"])
                     for e in (_load_ladder_for_team(team_name) or [])
                     if e.get("Name")}
    cur_set |= ladder_active
    aged_out    = _aged_out_norm(team_name, year)   # projected age >= 19: auto-dropped, no review
    roster_set  = _roster.roster_swimmer_set(team_name, year)
    # All swimmers who raced for this team in the immediately prior year
    swimmers    = defaultdict(lambda: {"years": [], "last_age": None, "n": 0})
    for y in range(year - 1, year - 2, -1):
        for wl, meets in (history.get(str(y)) or {}).items():
            for m in meets.values():
                for side in ("team_a", "team_b"):
                    if m[side].get("name") != team_name:
                        continue
                    for ev, d in m[side].get("lineup", {}).items():
                        age_grp = ev.split()[0] if ev else "?"
                        gender  = ev.split()[1] if " " in ev else "?"
                        for s in d.get("swimmers", []):
                            norm = normalize_name(s["name"])
                            swimmers[norm]["years"].append(y)
                            swimmers[norm]["last_age"] = f"{age_grp} {gender}"
                            swimmers[norm]["raw_name"] = s["name"]
                            swimmers[norm]["n"] += 1
    out = []
    for norm, info in swimmers.items():
        if norm in cur_set:
            continue   # has current-year data — no review needed
        if norm in aged_out:
            continue   # aged out (projected age >= 19) — auto-dropped from fielding, no review
        if not info["years"]:
            continue
        if roster_set is not None and norm not in roster_set:
            continue   # not on the current-year roster — aged out, quit, etc.
        # Aged-out heuristic for teams WITHOUT a roster.
        # NVSL caps at age 18 (as-of-June-1). Only swimmers who'd be ≥19 in the
        # current season have aged out. We infer this from career trajectory:
        #   First year in 15-18 group → was ≥15 then → now ≥(15 + years since)
        # So if their first 15-18 appearance was ≥4 years ago, they're ≥19 now.
        # 17- and 18-year-olds in 2024 are still eligible for 2025, so we keep
        # them in the review.
        if roster_set is None and info.get("last_age", "").startswith("15-18"):
            first_1518 = _first_year_in_age_group(team_name, norm, "15-18")
            if first_1518 is not None and (year - first_1518) >= 4:
                continue
        out.append({
            "name":          info.get("raw_name", norm),
            "norm_name":     norm,
            "last_raced":    max(info["years"]),
            "last_age_grp":  info["last_age"],
            "n_prior_races": info["n"],
        })

    # Note: roster newcomers (on team, zero race history) are NOT added to the
    # review list. There's no user judgment to apply for them — they're on the
    # roster (so they're on the team) and they'll be imputed via league-avg.
    # Showing 60+ newcomers would clutter the list without giving the user any
    # actionable decision to make.
    # Sort: most recent first, then by name. Roster newcomers (last_raced=None)
    # come last (treated as last_raced = 0).
    out.sort(key=lambda x: (-(x["last_raced"] or 0), x["name"]))
    return out


def augment_profiles_with_imputation(profiles, team_name, year,
                                      excludes_norm=None, baselines=None,
                                      impute_8u_cross=True):
    """
    Augment a team's profiles with prior-year-z-score imputation for swimmers
    who lack current-year data. Returns (enhanced_profiles, n_prior_added, n_league_added).

    excludes_norm: set of normalized swimmer names to EXCLUDE from imputation
                   (e.g., aged out, quit team).
    baselines:     {year: {event: {mean, std}}} — computed if None.

    If a SwimTopia roster file exists for this team (see roster.py), it overrides
    the race-history-derived active set, giving authoritative truth for who's
    actually on the team this year (especially valuable pre-season when no
    current-year race data exists yet).
    """
    from copy import deepcopy
    import roster as _roster
    excludes_norm = set(excludes_norm or [])
    baselines = baselines or _all_league_baselines()
    enhanced = deepcopy(profiles)

    # Set of swimmers who raced for this team in the current year (active roster).
    # Used to apply tiered league-avg penalty: active swimmers get +1.0 std,
    # prior-year-only swimmers get +1.5 std (less trusted in new events).
    active_2025 = _team_swimmers_2025(team_name, year)

    # Candidate pool = everyone likely on the team this season:
    #   - the uploaded ladder (already raced this year — time trials / B / A meets),
    #   - this-year race data, AND
    #   - immediately-prior-year RETURNERS — so kids who haven't done time trials
    #     yet still appear and get imputed (the ladder only lists who's raced).
    # A year-matched SwimTopia roster, if present, is the gold standard that trims
    # quitters/aged-out kids; the ladder is ALWAYS kept (it's ground truth).
    ladder_active = {normalize_name(e["Name"])
                     for e in (_load_ladder_for_team(team_name) or [])
                     if e.get("Name")}
    active_2025 = set(active_2025) | ladder_active   # ladder swimmers count as active
    roster_set     = _roster.roster_swimmer_set(team_name, year)
    roster_ages    = _roster.roster_age_groups(team_name, year)
    roster_genders = _roster.roster_genders(team_name, year)
    if roster_set is not None:
        all_team_swimmers = set(roster_set) | ladder_active
        print(f"[roster] {team_name}: SwimTopia roster ({len(roster_set)}) + ladder ({len(ladder_active)}) as pool", flush=True)
        # DROP race-history swimmers on NEITHER the roster nor the ladder (quit/aged out).
        n_dropped = 0
        for sw_raw in list(enhanced.keys()):
            if normalize_name(sw_raw) not in all_team_swimmers:
                del enhanced[sw_raw]
                n_dropped += 1
        if n_dropped:
            print(f"[roster] {team_name}: dropped {n_dropped} swimmer(s) not on roster or ladder", flush=True)
        # Override stale home_age_group AND gender with roster's authoritative data.
        n_corrected_age = n_corrected_gender = 0
        for sw_raw in list(enhanced.keys()):
            sw_norm = normalize_name(sw_raw)
            roster_age    = roster_ages.get(sw_norm) if roster_ages else None
            roster_gender = roster_genders.get(sw_norm) if roster_genders else None
            if roster_age and enhanced[sw_raw].get("home_age_group") != roster_age:
                enhanced[sw_raw]["home_age_group"] = roster_age
                n_corrected_age += 1
            if roster_gender and enhanced[sw_raw].get("gender") != roster_gender:
                enhanced[sw_raw]["gender"] = roster_gender
                n_corrected_gender += 1
        if n_corrected_age:
            print(f"[roster] {team_name}: corrected home_age_group for {n_corrected_age} swimmer(s)", flush=True)
        if n_corrected_gender:
            print(f"[roster] {team_name}: corrected gender for {n_corrected_gender} swimmer(s) (race-history conflict)", flush=True)
    else:
        # No roster: ladder (current) ∪ this-year racers ∪ prior-year returners.
        all_team_swimmers = (ladder_active | set(active_2025)
                             | _team_swimmers_all_prior(team_name, year, max_back=1))

    all_events = set()
    for y_evs in baselines.values():
        all_events |= set(y_evs.keys())

    # TEAM_COND_IMPUTE=1: anchor league_avg_* fills to the team's observed median z
    # (see _impute_swimmer_event docstring). Off by default until validated.
    team_z = 0.0
    if os.environ.get("TEAM_COND_IMPUTE", "0") == "1":
        import statistics as _stats
        cur_b = baselines.get(year, {})
        zs = []
        for _sw, p in enhanced.items():
            band, g = p.get("home_age_group"), p.get("gender")
            if not band or not g:
                continue
            for sk, s in (p.get("strokes") or {}).items():
                if s.get("source") is not None or s.get("mean") is None:
                    continue                      # only real observed entries
                b = cur_b.get(f"{band} {g} {sk}")
                if b and b.get("std", 0) > 0:
                    zs.append((s["mean"] - b["mean"]) / b["std"])
        if len(zs) >= 8:                          # too few observations → league anchor
            team_z = float(_stats.median(zs))
            print(f"[impute] {team_name}: team-conditioned anchor z={team_z:+.2f} "
                  f"(n={len(zs)} observed entries)", flush=True)

    n_prior_added = n_league_added = 0
    for swimmer in all_team_swimmers:
        if swimmer in excludes_norm:
            continue
        is_active = swimmer in active_2025
        # Authoritative gender + age from roster, when available — used for both
        # profile-shell creation AND eligibility filtering inside the event loop.
        sw_roster_gender = roster_genders.get(swimmer) if roster_genders else None
        sw_roster_age    = roster_ages.get(swimmer)    if roster_ages    else None
        for event in all_events:
            age_group, gender, stroke = parse_event(event)
            # 8U cross-stroke imputation off (default in production): don't fabricate a
            # time for an 8U swimmer in a stroke they've never swum. 8U kids are new and
            # breast/fly are DQ-prone — fielding a no-time 8U entry makes no coaching sense
            # and isn't backed by the ladder. (Real prior 8U times are already in the
            # profile, so this only blocks fabricated entries, never legitimate ones.)
            if not impute_8u_cross and age_group == "8U":
                continue
            # Hard skip: event gender or age doesn't match this swimmer per roster.
            # Prevents girls being slotted in Boys events (and vice versa) when
            # they have no race data to anchor the profile.
            if sw_roster_gender and gender != sw_roster_gender:
                continue
            if sw_roster_age and age_group != sw_roster_age:
                continue
            # Create profile shell. Prefer roster gender/age when available,
            # else fall back to event labels (race-history-only flow).
            if swimmer not in enhanced:
                enhanced[swimmer] = {"home_age_group": sw_roster_age or age_group,
                                     "gender":         sw_roster_gender or gender,
                                     "strokes":        {}}
            elif "gender" not in enhanced[swimmer]:
                enhanced[swimmer]["gender"] = sw_roster_gender or gender
            elif "home_age_group" not in enhanced[swimmer]:
                enhanced[swimmer]["home_age_group"] = sw_roster_age or age_group

            if enhanced[swimmer].get("gender") != gender: continue
            if enhanced[swimmer].get("home_age_group") != age_group: continue
            if stroke in enhanced[swimmer].get("strokes", {}): continue
            imp = _impute_swimmer_event(team_name, swimmer, event, year, baselines,
                                         is_active_2025=is_active, team_z=team_z)
            if imp is None: continue
            enhanced[swimmer]["strokes"][stroke] = imp
            if imp["source"] == "prior_year_z": n_prior_added += 1
            else: n_league_added += 1
    return enhanced, n_prior_added, n_league_added


def _load_leaders_cache():
    """Load leaders_cache.json (in-memory cached). Returns None if missing."""
    path = os.path.join(BASE_DIR, "leaders_cache.json")
    if "_leaders" in _HISTORY_CACHE:
        return _HISTORY_CACHE["_leaders"]
    if not os.path.exists(path):
        return None
    try:
        with open(path) as f:
            data = json.load(f)
        _HISTORY_CACHE["_leaders"] = data
        return data
    except Exception:
        return None


def _build_team_dated_from_leaders(team_name, year, max_date=None):
    """
    Build dated team_results from leaders_cache.json, filtering by date.
    max_date: ISO 'YYYY-MM-DD'. Only include times STRICTLY BEFORE this date.
    (We're predicting the meet on max_date, so don't leak its own times.)
    Returns ({event: {swimmer: [(date, time), ...]}}, sorted_unique_dates) or (None, []).
    """
    cache = _load_leaders_cache()
    if not cache:
        return None, []
    # The leaders cache is scoped to a single season (metadata.year). Using it
    # for a different year leaks 2025 race data into 2024 evaluations — and the
    # max_date filter only partially catches it (entries with dates earlier than
    # max_date slip through for some teams but not others, producing seemingly-
    # random divisions where some teams predict normally and others predict
    # near-zero). Mirrors the same year-match check in _leaders_swimmers_with_max_age.
    if cache.get("metadata", {}).get("year") != year:
        return None, []
    # Case-insensitive team lookup
    team_data = None
    for k, v in cache.get("teams", {}).items():
        if k.lower() == team_name.lower():
            team_data = v
            break
    if not team_data:
        return None, []
    out   = defaultdict(lambda: defaultdict(list))
    dates = set()
    for event_label, entries in team_data.items():
        for e in entries:
            t = e.get("time"); d = e.get("date"); n = e.get("name")
            if t is None or d is None or n is None:
                continue
            if max_date and d >= max_date:
                continue
            out[event_label][normalize_name(n)].append((d, t))
            dates.add(d)
    if not out:
        return None, []
    return {k: dict(v) for k, v in out.items()}, sorted(dates)


def _date_for_week(year, week_num):
    """Return ISO 'YYYY-MM-DD' for the start of the given NVSL week."""
    weeks = SEASON_WEEKS.get(year, [])
    for w in weeks:
        if w["week"] == week_num:
            d = w["date"]  # 'YYYYMMDD'
            return f"{d[0:4]}-{d[4:6]}-{d[6:8]}"
    return None


def _load_history():
    """Load nvsl_meet_history.json, cached after first call."""
    path = os.path.join(BASE_DIR, "nvsl_meet_history.json")
    if "_data" in _HISTORY_CACHE:
        return _HISTORY_CACHE["_data"]
    if not os.path.exists(path):
        return None
    try:
        with open(path) as f:
            data = json.load(f)
        # Backtest leakage guard (env HISTORY_MAX_YEAR, off by default): drop
        # season years AFTER the given year so an eval of a past season never
        # sees the future. Production never sets this — at real predict time
        # the future isn't in the file. Validated 2026-08-28: with 2026 data
        # present, a 2025-season backtest ran ~+5 hot from future-year leakage.
        _max_y = os.environ.get("HISTORY_MAX_YEAR")
        if _max_y:
            data = {k: v for k, v in data.items()
                    if not k.isdigit() or int(k) <= int(_max_y)}
        _HISTORY_CACHE["_data"] = data
        return data
    except Exception:
        return None


def _apply_age_up_correction(profiles, team_name, year):
    """Override home_age_group for swimmers who aged up between seasons.

    When we use prior-year history as the W1 baseline, swimmers get tagged with
    their OLD band (e.g., 8U). For 2025 predictions, an 8-year-old in 2024 may
    be 9 (and thus 9-10) in 2025. NVSL's age cutoff is June 1 of the season year.

    Sources of current-season band info (priority order):
      1. SwimTopia roster (nvsl_age_group field) — most authoritative
      2. Ladder (AgeGroup field on uploaded HY-TEK ladder PDF)
      3. Leaders cache (events from this year — band derived from event label)

    For aged-up swimmers, we update home_age_group AND drop their prior-band
    stroke data (since 25-distance times don't directly translate to 50-distance
    events). They effectively become unknowns in the new band — better than
    being placed in the wrong band with confidently-wrong times.

    Returns a NEW profile dict (does not mutate input).
    """
    from Optimizer import AGE_GROUP_ORDER

    if not profiles:
        return profiles

    # Build map: normalized_name -> current_band (highest seen across all signals)
    current_band = {}

    # 1. SwimTopia roster (if available)
    try:
        import roster as _roster_mod
        roster_data = _roster_mod.get_roster(team_name) or {}
        for sw in roster_data.get("swimmers", []):
            nm = normalize_name(sw.get("name", ""))
            band_raw = (sw.get("nvsl_age_group") or "").strip()
            # Could be "11-12 Boys" or just "11-12" — strip the gender suffix
            for b in AGE_GROUP_ORDER:
                if b in band_raw:
                    current_band[nm] = b
                    break
    except Exception:
        pass

    # 2. Ladder (uploaded PDF)
    try:
        ladder_entries = _load_ladder_for_team(team_name) or []
        for e in ladder_entries:
            nm = normalize_name(e.get("Name", ""))
            band = (e.get("AgeGroup") or "").strip()
            if band not in AGE_GROUP_ORDER: continue
            if nm in current_band:
                # Already from SwimTopia — keep the higher of the two
                cur_idx = AGE_GROUP_ORDER.index(current_band[nm])
                new_idx = AGE_GROUP_ORDER.index(band)
                if new_idx > cur_idx:
                    current_band[nm] = band
            else:
                current_band[nm] = band
    except Exception:
        pass

    # 3. Leaders cache (UNFILTERED — we want age info even from "future" weeks)
    try:
        lc = _load_leaders_cache()
        if lc:
            for tname, tdata in lc.get("teams", {}).items():
                if tname.lower() != team_name.lower(): continue
                for event_label, entries in tdata.items():
                    try:
                        from Optimizer import parse_event
                        band, _, _ = parse_event(event_label)
                    except Exception:
                        continue
                    if band not in AGE_GROUP_ORDER: continue
                    for e in entries:
                        nm = normalize_name(e.get("name", ""))
                        if nm in current_band:
                            cur_idx = AGE_GROUP_ORDER.index(current_band[nm])
                            new_idx = AGE_GROUP_ORDER.index(band)
                            if new_idx > cur_idx:
                                current_band[nm] = band
                        else:
                            current_band[nm] = band
                break
    except Exception:
        pass

    # 4. Projected age from stored NVSL ages (leaders cache). The band signals
    #    above only know what band a swimmer currently RACES in — useless for a
    #    returning swimmer who hasn't raced this season (every signal still shows
    #    last year's band). But mynvsl records each swimmer's age every season, so
    #    project the most recent one forward to find their CURRENT band. Catches
    #    kids like Ada Kowalski: age 8 in 2025 -> 9 in 2026 -> 9-10, so she's no
    #    longer eligible for (and stacked into) 8U events.
    proj_age_info = {}     # norm_name -> (from_age, obs_year), for the age curve
    try:
        for nm, obs in _team_actual_ages(team_name).items():
            if not obs:
                continue
            yr, age = max(obs)                     # most recent season on record
            proj_age_info[nm] = (age, yr)
            band = _age_to_band(age + (year - yr))
            if band not in AGE_GROUP_ORDER:
                continue
            if nm in current_band:
                if AGE_GROUP_ORDER.index(band) > AGE_GROUP_ORDER.index(current_band[nm]):
                    current_band[nm] = band
            else:
                current_band[nm] = band
    except Exception:
        pass

    # ── Individual-age improvement curve (gated by USE_AGE_CURVE; see age_curve.py) ─
    import age_curve
    # Swimmers WITH current-season data (uploaded ladder, this-year races, or
    # leaders observed in the target year). The within-band bump is applied ONLY to
    # swimmers WITHOUT current data, so we never improve a time that's already current.
    has_current = set()
    try:
        has_current |= {normalize_name(e["Name"])
                        for e in (_load_ladder_for_team(team_name) or []) if e.get("Name")}
    except Exception:
        pass
    try:
        has_current |= {sw for (sw, _ev) in _team_year_times_index(team_name, year).keys()}
    except Exception:
        pass
    for _nm, (_a, _yr) in proj_age_info.items():
        if _yr >= year:
            has_current.add(_nm)

    def _within_band_improve(nm, p, from_age, to_age):
        """Apply the same-distance age curve to an in-band swimmer's existing
        strokes — e.g. Victoria stays 13-14 but ages 13->14 and gets faster.
        Without this, same-band agers get zero improvement (the early return)."""
        ns = {}
        for sk, stt in (p.get("strokes") or {}).items():
            if not isinstance(sk, str) or "-" not in sk:
                ns[sk] = stt; continue
            ds, stroke = sk.split("-", 1)
            try:
                dist = int(ds)
            except Exception:
                ns[sk] = stt; continue
            m = stt.get("mean")
            if m is None:
                ns[sk] = stt; continue
            be = age_curve.best_ever(nm, stroke, dist)
            proj = age_curve.project_same_dist(m, stroke, dist, from_age, to_age, best_ever_time=be)
            if proj is None:
                ns[sk] = stt; continue
            fac = proj / m
            ns[sk] = {**stt, "mean": proj,
                      "std": max(stt.get("std", 0) * fac, proj * 0.02),
                      "_age_within_band": True, "_age_factor": round(fac, 3)}
        return {**p, "strokes": ns}

    # Apply corrections
    from Optimizer import parse_event as _parse_event
    out = {}
    n_corrected = 0
    for name, p in profiles.items():
        cur_home = p.get("home_age_group")
        target = current_band.get(name)
        nm = normalize_name(name)
        if target is None:
            out[name] = p
            continue
        if cur_home == target:
            # Same band: only the within-band age curve changes anything, and only
            # for a swimmer with NO current-season data (else we'd double-count an
            # already-current time). This is the within-band-improvement (Victoria) fix.
            fa = proj_age_info.get(nm)
            if age_curve.USE_AGE_CURVE and fa and fa[1] < year and nm not in has_current:
                out[name] = _within_band_improve(nm, p, fa[0], fa[0] + (year - fa[1]))
            else:
                out[name] = p
            continue
        # Only correct UPWARD age (don't move kids backward)
        cur_idx = AGE_GROUP_ORDER.index(cur_home) if cur_home in AGE_GROUP_ORDER else 0
        new_idx = AGE_GROUP_ORDER.index(target)
        if new_idx <= cur_idx:
            out[name] = p
            continue

        # Aged up: apply EMPIRICAL year-over-year scaling to prior-band times.
        #
        # Used as a fallback only — if the swimmer had CURRENT-band data
        # anywhere (ladder / SwimTopia / leaders before max_date),
        # build_profiles_recency_weighted would have set home_age_group
        # to the new band and this branch wouldn't trigger.
        #
        # Two scaling cases:
        #   (a) DISTANCE + GROWTH (25→50, e.g. free/back/breast across 8U→9-10):
        #       use combined factors from /tmp/measure_scaling_fast.py
        #       (n=2393, includes 1 year of growth on top of distance change)
        #   (b) GROWTH-ONLY (same distance, e.g. 25-fly across 8U→9-10,
        #       50-* across 9-10→11-12, etc.): use growth ratios from
        #       /tmp/measure_growth_only.py (per from-band/to-band/stroke)
        #
        # For multi-band skips (e.g. 8U→11-12 — uncommon), we compose: scale
        # one band at a time through the chain.
        n_corrected += 1
        DIST_GROWTH_25_TO_50 = {"free": 2.03, "back": 1.97, "breast": 1.96, "fly": 2.12}
        # Growth-only factors per (from_band, to_band, stroke). Measured medians.
        GROWTH_ONLY = {
            ("8U",    "9-10",  "fly"):    0.85,
            ("9-10",  "11-12", "free"):   0.928,
            ("9-10",  "11-12", "back"):   0.924,
            ("9-10",  "11-12", "breast"): 0.919,
            ("11-12", "13-14", "free"):   0.950,
            ("11-12", "13-14", "back"):   0.945,
            ("11-12", "13-14", "breast"): 0.938,
            ("11-12", "13-14", "fly"):    0.931,
            ("13-14", "15-18", "free"):   0.981,
            ("13-14", "15-18", "back"):   0.976,
            ("13-14", "15-18", "breast"): 0.978,
            ("13-14", "15-18", "fly"):    0.974,
        }
        def _growth_factor(from_b, to_b, stroke):
            """Compose growth-only scaling across one or more single-band steps."""
            from Optimizer import AGE_GROUP_ORDER as _ord
            fi, ti = _ord.index(from_b), _ord.index(to_b)
            if ti <= fi: return 1.0
            f = 1.0
            for i in range(fi, ti):
                step = GROWTH_ONLY.get((_ord[i], _ord[i+1], stroke))
                if step is None: step = 0.95  # conservative fallback
                f *= step
            return f

        new_strokes = {}
        for stroke_key, stats in (p.get("strokes") or {}).items():
            if not isinstance(stroke_key, str) or "-" not in stroke_key: continue
            dist_str, stroke = stroke_key.split("-", 1)
            try: dist = int(dist_str)
            except Exception: continue
            # What distance does the NEW band use for this stroke?
            if target == "8U": exp_dist = 25
            elif target == "9-10": exp_dist = 25 if stroke == "fly" else 50
            else: exp_dist = 50
            new_key = f"{exp_dist}-{stroke}"
            m = stats.get("mean")
            s = stats.get("std", 0)
            if m is None: continue

            if dist == exp_dist:
                # Same distance: prefer the individual-age curve (gated), else the
                # band growth factor; then clamp so we never project a swimmer
                # faster than their best-ever (the Ada over-projection fix).
                growth = None
                if age_curve.USE_AGE_CURVE:
                    fa = proj_age_info.get(nm)
                    if fa:
                        growth = age_curve.same_dist_factor(
                            stroke, exp_dist, fa[0], fa[0] + (year - fa[1]))
                if growth is None:
                    growth = _growth_factor(cur_home, target, stroke)
                proj = m * growth
                if age_curve.USE_AGE_CURVE:
                    be = age_curve.best_ever(nm, stroke, exp_dist)
                    if be:
                        proj = max(proj, be * (1 - age_curve.CLAMP_FLOOR))
                new_strokes[new_key] = {
                    "mean": proj,
                    "std":  max(s * growth, proj * 0.02),
                    "_age_up_scaled": True,
                    "_growth_factor": growth,
                }
            elif dist == 25 and exp_dist == 50:
                # Distance changes (25→50): use combined factor for the FIRST
                # band step (which includes 1 year of growth), then compose
                # growth-only for any remaining steps.
                combined = DIST_GROWTH_25_TO_50.get(stroke, 2.02)
                # If new_idx - cur_idx > 1, we've skipped band(s); apply
                # additional growth-only steps.
                extra_growth = 1.0
                from Optimizer import AGE_GROUP_ORDER as _ord
                # combined already covers cur_home → (cur_home + 1).
                # If target > cur_home + 1, compose additional growth steps.
                intermediate = _ord[cur_idx + 1]
                if target != intermediate:
                    extra_growth = _growth_factor(intermediate, target, stroke)
                total = combined * extra_growth
                new_strokes[new_key] = {
                    "mean": m * total,
                    "std":  max(s * total, m * total * 0.02),
                    "_age_up_scaled": True,
                    "_scale_factor": total,
                }
            elif dist == 50 and exp_dist == 25:
                # Aging DOWN shouldn't happen (we only correct upward).
                pass

        out[name] = {
            "home_age_group": target,
            "gender": p.get("gender"),
            "strokes": new_strokes,
        }
    if n_corrected:
        print(f"[age_up] {team_name}: corrected {n_corrected} swimmer(s) "
              f"to current-season band", flush=True)
    return out


def _build_team_dated_from_history(team_name, year, max_week=None):
    """
    Build dated team_results for a team from nvsl_meet_history.json.
    Returns ({event: {swimmer: [(date, time), ...]}}, weeks_used) or (None, [])
    if the team isn't in history.
    """
    history = _load_history()
    if not history:
        return None, []
    year_data = history.get(str(year))
    if not year_data:
        return None, []
    out = defaultdict(lambda: defaultdict(list))
    weeks_used = []
    for wk_label in sorted(year_data.keys()):
        wk_num = int(wk_label.split()[1])
        if max_week is not None and wk_num >= max_week:
            continue
        meets = year_data[wk_label]
        had_team_this_week = False
        for mid, meet in meets.items():
            for side in ("team_a", "team_b"):
                team = meet[side]
                if team["name"].lower() != team_name.lower():
                    continue
                had_team_this_week = True
                for ev_label, ev_data in team["lineup"].items():
                    for sw in ev_data["swimmers"]:
                        if sw.get("time_sec") and sw["time_sec"] > 0:
                            out[ev_label][normalize_name(sw["name"])].append(
                                (meet["date"], sw["time_sec"])
                            )
        if had_team_this_week:
            weeks_used.append(wk_label)
    if not out:
        return None, []
    return {k: dict(v) for k, v in out.items()}, weeks_used


# Below this many distinct swimmers, a leaders_cache result is treated as "thin"
# (a stray pre-season time-trial entry, typical at W1) rather than real coverage.
LEADERS_MIN_SWIMMERS = 20

def _dated_swimmer_count(dated):
    """Distinct swimmers across all events in a dated_results dict."""
    if not dated:
        return 0
    sw = set()
    for _ev, d in dated.items():
        sw |= set(d.keys())
    return len(sw)

def _merge_dated(base, extra):
    """Merge `extra` {event:{swimmer:[(date,time),...]}} into `base` (in place)."""
    for ev, d in (extra or {}).items():
        bev = base.setdefault(ev, {})
        for sw, times in d.items():
            bev.setdefault(sw, []).extend(times)
    return base

def _history_baseline(team, year, max_week):
    """Best available history baseline: current-year (pre-max_week) if present,
    else prior-year full season. Returns (dated, weeks, source) or (None, [], None)."""
    d, w = _build_team_dated_from_history(team, year, max_week=max_week)
    if d is not None:
        return d, w, "history"
    d, w = _build_team_dated_from_history(team, year - 1, max_week=None)
    if d is not None:
        return d, w, f"history_{year-1}"
    return None, [], None


@app.route("/api/load_setup", methods=["POST"])
def api_load_setup():
    """
    Load profiles for both teams.

    Priority:
      1. leaders_cache.json  (best — full season w/ dates, filtered to before target meet)
      2. nvsl_meet_history.json  (fallback — only the weeks we captured)
      3. live scrape via Scraper.py  (slow last resort)
    """
    try:
        from Optimizer import build_profiles_recency_weighted

        body      = request.get_json()
        your_team = body["your_team"]
        opp_team  = body["opp_team"]
        year      = int(body.get("year") or 2026)
        week_num  = int(body.get("week") or 1)

        warnings = []
        your_source = None
        opp_source  = None

        # Compute date threshold: don't include times from the target meet itself
        max_date = _date_for_week(year, week_num)

        your_dated = opp_dated = None
        your_weeks = opp_weeks = []

        # === Source 1: leaders_cache.json (preferred) ===
        your_dated, your_weeks_dates = _build_team_dated_from_leaders(your_team, year, max_date=max_date)
        opp_dated,  opp_weeks_dates  = _build_team_dated_from_leaders(opp_team,  year, max_date=max_date)
        if your_dated is not None:
            your_source = "leaders"
            your_weeks  = your_weeks_dates
        if opp_dated is not None:
            opp_source = "leaders"
            opp_weeks  = opp_weeks_dates

        # === Source 1.5: thin-leaders guard (fixes W1 blowouts) ===
        # A stray pre-season time-trial in leaders_cache gives a 1-few-swimmer
        # profile that otherwise SUPPRESSES the full prior-year record → a near-
        # empty profile → ±200 score blowouts (13 of 24 W1 sides). When leaders
        # coverage is thin, merge the prior-year (full-season) baseline underneath
        # the stray current entries so the team has a real profile.
        if your_dated is not None and _dated_swimmer_count(your_dated) < LEADERS_MIN_SWIMMERS:
            n_thin = _dated_swimmer_count(your_dated)
            base, w, src = _history_baseline(your_team, year, week_num)
            if base is not None:
                your_dated = _merge_dated(base, your_dated)
                your_source = f"{src}+thin_leaders"; your_weeks = w
                warnings.append(f"{your_team}: thin leaders ({n_thin} swimmers) — merged {src} baseline")
        if opp_dated is not None and _dated_swimmer_count(opp_dated) < LEADERS_MIN_SWIMMERS:
            n_thin = _dated_swimmer_count(opp_dated)
            base, w, src = _history_baseline(opp_team, year, week_num)
            if base is not None:
                opp_dated = _merge_dated(base, opp_dated)
                opp_source = f"{src}+thin_leaders"; opp_weeks = w
                warnings.append(f"{opp_team}: thin leaders ({n_thin} swimmers) — merged {src} baseline")

        # === Source 2: meet history JSON (fallback) ===
        if your_dated is None:
            d, w = _build_team_dated_from_history(your_team, year, max_week=week_num)
            if d is not None:
                your_dated = d; your_weeks = w; your_source = "history"
        if opp_dated is None:
            d, w = _build_team_dated_from_history(opp_team, year, max_week=week_num)
            if d is not None:
                opp_dated = d; opp_weeks = w; opp_source = "history"

        # === Source 2.5: PRIOR-YEAR history fallback (W1 / no in-season data) ===
        # When there's no in-season data yet (typically W1, or a team with no
        # results uploaded for the current year), fall back to last year's
        # full-season history. This gives us a real baseline instead of forcing
        # live scrape (which would leak FUTURE-season times into a past prediction
        # — see fix below).
        if your_dated is None:
            d, w = _build_team_dated_from_history(your_team, year - 1, max_week=None)
            if d is not None:
                your_dated = d; your_weeks = w; your_source = f"history_{year-1}"
                warnings.append(
                    f"{your_team}: no {year} data yet — using {year-1} baseline")
        if opp_dated is None:
            d, w = _build_team_dated_from_history(opp_team, year - 1, max_week=None)
            if d is not None:
                opp_dated = d; opp_weeks = w; opp_source = f"history_{year-1}"
                warnings.append(
                    f"{opp_team}: no {year} data yet — using {year-1} baseline")

        # === Source 3: live scrape (last resort) ===
        # CRITICAL: skip live scrape when we're predicting a meet in the past
        # (max_date <= today). Live scrape returns CURRENT website state, which
        # includes all weeks of the current season — those would leak as
        # "future" data into a past prediction. _build_dated_results uses a
        # default_date='2099-01-01' that bypasses any downstream date filter,
        # so scraped data weights highest regardless of when it actually happened.
        import datetime as _dt
        today_iso = _dt.date.today().isoformat()
        scrape_safe = (max_date is None) or (max_date > today_iso)

        scrape_targets = []
        if scrape_safe:
            if your_dated is None: scrape_targets.append(your_team)
            if opp_dated  is None: scrape_targets.append(opp_team)
        else:
            if your_dated is None or opp_dated is None:
                warnings.append(
                    f"Past prediction (max_date={max_date} <= today={today_iso}); "
                    f"skipping live scrape to prevent future-data leak")
        if scrape_targets:
            print(f"[setup] live-scraping {scrape_targets} (no cache hit)...", flush=True)
            try:
                from Scraper import scrape_teams
                fresh = scrape_teams(scrape_targets, year=year)
                if your_dated is None and fresh.get(your_team):
                    your_dated = _build_dated_results(fresh[your_team])
                    your_source = "scrape"
                    warnings.append(f"{your_team}: not in leaders or history cache, used live scrape")
                if opp_dated is None and fresh.get(opp_team):
                    opp_dated = _build_dated_results(fresh[opp_team])
                    opp_source = "scrape"
                    warnings.append(f"{opp_team}: not in leaders or history cache, used live scrape")
            except Exception as e:
                warnings.append(f"Live scrape failed ({e})")

        if your_dated is None:
            return jsonify({"error": f"No data found for {your_team} anywhere"}), 400
        if opp_dated is None:
            return jsonify({"error": f"No data found for {opp_team} anywhere"}), 400

        # Coverage warnings
        if your_source == "history" and len(your_weeks) <= 1:
            warnings.append(f"{your_team}: only {your_weeks[0] if your_weeks else 'no'} weeks in history — consider running scrape_leaders.py")
        if opp_source == "history" and len(opp_weeks) <= 1:
            warnings.append(f"{opp_team}: only {opp_weeks[0] if opp_weeks else 'no'} weeks in history — consider running scrape_leaders.py")

        # Apply ladder data (with date filter — don't leak future ladder entries
        # into a prediction for an earlier meet)
        ladder_status = {"our": None, "opp": None}
        if (lu := _load_ladder_for_team(your_team)):
            added, skipped, future = _merge_ladder_into_dated(your_dated, lu, max_date=max_date)
            ladder_status["our"] = {"added": added, "skipped": skipped, "future_filtered": future, "total": len(lu)}
        if (lo := _load_ladder_for_team(opp_team)):
            added, skipped, future = _merge_ladder_into_dated(opp_dated, lo, max_date=max_date)
            ladder_status["opp"] = {"added": added, "skipped": skipped, "future_filtered": future, "total": len(lo)}

        # Manually-entered times (e.g. a DQ'd swim) count as real data — merge them in.
        if (mtu := _manual_times_as_ladder(your_team)):
            _merge_ladder_into_dated(your_dated, mtu, max_date=None)
        if (mto := _manual_times_as_ladder(opp_team)):
            _merge_ladder_into_dated(opp_dated, mto, max_date=None)

        # Revise the "no {year} data yet" warning: it's appended during source
        # selection (Source 2.5), BEFORE the ladder is merged above. If a ladder
        # then added current-season times, that warning is stale/misleading — the
        # ladder IS being used (2024 only backfills events it doesn't cover). Drop
        # the stale line and replace it with an accurate one.
        for side_key, team_nm in (("our", your_team), ("opp", opp_team)):
            st = ladder_status.get(side_key)
            if st and st.get("added"):
                stale = f"{team_nm}: no {year} data yet — using {year-1} baseline"
                warnings = [w for w in warnings if w != stale]
                warnings.append(
                    f"{team_nm}: using uploaded ladder ({st['added']} current-season "
                    f"times) + {year-1} baseline backfill")

        # Build profiles + cache for Run
        your_profiles = build_profiles_recency_weighted(your_dated, decay=0.7, cur_year=year)
        opp_profiles  = build_profiles_recency_weighted(opp_dated, decay=0.7, cur_year=year)

        # Age-up correction: when we use prior-year history (W1 with no current
        # data), home_age_group reflects the swimmer's OLD band. Override using
        # current-season signals (leaders_cache > ladder > SwimTopia roster).
        # This catches kids like Henry Okafor who were 8U in 2024 but are
        # 9-10 in 2025.
        your_profiles = _apply_age_up_correction(your_profiles, your_team, year)
        opp_profiles  = _apply_age_up_correction(opp_profiles,  opp_team,  year)

        events        = sorted(set(your_dated.keys()) | set(opp_dated.keys()))

        _cache["your_profiles"] = your_profiles
        _cache["opp_profiles"]  = opp_profiles
        _cache["events"]        = events
        _cache["ladder_status"] = ladder_status
        _cache["setup"] = {
            "your_team": your_team, "opp_team": opp_team,
            "year": year, "week_num": week_num,
            "your_source": your_source, "opp_source": opp_source,
            "your_weeks": your_weeks, "opp_weeks": opp_weeks,
            "warnings": warnings,
        }

        # Rosters for the availability UI (your team + opponent)
        your_roster = sorted(your_profiles.keys())
        opp_roster  = sorted(opp_profiles.keys())

        # Per-team time totals for transparency
        your_times = sum(len(ts) for ev in your_dated.values() for ts in ev.values())
        opp_times  = sum(len(ts) for ev in opp_dated.values()  for ts in ev.values())

        # Imputation review list: swimmers in prior years with no current-year data.
        # Also pull saved exclusions for both teams so the UI can pre-check them.
        try:
            your_review    = _swimmers_needing_review(your_team, year)
            opp_review     = _swimmers_needing_review(opp_team,  year)
            your_excludes  = sorted(_load_roster_excludes(your_team))
            opp_excludes   = sorted(_load_roster_excludes(opp_team))
        except Exception:
            your_review = opp_review = []
            your_excludes = opp_excludes = []

        # Ambiguous-name review: ladder names with a plausible-but-uncertain
        # match in last year's results (siblings vs. nicknames). Your team only —
        # the opponent's ladder isn't uploaded.
        try:
            your_name_flags, your_name_auto = _flag_name_matches(your_team, year)
        except Exception:
            your_name_flags, your_name_auto = [], 0

        return jsonify({
            "ok": True,
            "your_team":     your_team,
            "opp_team":      opp_team,
            "your_source":   your_source,
            "opp_source":    opp_source,
            "your_weeks":    your_weeks,
            "opp_weeks":     opp_weeks,
            "your_roster":   your_roster,
            "opp_roster":    opp_roster,
            "your_n":        len(your_profiles),
            "opp_n":         len(opp_profiles),
            "your_times":    your_times,
            "opp_times":     opp_times,
            "max_date":      max_date,
            "ladder_status": ladder_status,
            "warnings":      warnings,
            # NEW: imputation review (per team)
            "your_imp_review":   your_review,
            "opp_imp_review":    opp_review,
            "your_excludes_pre": your_excludes,   # already-saved excludes (pre-checked in UI)
            "opp_excludes_pre":  opp_excludes,
            # Per-(team, week) saved setup: absences (pre-unchecked) + imputation toggle
            "your_absent_pre":   _load_meet_prefs(your_team, year, week_num).get("absent", []),
            "use_imputation_pre": _load_meet_prefs(your_team, year, week_num).get("use_imputation"),
            # NEW: ambiguous-name review (your team)
            "your_name_flags":   your_name_flags,
            "your_name_auto":    your_name_auto,   # pairs auto-resolved by age
            # Manual time entry (Build page): the meet's events + already-saved times
            "events":            events,
            "your_manual_times": _load_manual_times(your_team),
        })
    except Exception:
        return jsonify({"error": traceback.format_exc()}), 500


@app.route("/api/team_roster", methods=["GET"])
def api_team_roster():
    """Return sorted list of unique swimmer names for a team, combining
    cached scrape data (results.json) and uploaded ladder data."""
    try:
        team = request.args.get("team", "").strip()
        if not team:
            return jsonify({"swimmers": []})

        names = set()

        # From cached scrape
        cache_file = os.path.join(BASE_DIR, "results.json")
        if os.path.exists(cache_file):
            try:
                with open(cache_file) as f:
                    cached = json.load(f)
                team_data = None
                for k, v in cached.items():
                    if k.lower() == team.lower():
                        team_data = v
                        break
                if team_data:
                    for ev, swimmers in team_data.items():
                        for name in swimmers.keys():
                            names.add(normalize_name(name))
            except Exception:
                pass

        # From ladder
        for entry in _load_ladder_for_team(team):
            names.add(normalize_name(entry["Name"]))

        return jsonify({"swimmers": sorted(names), "source": "scrape+ladder"})
    except Exception:
        return jsonify({"error": traceback.format_exc()}), 500


@app.route("/api/upload_swimtopia_roster", methods=["POST"])
def api_upload_swimtopia_roster():
    """
    Accept a SwimTopia '/manage/people' HTML paste and parse into roster/<slug>.json.
    Form fields: team (str), html (str), season (str, optional, default current year+1)
    """
    try:
        import tempfile, subprocess, datetime as _dt
        team_name = (request.form.get("team") or "").strip()
        html_text = request.form.get("html") or ""
        season    = (request.form.get("season") or str(_dt.date.today().year)).strip()
        if not team_name:
            return jsonify({"error": "team name required"}), 400
        if not html_text or len(html_text) < 1000:
            return jsonify({"error": "html paste looks empty/incomplete"}), 400

        # Save HTML to a tmp file, run the parser
        with tempfile.NamedTemporaryFile(mode="w", delete=False, suffix=".html") as tmp:
            tmp.write(html_text)
            tmp_path = tmp.name
        try:
            result = subprocess.run(
                ["python3", os.path.join(BASE_DIR, "parse_swimtopia_roster.py"),
                 tmp_path, team_name, season],
                capture_output=True, text=True, cwd=BASE_DIR, timeout=60,
            )
            if result.returncode != 0:
                return jsonify({"error": "parser failed",
                                "stdout": result.stdout, "stderr": result.stderr}), 500
        finally:
            try: os.unlink(tmp_path)
            except Exception: pass

        # Reload roster cache so the new file is visible
        import roster as _roster
        _roster.clear_cache()
        loaded = _roster.get_roster(team_name)
        n = len(loaded.get("swimmers", [])) if loaded else 0
        return jsonify({"ok": True, "team": team_name, "n_swimmers": n,
                        "stdout_tail": result.stdout[-800:]})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route("/api/upload_ladder", methods=["POST"])
def api_upload_ladder():
    """
    Accept a HY-TEK Team Manager 'ladder' PDF upload, parse it, and save the
    extracted time-trial / per-swimmer best-times data for the given team.
    Form fields: team (str), file (PDF)
    """
    try:
        import tempfile, time as _time
        if _parse_ladder_pdf is None:
            return jsonify({"error": "pymupdf not installed — run: pip3 install pymupdf"}), 500

        if "file" not in request.files:
            return jsonify({"error": "no file uploaded"}), 400
        team_name = (request.form.get("team") or "").strip()
        if not team_name:
            return jsonify({"error": "team name required"}), 400

        pdf_file = request.files["file"]
        if not pdf_file.filename.lower().endswith(".pdf"):
            return jsonify({"error": "file must be a PDF"}), 400

        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            pdf_file.save(tmp.name)
            tmp_path = tmp.name
        try:
            t0 = _time.time()
            entries = _parse_ladder_pdf(tmp_path)
            parse_ms = int((_time.time() - t0) * 1000)
            print(f"[upload_ladder] parsed {len(entries)} entries in {parse_ms}ms", flush=True)
        finally:
            try: os.unlink(tmp_path)
            except Exception: pass

        if not entries:
            return jsonify({"error": "no data parsed from PDF — wrong format?"}), 400

        # Apply alias rules + middle-initial normalization
        aliases = _load_aliases(team_name)
        entries = _apply_aliases(entries, aliases)

        # Save
        out = {
            "team": team_name,
            "entries": entries,
            "uploaded_at": __import__("datetime").datetime.now().isoformat(timespec="seconds"),
            "aliases_applied": len(aliases),
        }
        with open(_ladder_path(team_name), "w") as f:
            json.dump(out, f, indent=2)

        swimmers = sorted({e["Name"] for e in entries})
        events   = sorted({(e["AgeGroup"], e["Gender"], e["Stroke"], e["Distance"]) for e in entries})
        sources  = {}
        for e in entries:
            sources[e["Source"]] = sources.get(e["Source"], 0) + 1

        return jsonify({
            "ok":         True,
            "entries":    len(entries),
            "swimmers":   len(swimmers),
            "events":     len(events),
            "sources":    sources,
            "aliases_used": len(aliases),
            "saved_to":   os.path.basename(_ladder_path(team_name)),
        })
    except Exception:
        return jsonify({"error": traceback.format_exc()}), 500


@app.route("/api/ladder_info", methods=["GET"])
def api_ladder_info():
    """Report whether a saved ladder/time-trial PDF is already on file for a team.
    Ladders persist per-team (time_trials/<team>.json) and are reused automatically
    every session — the UI calls this so the user knows the ladder is still there
    and doesn't need to re-upload it each time."""
    team = (request.args.get("team") or "").strip()
    path = _ladder_path(team) if team else None
    if not team or not path or not os.path.exists(path):
        return jsonify({"exists": False})
    try:
        entries  = _load_ladder_for_team(team) or []
        swimmers = len({normalize_name(e.get("Name", "")) for e in entries if e.get("Name")})
        return jsonify({
            "exists":   True,
            "total":    len(entries),
            "swimmers": swimmers,
            "mtime":    os.path.getmtime(path),
        })
    except Exception:
        return jsonify({"exists": True, "total": 0, "swimmers": 0})


@app.route("/api/clear_cache", methods=["POST"])
def api_clear_cache():
    """Wipe in-memory cache and delete on-disk snapshot.

    Equivalent to restarting the app: forces a fresh build of profiles +
    fresh optimizer run on the next prediction. Useful after uploading
    new data (ladder, SwimTopia roster) or when results look stale.
    """
    n_keys = len(_cache)
    _cache.clear()
    snapshot_deleted = False
    try:
        if os.path.exists(_LINEUP_SNAPSHOT_PATH):
            os.remove(_LINEUP_SNAPSHOT_PATH)
            snapshot_deleted = True
    except Exception as e:
        return jsonify({"ok": False, "error": f"could not delete snapshot: {e}"}), 500
    return jsonify({
        "ok":                True,
        "cache_keys_cleared": n_keys,
        "snapshot_deleted":   snapshot_deleted,
    })


@app.route("/api/save_name_matches", methods=["POST"])
def api_save_name_matches():
    """Persist the user's 'same person' / 'different' answers for ambiguous ladder
    names. The frontend re-runs Load Data afterward so profiles rebuild with the
    merged identities (no re-scrape needed)."""
    try:
        body = request.get_json(force=True) or {}
        team = (body.get("team") or "").strip()
        decisions = body.get("decisions") or []
        if not team:
            return jsonify({"error": "team required"}), 400
        merged, distinct_total = _save_name_decisions(team, decisions)
        return jsonify({"ok": True, "merged": merged, "distinct_total": distinct_total})
    except Exception:
        return jsonify({"error": traceback.format_exc()}), 500


@app.route("/api/run", methods=["POST"])
def api_run():
    try:
        body      = request.get_json()
        your_team = body["your_team"]
        opp_team  = body["opp_team"]
        year      = int(body.get("year") or 2026)
        week_num  = int(body.get("week") or 1)
        use_opp_fingerprint = bool(body.get("use_opp_fingerprint", True))
        absent_list      = body.get("absent") or []
        opp_absent_list  = body.get("opp_absent") or []
        use_imputation   = bool(body.get("use_imputation", True))   # default ON
        swimup_only_if_scoring = bool(body.get("swimup_only_if_scoring", True))   # default ON
        your_excludes    = body.get("your_excludes") or []
        # Opp imputation panel was removed from the UI, so opp_excludes usually
        # isn't sent (None). When absent, we DON'T overwrite the saved file —
        # we load it instead, so previously-configured opp exclusions persist.
        opp_excludes     = body.get("opp_excludes")
        your_is_home     = bool(body.get("your_is_home", True))

        # Persist exclusion lists to disk so they survive across runs. The review
        # UI only lists returning swimmers with NO current-year data, so its
        # exclude payload omits anyone it doesn't show — e.g. a departed swimmer
        # who still has seed/leaders data (and so can never appear in the review).
        # Replacing the file wholesale would silently re-add those. So PRESERVE
        # excludes for swimmers OUTSIDE the current review list; the client's list
        # governs only the swimmers it actually showed (so review toggles still work).
        try:
            if your_excludes is not None:
                try:
                    shown = {c["norm_name"] for c in _swimmers_needing_review(your_team, year)}
                    existing = _load_roster_excludes(your_team)          # normalized set
                    client   = {normalize_name(n) for n in your_excludes}
                    your_excludes = sorted(client | {n for n in existing if n not in shown})
                except Exception:
                    pass  # fall back to the client list as-is
                _save_roster_excludes(your_team, your_excludes)
            if opp_excludes is not None:
                _save_roster_excludes(opp_team,  opp_excludes)
        except Exception:
            pass  # best-effort; don't fail run because save couldn't write

        # Fall back to the saved opp exclusions when the UI didn't send any.
        if opp_excludes is None:
            try:
                opp_excludes = sorted(_load_roster_excludes(opp_team))
            except Exception:
                opp_excludes = []

        weeks      = SEASON_WEEKS.get(year, [])
        week_info  = next((w for w in weeks if w["week"] == week_num), None)
        week_label = week_info["label"] if week_info else f"Week {week_num}"

        _cache["config"] = {
            "your_team": your_team, "opp_team": opp_team,
            "year": year, "week_num": week_num, "week_label": week_label,
            "use_opp_fingerprint": use_opp_fingerprint,
            "use_imputation":  use_imputation,
            "your_is_home":    your_is_home,
        }

        if week_num > 1:
            try:
                _fetch_opp_prev_lineup(opp_team, week_num, year)
            except Exception:
                pass  # best-effort; falls back gracefully if file is missing

        # Persist this meet's absences + imputation toggle ONLY when the client sent
        # them (the UI does). /api/save_prefs is the primary, immediate persister now,
        # so a run must never clobber a saved list — e.g. a bare API/test call with no
        # "absent" field would otherwise overwrite the coach's absences with an empty
        # one. Save just the fields that were actually provided.
        _pref_fields = {}
        if "absent" in body:
            _pref_fields["absent"] = absent_list
        if "use_imputation" in body:
            _pref_fields["use_imputation"] = use_imputation
        if _pref_fields:
            _save_meet_prefs(your_team, year, week_num, **_pref_fields)

        _run_and_cache(your_team, opp_team, year,
                       use_opp_fingerprint=use_opp_fingerprint,
                       absent_swimmers=absent_list,
                       opp_absent_swimmers=opp_absent_list,
                       use_imputation=use_imputation,
                       your_excludes=your_excludes,
                       opp_excludes=opp_excludes,
                       swimup_only_if_scoring=swimup_only_if_scoring)
        return jsonify({"ok": True})
    except Exception:
        return jsonify({"error": traceback.format_exc()}), 500


@app.route("/api/save_prefs", methods=["POST"])
def api_save_prefs():
    """Persist availability (absences + imputation toggle) and returning-swimmer
    excludes the MOMENT the coach changes them, so choices stick without having to
    run the optimizer. Each field is optional and merge-saved — a payload that omits
    a field leaves the stored value untouched (so a not-yet-rendered panel can't wipe
    it)."""
    try:
        body      = request.get_json(silent=True) or {}
        your_team = body.get("your_team")
        if not your_team:
            return jsonify({"error": "your_team required"}), 400
        year = int(body.get("year") or 2026)
        week = int(body.get("week") or 1)

        saved = []
        # Absences + imputation toggle (scoped per team + week)
        fields = {}
        if "absent" in body:
            fields["absent"] = body.get("absent") or []
        if "use_imputation" in body:
            fields["use_imputation"] = bool(body.get("use_imputation"))
        if fields:
            _save_meet_prefs(your_team, year, week, **fields)
            saved.append("availability")

        # Returning-swimmer excludes (per team). Merge with excludes for swimmers NOT
        # in the current review list, so we never silently re-add a departed swimmer
        # who has seed data (mirrors the merge in /api/run).
        if body.get("your_excludes") is not None:
            try:
                shown    = {c["norm_name"] for c in _swimmers_needing_review(your_team, year)}
                existing = _load_roster_excludes(your_team)
                client   = {normalize_name(n) for n in body["your_excludes"]}
                merged   = sorted(client | {n for n in existing if n not in shown})
            except Exception:
                merged = body["your_excludes"]
            _save_roster_excludes(your_team, merged)
            saved.append("returning")

        return jsonify({"ok": True, "saved": saved})
    except Exception:
        return jsonify({"error": traceback.format_exc()}), 500


@app.route("/api/manual_times", methods=["GET"])
def api_manual_times():
    """List a team's manually-entered times."""
    return jsonify({"times": _load_manual_times(request.args.get("team", ""))})


@app.route("/api/save_manual_time", methods=["POST"])
def api_save_manual_time():
    """Add (or replace) a manual time for one swimmer in one event. Merged into the
    profiles as a real time on the next load/run."""
    try:
        body  = request.get_json(silent=True) or {}
        team  = body.get("team")
        name  = (body.get("name") or "").strip()
        event = (body.get("event") or "").strip()
        tstr  = (body.get("time") or "").strip()
        if not (team and name and event and tstr):
            return jsonify({"error": "team, name, event, and time are all required"}), 400
        parts = event.split()
        if len(parts) < 3 or "-" not in parts[-1]:
            return jsonify({"error": f"unrecognized event '{event}'"}), 400
        sec = _parse_time_str(tstr)
        if sec is None or sec <= 0:
            return jsonify({"error": f"couldn't read the time '{tstr}' — use e.g. 33.5 or 1:05.2"}), 400
        times = [t for t in _load_manual_times(team)
                 if not (normalize_name(t.get("name", "")) == normalize_name(name)
                         and t.get("event") == event)]
        times.append({"name": name, "event": event, "time": tstr,
                      "time_sec": round(sec, 2)})
        times.sort(key=lambda t: (t.get("event", ""), t.get("name", "")))
        _save_manual_times(team, times)
        return jsonify({"ok": True, "times": times})
    except Exception:
        return jsonify({"error": traceback.format_exc()}), 500


@app.route("/api/delete_manual_time", methods=["POST"])
def api_delete_manual_time():
    """Remove a manual time for one swimmer + event."""
    try:
        body  = request.get_json(silent=True) or {}
        team  = body.get("team")
        name  = body.get("name") or ""
        event = body.get("event") or ""
        times = [t for t in _load_manual_times(team)
                 if not (normalize_name(t.get("name", "")) == normalize_name(name)
                         and t.get("event") == event)]
        _save_manual_times(team, times)
        return jsonify({"ok": True, "times": times})
    except Exception:
        return jsonify({"error": traceback.format_exc()}), 500


# ── Lineup ─────────────────────────────────────────────────────────────────────

@app.route("/lineup")
def lineup():
    data = _cache.get("lineup_data")
    if not data:
        # In-memory cache is empty (e.g. after a server restart) — try the
        # on-disk snapshot of the last optimize before giving up.
        if _load_lineup_snapshot():
            data = _cache.get("lineup_data")
    if not data:
        # Nothing server-side and no snapshot — show the empty CTA.
        return render_template("lineup.html", active_page="lineup", payload=None)

    cfg          = _cache.get("config", {})
    mstats       = data.get("match_stats") or {}
    analytical   = data.get("analytical_winp")
    robust_norm  = data.get("robust_norm_winp")
    robust_ewp   = data.get("robust_ewp")
    week_num     = data.get("week_num")

    if week_num == 1:
        # W1 uses its own calibrated win-prob (_w1_winprob — probit with intercept,
        # fit on W1 margins). It's correctly de-biased but only weakly informative
        # (W1 is inherently noisy with no current-season data), so we show it WITH a
        # low-confidence caveat rather than presenting it like a W2-5 number.
        if analytical is not None:
            headline_wp    = analytical * 100
            headline_label = ("chance of winning — rough guess this early "
                              "in the season, before anyone has raced")
        else:
            headline_wp    = None
            headline_label = "win probability unavailable at Week 1"
    elif analytical is not None:
        # Issue #2: shrinkage-corrected headline win prob (de-biases the optimizer's
        # inflated predicted margin before mapping to a probability).
        headline_wp    = analytical * 100
        headline_label = "chance of winning"
    elif robust_norm is not None:
        headline_wp    = robust_norm * 100
        headline_label = "chance of winning"
    elif robust_ewp is not None:
        headline_wp    = robust_ewp * 100
        headline_label = "average chance of winning across the other team's likely lineups"
    elif mstats.get("win_prob") is not None:
        headline_wp    = mstats["win_prob"] * 100
        headline_label = "chance of winning if they swim their best lineup"
    else:
        headline_wp    = None
        headline_label = None

    payload = {
        "your_team":         cfg.get("your_team", ""),
        "opp_team":          cfg.get("opp_team", ""),
        "week_label":        cfg.get("week_label", ""),
        "your_is_home":      data.get("your_is_home", True),
        "headline_wp":       headline_wp,
        "headline_wp_lastweek": data.get("headline_wp_lastweek"),
        "headline_label":    headline_label,
        "events":            data.get("events", []),
        "has_lastweek_opp":  data.get("has_lastweek_opp", False),
        "opp_median_lastweek": data.get("opp_median_lastweek"),
        "opt_score":         data.get("opt_score", 0),
        "mc_total":          data.get("mc_total", 0),
        "our_median":        (data.get("match_stats") or {}).get("our_median"),
        "opp_median":        (data.get("match_stats") or {}).get("opp_median"),
        "margin":            (data.get("match_stats") or {}).get("margin"),
        "relay_exp_pts":     data.get("relay_exp_pts"),
        "relay_data":        _cache.get("relay_data", {}),
        "scenarios":         data.get("scenarios", []),
        "pct_needed":        (data.get("score_stats") or {}).get("pct_needed"),
        "score_target":      (data.get("score_stats") or {}).get("target"),
        "robust_breakdown":  data.get("robust_breakdown"),
    }

    # Prepare the Results-page "pick a swimmer" lanes. Every OUR lane with no real
    # time (a league-average fill or a filler body) stays SHOWN but becomes an
    # editable dropdown limited to under-2-event swimmers; truly empty lanes get the
    # dropdown too. Nothing is blanked — the coach always sees who's slotted.
    from Optimizer import MAX_EVENTS as _MAX_EV
    _elig = _cache.get("eligible_by_event") or _eligible_by_event(
        _cache.get("your_profiles"), _cache.get("events"))

    for _ev in payload["events"]:
        for _ln in (_ev.get("lanes_predicted") or []):
            if _ln.get("team") != "us":
                continue
            if _ln.get("no_real_time") or not _ln.get("swimmer"):
                _ln["pick_one"] = True  # no-time OR empty OUR lane → coach-pickable
                _ln["no_real_time"] = True

    # Count each OUR swimmer's current individual-event placements (real + picked).
    _usage = {}
    for _ev in payload["events"]:
        for _ln in (_ev.get("lanes_predicted") or []):
            if _ln.get("team") == "us" and _ln.get("swimmer"):
                _usage[_ln["swimmer"]] = _usage.get(_ln["swimmer"], 0) + 1

    for _ev in payload["events"]:
        _lanes = _ev.get("lanes_predicted") or []
        _pick = [ln for ln in _lanes if ln.get("pick_one")]
        if not _pick:
            continue
        _in_event = {ln.get("swimmer") for ln in _lanes
                     if ln.get("team") == "us" and ln.get("swimmer")}
        _all = _elig.get(_ev["event"]) or []
        for _ln in _pick:
            _own = _ln.get("swimmer")
            _cands = [_own] if _own else []         # always keep the slotted swimmer selectable
            for n in _all:
                if n == _own:
                    continue
                if _usage.get(n, 0) >= _MAX_EV:      # already in the max events
                    continue
                if n in _in_event:                  # already swimming this event
                    continue
                _cands.append(n)
            _ln["pick_candidates"] = _cands

    return render_template("lineup.html", active_page="lineup", payload=payload)


@app.route("/api/set_lane_pick", methods=["POST"])
def api_set_lane_pick():
    """Coach picks the swimmer for a 'no real time — pick one' lane directly on the
    Results page. Persists the choice on the cached lineup + snapshot so both the
    Results page and the Excel export show it. It's a meet-day roster choice (the
    swimmer has no seed time), so it isn't scored — like the other no-time fillers."""
    body    = request.get_json(silent=True) or {}
    event   = body.get("event")
    lane    = body.get("lane")
    swimmer = (body.get("swimmer") or "").strip() or None

    data = _cache.get("lineup_data")
    if not data:
        _load_lineup_snapshot()
        data = _cache.get("lineup_data")
    if not data:
        return jsonify({"error": "no active lineup"}), 409

    # Don't let a manual pick push a swimmer over the 2-event cap (guards against a
    # stale dropdown). Count their placements in OTHER events / lanes.
    if swimmer:
        from Optimizer import MAX_EVENTS as _MX
        cnt = 0
        for ev in data.get("events", []):
            for ln in (ev.get("lanes_predicted") or []):
                if ln.get("team") == "us" and ln.get("swimmer") == swimmer \
                        and not (ev.get("event") == event and ln.get("lane") == lane):
                    cnt += 1
        if cnt >= _MX:
            return jsonify({"error": f"{swimmer} is already in {cnt} events (max {_MX})."}), 400

    updated = False
    for ev in data.get("events", []):
        if ev.get("event") != event:
            continue
        for grid_key in ("lanes_predicted", "lanes_lastweek"):
            for ln in (ev.get(grid_key) or []):
                if ln.get("lane") == lane and ln.get("team") == "us" \
                        and (ln.get("pick_one") or ln.get("no_real_time")):
                    ln["swimmer"]      = swimmer
                    ln["manual_pick"]  = bool(swimmer)
                    ln["pick_one"]     = True
                    ln["no_real_time"] = True
                    updated = True
        break
    if not updated:
        return jsonify({"error": "pick-one lane not found"}), 404

    # Persist to the snapshot, preserving its precomputed maps.
    try:
        snap = {}
        if os.path.exists(_LINEUP_SNAPSHOT_PATH):
            with open(_LINEUP_SNAPSHOT_PATH) as f:
                snap = json.load(f) or {}
        snap["lineup_data"] = data
        with open(_LINEUP_SNAPSHOT_PATH, "w") as f:
            json.dump(snap, f)
    except Exception as e:
        print(f"[lane_pick] persist failed: {e}", flush=True)

    return jsonify({"ok": True, "event": event, "lane": lane, "swimmer": swimmer})


@app.route("/check")
def check():
    """What-if lineup lab: edit the optimal lineup and compare. Most of the
    interactivity is client-side; live re-scoring + swap suggestions need a
    backend endpoint (stubbed with hints in the UI for now)."""
    data = _cache.get("lineup_data")
    if not data:
        if _load_lineup_snapshot():
            data = _cache.get("lineup_data")
    if not data:
        return render_template("check.html", active_page="check", payload=None)

    cfg    = _cache.get("config", {})
    # Baseline win% on the SAME basis the score endpoint uses: a head-to-head
    # sim of our lineup vs the opponent's PREDICTED lineup (already computed as
    # a scenario). Keeps the optimal-vs-edited delta apples-to-apples.
    scen   = {s.get("label"): s for s in (data.get("scenarios") or [])}
    pred   = scen.get("Their predicted lineup") or {}
    mstats = data.get("match_stats") or {}
    headline_wp = (pred["win_prob"] * 100) if pred.get("win_prob") is not None else (
        mstats["win_prob"] * 100 if mstats.get("win_prob") is not None else None)

    # Derive, per event, OUR entered swimmers (from the lane layout), plus a
    # per-gender pool of our swimmers so the editor's dropdowns stay same-gender.
    # (Pure data reshaping — no optimizer logic.)
    events = data.get("events", [])
    # Per-event eligible swimmers for the editor dropdowns. Compute fresh from
    # the in-memory profiles when available; otherwise use the copy saved in the
    # snapshot (survives restarts). Empty → template falls back to gender pool.
    your_profiles = _cache.get("your_profiles")
    if your_profiles:
        elig_map = _eligible_by_event(your_profiles, _cache.get("events")
                                      or [ev.get("event") for ev in events])
    else:
        elig_map = _cache.get("eligible_by_event") or {}

    check_events = []
    boys_pool, girls_pool = set(), set()
    _yp = _cache.get("your_profiles") or {}
    _yp_norm = {normalize_name(k): v for k, v in _yp.items()}
    for ev in events:
        label = ev.get("event", "")
        gender = label.split()[1] if len(label.split()) > 1 else ""
        ours = [ln["swimmer"] for ln in (ev.get("lanes_predicted") or [])
                if ln.get("team") == "us" and ln.get("swimmer")]
        for s in ours:
            (boys_pool if gender == "Boys" else girls_pool).add(s)
        # no_time = fielded swimmers with NO real current-season time in this stroke
        # (imputed/league-avg/prior-year, or absent) — surfaced so the coach can see
        # which entries aren't backed by a real seed time.
        try:
            _stroke = parse_event(label)[2]
        except Exception:
            _stroke = None
        if _yp:
            # Raw profiles in memory (fresh optimize): authoritative — also flags
            # league-avg / prior-year imputed times, not just absent ones.
            no_time = []
            if _stroke:
                for s in ours:
                    prof = _yp.get(s) or _yp_norm.get(normalize_name(s)) or {}
                    sd = (prof.get("strokes") or {}).get(_stroke) or {}
                    if sd.get("mean") is None or sd.get("source"):
                        no_time.append(s)
        elif _cache.get("no_time_by_event"):
            # Rehydrated from a snapshot that carried the precomputed map.
            no_time = _cache["no_time_by_event"].get(label, [])
        else:
            # Older snapshot without the map: fall back to the per-lane flag the
            # optimizer already persisted, so we surface true fillers instead of
            # (with no profiles to read) flagging EVERY swimmer.
            no_time = [ln["swimmer"] for ln in (ev.get("lanes_predicted") or [])
                       if ln.get("team") == "us" and ln.get("swimmer")
                       and (ln.get("no_real_time") or ln.get("time") is None)]
        # Always show 3 slots per event (NVSL lanes per team); pad the unfilled ones
        # with empty selects so the coach can drop a swimmer into a blank lane.
        padded = list(ours)[:3] + [""] * max(0, 3 - len(ours))
        check_events.append({
            "event":    label,
            "swimmers": padded,
            "eligible": elig_map.get(label, []),
            "no_time":  no_time,
            "win_pct":  ev.get("win_pct"),
            "mc_pts":   ev.get("mc_pts"),
        })

    payload = {
        "your_team":     data.get("your_team", cfg.get("your_team", "")),
        "opp_team":      data.get("opp_team", cfg.get("opp_team", "")),
        "check_events":  check_events,
        "boys_pool":     sorted(boys_pool),
        "girls_pool":    sorted(girls_pool),
        "mc_total":      data.get("mc_total", 0),
        "relay_exp_pts": data.get("relay_exp_pts"),
        "headline_wp":   headline_wp,
        "elig_ready":    bool(elig_map),   # False → dropdowns fall back to gender pool
    }
    return render_template("check.html", active_page="check", payload=payload)


# ── Anomaly reporting ─────────────────────────────────────────────────────────
# Records noteworthy events (a coach lineup beating "optimal"; an .xlsx we couldn't
# parse) so we can spot bugs / format gaps. PRIVACY: payloads are anonymized by the
# CALLER — never pass swimmer names (often minors) or uploaded file contents. We log
# only scores/deltas, event counts, sheet names, and column-header labels. Always
# local (anomalies.jsonl); also POSTs to LANELAB_ANOMALY_WEBHOOK if that env is set
# (fire-and-forget, so it never blocks or breaks a request).
ANOMALY_BEAT_OPTIMAL_PTS = 5.0   # min margin over optimal to flag (filters MC noise)

def _log_anomaly(kind, payload=None):
    import datetime as _dt
    rec = {"kind": kind, "ts": _dt.datetime.now().isoformat(timespec="seconds"),
           **(payload or {})}
    try:
        with open(os.path.join(BASE_DIR, "anomalies.jsonl"), "a") as f:
            f.write(json.dumps(rec) + "\n")
    except Exception:
        pass
    url = os.environ.get("LANELAB_ANOMALY_WEBHOOK")
    if url:
        def _post():
            try:
                import urllib.request
                req = urllib.request.Request(
                    url, data=json.dumps(rec).encode("utf-8"),
                    headers={"Content-Type": "application/json"})
                urllib.request.urlopen(req, timeout=4)
            except Exception:
                pass
        try:
            import threading
            threading.Thread(target=_post, daemon=True).start()
        except Exception:
            pass


# ── Fuzzy lineup-xlsx matching (tolerate foreign coach spreadsheets) ───────────
_STROKE_SYN = {
    "free": "free", "freestyle": "free", "fr": "free",
    "back": "back", "backstroke": "back", "bk": "back",
    "breast": "breast", "breaststroke": "breast", "br": "breast",
    "fly": "fly", "butterfly": "fly", "fl": "fly",
    "im": "im", "medley": "im",
}

def _canon_event_key(s):
    """Loose (age, gender, dist, stroke) key so coach strings like
    '8 & Under Boys 25 Freestyle' match the app's '8U Boys 25-free'."""
    import re
    # keep hyphens (word boundaries still split '50-free'); only expand '&'
    t = " " + str(s).lower().replace("&", " and ") + " "
    if re.search(r"\b8\s*(?:and\s*)?(?:u|under)\b|\b8u\b", t):
        age = "8u"
    else:
        m = re.search(r"\b(9|11|13|15)\s*(?:to|through|[-–]|\s)\s*(10|12|14|18)\b", t)
        age = f"{m.group(1)}-{m.group(2)}" if m else ""
    if re.search(r"\bgirl|\bwomen\b|\bwoman\b|\bfemale\b", t):
        gender = "Girls"
    elif re.search(r"\bboy|\bmen\b|\bman\b|\bmale\b", t):
        gender = "Boys"
    else:
        gender = ""
    dm = re.search(r"\b(25|50|100|200)\b", t)
    dist = dm.group(1) if dm else ""
    stroke = ""
    for k in sorted(_STROKE_SYN, key=len, reverse=True):
        if re.search(r"\b" + re.escape(k) + r"\b", t):
            stroke = _STROKE_SYN[k]; break
    return (age, gender, dist, stroke)

def _match_name_to_roster(raw, by_norm):
    """Coach name -> roster display name (or None). Handles 'Last, First'."""
    nn = normalize_name(raw)
    if nn in by_norm:
        return by_norm[nn]
    if "," in raw:
        a, b = (raw.split(",", 1) + [""])[:2]
        nn2 = normalize_name(f"{b.strip()} {a.strip()}")
        if nn2 in by_norm:
            return by_norm[nn2]
    return None


@app.route("/api/export_lineup.xlsx")
def api_export_lineup_xlsx():
    """Download the current cached lineup (individual events + relays) as .xlsx."""
    data = _cache.get("lineup_data")
    if not data:
        return jsonify({"error": "No lineup in memory — run the optimizer first."}), 409
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO
        cfg       = _cache.get("config", {})
        your_team = cfg.get("your_team", "Your team")
        opp_team  = cfg.get("opp_team", "Opponent")
        week_lbl  = cfg.get("week_label", "")
        bold = Font(bold=True)
        hfill = PatternFill("solid", fgColor="DDE6F0")

        # Section-per-event layout: a bold event row, then its swimmers below,
        # fastest first (seed-time order, not lane order). The Event/Swimmer
        # header row is kept so the Check page's xlsx parser (which carries the
        # last seen event forward over blank cells) round-trips this exactly.
        wb = Workbook()
        ws = wb.active
        ws.title = "Lineup"
        ws["A1"] = "Lane Lab — Lineup"; ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"{your_team} vs {opp_team}"; ws["A2"].font = bold
        ws["A3"] = week_lbl
        # Two side-by-side blocks: BOYS in cols A-C, GIRLS in cols E-G (D is a
        # spacer). Each block: Event | Swimmer | Time, swimmers fastest-first with
        # their seed time shown (↑ marks a swim-up; "no time" marks a fill lane).
        r0 = 5  # column-header row; gender labels sit on r0-1, swimmers from r0+1

        def _fmt_time(t):
            if t is None:
                return ""
            t = float(t)
            if t >= 60:
                m = int(t // 60)
                return f"{m}:{t - 60 * m:05.2f}"
            return f"{t:.2f}"

        def _write_block(events_list, c0, label):
            ws.cell(row=r0 - 1, column=c0, value=label).font = Font(bold=True, size=12)
            for c, h in enumerate(["Event", "Swimmer", "Time"], c0):
                cell = ws.cell(row=r0, column=c, value=h); cell.font = bold; cell.fill = hfill
            r = r0 + 1
            for ev in events_list:
                cell = ws.cell(row=r, column=c0, value=ev.get("event"))
                cell.font = bold; cell.fill = hfill
                ws.cell(row=r, column=c0 + 1).fill = hfill
                ws.cell(row=r, column=c0 + 2).fill = hfill
                r += 1
                our = [l for l in (ev.get("lanes_predicted") or [])
                       if l.get("team") == "us" and l.get("swimmer")]
                if not our:
                    ws.cell(row=r, column=c0 + 1, value="(no entry)"); r += 1; continue
                our.sort(key=lambda l: (l.get("time") is None, l.get("time")))
                for lane in our:
                    ws.cell(row=r, column=c0 + 1, value=lane.get("swimmer"))
                    tstr = "no time" if lane.get("no_real_time") else _fmt_time(lane.get("time"))
                    if lane.get("is_swim_up") and tstr:
                        tstr += " ↑"
                    ws.cell(row=r, column=c0 + 2, value=tstr)
                    r += 1

        all_events = data.get("events") or []
        from Optimizer import AGE_GROUP_ORDER as _AGO
        _STROKE_ORDER = {"free": 0, "back": 1, "breast": 2, "fly": 3}
        def _ev_parts(ev):
            p = (ev.get("event") or "").split()
            age    = p[0] if p else ""
            gender = p[1] if len(p) > 1 else ""
            raw    = p[2] if len(p) > 2 else ""
            stroke = raw.split("-")[-1] if "-" in raw else raw
            return age, gender, stroke
        def _ev_sort_key(ev):
            age, _, stroke = _ev_parts(ev)
            return (_AGO.index(age) if age in _AGO else 99,
                    _STROKE_ORDER.get(stroke, 9))   # free → back → breast → fly
        def _gender(ev):
            return _ev_parts(ev)[1]
        boys  = sorted([e for e in all_events if _gender(e) == "Boys"],  key=_ev_sort_key)
        girls = sorted([e for e in all_events if _gender(e) == "Girls"], key=_ev_sort_key)
        _write_block(boys, 1, "BOYS")
        _write_block(girls, 5, "GIRLS")
        for col, w in zip(["A", "B", "C", "D", "E", "F", "G"],
                          [22, 26, 12, 3, 22, 26, 12]):
            ws.column_dimensions[col].width = w

        relay_data = _cache.get("relay_data")
        if isinstance(relay_data, dict) and relay_data:
            try:
                w2 = wb.create_sheet("Relays")
                for c, h in enumerate(["Relay", "Swimmers", "Exp Pts", "Win %"], 1):
                    cell = w2.cell(row=1, column=c, value=h); cell.font = bold; cell.fill = hfill
                rr = 2
                for gender, gdata in relay_data.items():
                    for relay in (gdata.get("age_relays") or []):
                        legs = ", ".join((l.get("swimmer") or l.get("name") or "")
                                         for l in (relay.get("legs") or []))
                        w2.cell(row=rr, column=1, value=f"{gender} {relay.get('name', '')}")
                        w2.cell(row=rr, column=2, value=legs)
                        w2.cell(row=rr, column=3, value=round(relay.get("exp_pts", 0) or 0, 1))
                        w2.cell(row=rr, column=4, value=round(relay.get("win_pct", 0) or 0))
                        rr += 1
                for col, w in zip("ABCD", [22, 52, 10, 10]):
                    w2.column_dimensions[col].width = w
            except Exception:
                pass

        buf = BytesIO(); wb.save(buf); buf.seek(0)
        fname = f"lane_lab_lineup_{_safe_team_path(your_team)}.xlsx"
        return Response(
            buf.read(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'})
    except Exception:
        return jsonify({"error": traceback.format_exc()}), 500


@app.route("/api/parse_lineup_xlsx", methods=["POST"])
def api_parse_lineup_xlsx():
    """Parse an uploaded .xlsx coach lineup into the app's CANONICAL {event: [swimmers]}.
    Tolerant of foreign formats: scans every sheet for an 'Event' column + a
    Swimmer/Name/Athlete column, fuzzy-matches event strings ('8 & Under Boys 25 Free'
    -> '8U Boys 25-free') and names (incl. 'Last, First') against the loaded meet's
    events + roster. Returns the canonical lineup plus warnings for anything unmatched.
    Round-trips the Export format exactly; degrades gracefully on anything else."""
    try:
        from openpyxl import load_workbook
        f = request.files.get("file")
        if not f:
            return jsonify({"error": "No file uploaded"}), 400
        # Editor's own event labels (so matched events line up with the dropdowns).
        try:
            posted_events = json.loads(request.form.get("events") or "[]")
        except Exception:
            posted_events = []
        canon_events = posted_events or list(_cache.get("events") or
                                             (_cache.get("opp_predicted_full") or {}).keys())

        wb = load_workbook(f, read_only=True, data_only=True)
        name_hdrs = ("swimmer", "name", "athlete", "swimmers")
        raw, found_header = {}, False
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            blocks, hdr = [], None      # blocks: list of (ev_col, nm_col)
            for i, row in enumerate(rows[:25]):
                cells = [(str(c).strip().lower() if c is not None else "") for c in row]
                ev_cols = [j for j, c in enumerate(cells) if c == "event"]
                nm_cols = [j for j, c in enumerate(cells) if c in name_hdrs]
                if ev_cols and nm_cols:
                    # Pair each Event column with the nearest Name column to its right
                    # (or nearest overall) — supports the two-gender side-by-side export.
                    for ec in ev_cols:
                        nc = min((n for n in nm_cols if n > ec), default=None)
                        if nc is None:
                            nc = min(nm_cols, key=lambda n: abs(n - ec))
                        blocks.append((ec, nc))
                    hdr = i; break
            if hdr is None:
                continue
            found_header = True
            for ev_col, nm_col in blocks:
                last_ev = None
                for row in rows[hdr + 1:]:
                    ev = row[ev_col] if ev_col < len(row) else None
                    nm = row[nm_col] if nm_col < len(row) else None
                    ev = str(ev).strip() if ev not in (None, "") else last_ev
                    if ev:
                        last_ev = ev
                    nm = str(nm).strip() if nm is not None else ""
                    if ev and nm and nm != "(no entry)":
                        raw.setdefault(ev, []).append(nm)

        if not found_header:
            _log_anomaly("xlsx_parse_fail", {"reason": "no Event + Swimmer/Name header",
                         "sheets": wb.sheetnames})
            return jsonify({"error": "Couldn't find an 'Event' column and a "
                                     "'Swimmer'/'Name' column. Tip: a file exported from "
                                     "Lane Lab always works as a template."}), 400

        # Fuzzy-match raw events -> canonical labels, raw names -> roster display names.
        ev_by_key = {}
        for ce in canon_events:
            ev_by_key.setdefault(_canon_event_key(ce), ce)
        prof = _cache.get("your_profiles") or {}
        by_norm = {normalize_name(k): k for k in prof}

        lineup, unmatched_events, unmatched_names = {}, [], []
        for rev, names in raw.items():
            ce = ev_by_key.get(_canon_event_key(rev))
            if not ce:
                unmatched_events.append(rev); continue
            for nm in names:
                disp = _match_name_to_roster(nm, by_norm) if by_norm else nm
                if disp:
                    lineup.setdefault(ce, []).append(disp)
                else:
                    unmatched_names.append(f"{nm} ({rev})")

        if not lineup:
            _log_anomaly("xlsx_parse_fail", {
                "reason": "parsed but nothing matched the loaded meet",
                "sheets": wb.sheetnames,
                "raw_events": len(raw),
                "unmatched_events": len(unmatched_events),
                "unmatched_names": len(unmatched_names)})
            return jsonify({"error": "Parsed the file but couldn't match any events to "
                                     "this meet. Check the events/teams match what you ran.",
                            "unmatched_events": unmatched_events[:20]}), 400

        return jsonify({"ok": True, "lineup": lineup, "n_events": len(lineup),
                        "unmatched_events": unmatched_events[:20],
                        "unmatched_names": unmatched_names[:20]})
    except Exception:
        _log_anomaly("xlsx_parse_fail", {"reason": "exception"})
        return jsonify({"error": traceback.format_exc()}), 500


@app.route("/api/score_lineup", methods=["POST"])
def api_score_lineup():
    """Score a coach-edited lineup against the opponent's predicted lineup,
    reusing the SAME sim/scoring used for the optimal lineup. No model or
    calibration changes — it just feeds an arbitrary lineup through the
    existing functions. Returns total/win%/medians + per-event numbers, and
    flags swimmers dropped for having no time in an event's stroke."""
    your_profiles = _cache.get("your_profiles")
    opp_profiles  = _cache.get("opp_profiles")
    events        = _cache.get("events")
    relay_data    = _cache.get("relay_data")
    opp_lineup    = _cache.get("opp_predicted_full")
    if not (your_profiles and opp_profiles and events and opp_lineup):
        return jsonify({"error": "No active lineup in memory — re-run the optimizer "
                                 "(swimmer profiles aren't kept after a server restart)."}), 409

    body   = request.get_json(silent=True) or {}
    posted = body.get("lineup") or {}

    # Build a scorable lineup: keep only swimmers who have a profile + a time in
    # the event's stroke. Dropped names come back as warnings (eligibility feedback).
    edited, warnings = {}, []
    for ev in posted:
        if ev not in opp_lineup:
            continue
        _, _, stroke = parse_event(ev)
        valid, dropped = [], []
        for s in (posted.get(ev) or []):
            if s and s in your_profiles and stroke in your_profiles[s].get("strokes", {}):
                valid.append(s)
            elif s:
                dropped.append(s)
        if dropped:
            warnings.append({"event": ev, "swimmers": dropped})
        edited[ev] = {"swimmers": valid, "expected_points": 0.0}

    N = 4000
    rows, mc_total = _build_lineup_rows(edited, opp_lineup, your_profiles, opp_profiles)
    mstats = match_stats(simulate_match(edited, opp_lineup, your_profiles, opp_profiles,
                                        events, relay_data, n=N))
    opt       = _cache.get("lineup_data") or {}
    relay_exp = opt.get("relay_exp_pts") or 0.0

    # CRN-PRECISE per-event delta vs the OPTIMAL lineup. Score the optimal and the
    # edited lineup against the SAME pre-drawn random numbers so the noise cancels
    # — a one-event coach edit then reads a trustworthy ±0.x instead of the ±2 of
    # comparing two independent MC runs (which sent us chasing phantom +0.2 swaps).
    import numpy as _np
    orig_lineup = {r["event"]: {"swimmers": r.get("swimmers", [])} for r in opt.get("rows", [])}
    deltas = {}
    try:
        our_names = sorted(your_profiles.keys()); opp_names = sorted(opp_profiles.keys())
        _rng = _np.random.default_rng(seed=42)
        crn = {"our_bank": _rng.standard_normal((len(our_names), N)),
               "opp_bank": _rng.standard_normal((len(opp_names), N)),
               "our_idx":  {n: i for i, n in enumerate(our_names)},
               "opp_idx":  {n: i for i, n in enumerate(opp_names)}}
        def _ev_pts(lineup, ev, age, stroke):
            oc = tuple(lineup.get(ev, {}).get("swimmers", []))
            pc = tuple(opp_lineup.get(ev, {}).get("swimmers", []))
            p, _ = _sim_event_both_pts(oc, pc, stroke, your_profiles, opp_profiles,
                                       N, age_group=age, crn=crn)
            return float(p.mean())
        for ev in events:
            try:
                age, _, stroke = parse_event(ev)
            except Exception:
                continue
            deltas[ev] = round(_ev_pts(edited, ev, age, stroke)
                               - _ev_pts(orig_lineup, ev, age, stroke), 2)
    except Exception:
        deltas = {}
    # Sanity signal: a coach-edited lineup should NOT out-score our "optimal" by much.
    # If it does (beyond MC noise), the optimizer likely left points on the table — flag
    # it (anonymized: scores/teams/week only, never swimmer names).
    try:
        opt_mc = opt.get("mc_total")
        if opt_mc is not None:
            delta = float(mc_total) - float(opt_mc)
            if delta > ANOMALY_BEAT_OPTIMAL_PTS:
                cfg = _cache.get("config", {})
                _log_anomaly("check_beats_optimal", {
                    "delta": round(delta, 1),
                    "check_mc": round(float(mc_total), 1),
                    "optimal_mc": round(float(opt_mc), 1),
                    "your_team": cfg.get("your_team"),
                    "opp_team": cfg.get("opp_team"),
                    "week": cfg.get("week_num") or cfg.get("week_label"),
                    "n_dropped": len(warnings)})
    except Exception:
        pass
    return jsonify({
        "ok":         True,
        "win_pct":    mstats["win_prob"] * 100,
        "our_median": mstats["our_median"],
        "opp_median": mstats["opp_median"],
        "mc_total":   mc_total,
        "total":      mc_total + relay_exp,
        "events":     [{"event": r["event"], "win_pct": r["win_pct"], "mc_pts": r["mc_pts"],
                        "delta": deltas.get(r["event"], 0.0)} for r in rows],
        "warnings":   warnings,
    })


def _robust_vs_mixture(your_lineup, your_profiles, opp_profiles, events, relay_data,
                       mixture, opp_lineup=None, n=3000):
    """Weighted expected (win_prob, margin) of your_lineup across the opponent
    set the optimizer hedged over. Against the single predicted opponent a lineup
    can look great yet be brittle if the opponent surprises you, so we average over
    the mixture. MARGIN is the discriminator — win_prob saturates in a blowout
    (a clearly worse lineup can still win ~100%), margin does not. Falls back to the
    single predicted opp_lineup when no mixture is available."""
    entries = ([(float(e[0]), e[1]) for e in mixture] if mixture
               else ([(1.0, opp_lineup)] if opp_lineup else []))
    if not entries:
        return None, None
    total_w = sum(w for w, _ in entries) or 1.0
    ewp = emargin = 0.0
    for w, opp_l in entries:
        ms = match_stats(simulate_match(your_lineup, opp_l, your_profiles,
                                        opp_profiles, events, relay_data, n=n))
        ewp     += (w / total_w) * float(ms.get("win_prob") or 0.0)
        emargin += (w / total_w) * float(ms.get("margin") or 0.0)
    return ewp, emargin


@app.route("/api/promote_lineup", methods=["POST"])
def api_promote_lineup():
    """Promote a coach-edited lineup to be the new 'optimal'. Guardrails:
      - valid lineup (no double-booking, nobody over the 2-event cap);
      - must hold up against the SAME opponent-uncertainty mixture the optimizer
        hedged over — a swap can beat the single predicted opp yet be more brittle,
        and we won't silently trade that away.
    On success it repackages the lineup (rows, lanes, scenarios) with the same
    helpers a normal run uses and replaces the cached + saved optimal. The
    calibrated headline forecast is anchored on coach behavior (not the displayed
    lineup), so it carries over unchanged."""
    your_profiles = _cache.get("your_profiles")
    opp_profiles  = _cache.get("opp_profiles")
    events        = _cache.get("events")
    relay_data    = _cache.get("relay_data")
    opp_lineup    = _cache.get("opp_predicted_full")
    opt           = _cache.get("lineup_data") or {}
    if not (your_profiles and opp_profiles and events and opp_lineup and opt):
        return jsonify({"error": "No active lineup in memory — re-run the optimizer "
                                 "(profiles aren't kept after a server restart)."}), 409

    from Optimizer import MAX_EVENTS
    posted = (request.get_json(silent=True) or {}).get("lineup") or {}

    # ── Validate + build a scorable lineup ───────────────────────────────────
    edited, warnings = {}, []
    ev_by_swimmer = defaultdict(set)
    for ev in posted:
        if ev not in opp_lineup:
            continue
        _, _, stroke = parse_event(ev)
        valid, seen = [], set()
        for s in (posted.get(ev) or []):
            if not s:
                continue
            if s in seen:
                return jsonify({"ok": True, "promoted": False,
                    "reason": f"{s} is entered twice in {ev} — fix the double-booking first."}), 200
            seen.add(s)
            if s in your_profiles and stroke in your_profiles[s].get("strokes", {}):
                valid.append(s); ev_by_swimmer[s].add(ev)
            else:
                warnings.append(s)
        edited[ev] = {"swimmers": valid, "expected_points": 0.0}
    over = sorted(n for n, evs in ev_by_swimmer.items() if len(evs) > MAX_EVENTS)
    if over:
        return jsonify({"ok": True, "promoted": False,
            "reason": f"{over[0]} is entered in more than {MAX_EVENTS} events — fix that first."}), 200

    # ── Robust gate: must not be LESS robust than the current optimal ─────────
    mixture    = _cache.get("robust_mixture")
    cur_lineup = _cache.get("your_lineup_full") or {
        r["event"]: {"swimmers": r.get("swimmers", [])} for r in opt.get("rows", [])}
    basis = "the opponent-uncertainty set" if mixture else "the predicted opponent"
    cand_wp, cand_m = _robust_vs_mixture(edited, your_profiles, opp_profiles, events,
                                         relay_data, mixture, opp_lineup=opp_lineup)
    opt_wp,  opt_m  = _robust_vs_mixture(cur_lineup, your_profiles, opp_profiles, events,
                                         relay_data, mixture, opp_lineup=opp_lineup)
    # Hold off if the lineup is robustly WORSE: less likely to win, or (the usual
    # case in a lopsided meet, where win% is pinned near 100) a lower expected
    # margin across the opponent set. Margin is the real discriminator.
    EPS_WP, EPS_M = 0.01, 0.5
    worse_wp = (cand_wp is not None and opt_wp is not None and cand_wp + EPS_WP < opt_wp)
    worse_m  = (cand_m  is not None and opt_m  is not None and cand_m  + EPS_M  < opt_m)
    if worse_wp or worse_m:
        detail = (f"it wins {cand_wp*100:.0f}% vs the current optimal's {opt_wp*100:.0f}%"
                  if worse_wp else
                  f"its expected margin is {cand_m:+.1f} pts vs the current optimal's {opt_m:+.1f}")
        return jsonify({"ok": True, "promoted": False, "basis": basis,
            "cand_winp": round((cand_wp or 0) * 100, 1), "opt_winp": round((opt_wp or 0) * 100, 1),
            "cand_margin": round(cand_m or 0, 1), "opt_margin": round(opt_m or 0, 1),
            "reason": (f"Held off — across {basis}, {detail}, so it isn't an improvement once "
                       f"the opponent's possible lineups are accounted for. (A swap can score "
                       f"better against the single predicted lineup yet do worse if they "
                       f"surprise you — that's what the optimizer was hedging against.)")}), 200

    # ── Passed: repackage the edited lineup as the new optimal ────────────────
    rows, mc_total_raw = _build_lineup_rows(edited, opp_lineup, your_profiles, opp_profiles)
    try:   # conserving scoresheet: opp_pts = points awarded − ours (mirror _run_and_cache)
        def _alloc(ev):
            st = parse_event(ev)[2]
            ny = sum(1 for s in edited.get(ev, {}).get("swimmers", [])
                     if st in your_profiles.get(s, {}).get("strokes", {}))
            no = sum(1 for s in opp_lineup.get(ev, {}).get("swimmers", [])
                     if st in opp_profiles.get(s, {}).get("strokes", {}))
            return sum((5, 3, 1)[i] for i in range(min(ny + no, 3)))
        for r in rows:
            r["opp_pts"] = max(0.0, _alloc(r["event"]) - float(r.get("mc_pts", 0.0)))
    except Exception:
        pass
    cfg = _cache.get("config") or {}
    scenarios = _compute_scenarios(edited, your_profiles, opp_profiles, events,
                                   relay_data=relay_data, opp_team_name=opt.get("opp_team"),
                                   year=cfg.get("year"), week=opt.get("week_num"))
    your_is_home  = bool(cfg.get("your_is_home", True))
    fillers       = _compute_event_fillers(edited, your_profiles)
    event_payload = _build_event_payload(edited, your_profiles, opp_lineup,
                                         _cache.get("opp_actual"), opp_profiles,
                                         rows, your_is_home, fillers)
    new_data = dict(opt)              # carry the calibrated headline (lineup-independent anchor)
    new_data["rows"]      = rows
    new_data["scenarios"] = scenarios
    new_data["events"]    = event_payload
    if cand_wp is not None:
        new_data["robust_ewp"] = cand_wp
    _cache["lineup_data"]      = new_data
    _cache["your_lineup_full"] = edited
    save_lineup(edited, os.path.join(BASE_DIR, "your_lineup.json"))
    _save_lineup_snapshot()
    pred = next((s for s in scenarios if s.get("label") == "Their predicted lineup"), None)
    return jsonify({"ok": True, "promoted": True, "basis": basis,
        "win_pct": ((pred or {}).get("win_prob") or 0.0) * 100,
        "dropped": warnings})


# ── Helpers ────────────────────────────────────────────────────────────────────

def _compute_scenarios(your_lineup, your_profiles, opp_profiles, events, relay_data=None,
                       opp_team_name=None, year=None, week=None):
    """
    Evaluate our lineup against 3 opponent assumptions.
    Returns list of {label, description, win_prob, our_median, opp_median, margin}.
    Both sides are simulated fully (individual + relay) — win_prob = P(our total > opp total).
    """
    if relay_data is None:
        relay_data = _build_relay_results(your_profiles, opp_profiles)

    N = 10000

    def _m(opp_lineup):
        match = simulate_match(your_lineup, opp_lineup, your_profiles, opp_profiles,
                               events, relay_data, n=N)
        return match_stats(match)

    scenarios = []

    # 1. vs their actual last week lineup (from saved file)
    last_week_path = os.path.join(BASE_DIR, "opp_lineup.json")
    if os.path.exists(last_week_path):
        with open(last_week_path) as f:
            last_week = _normalize_lineup(json.load(f))
        scenarios.append({
            "label": "Their last week lineup",
            "desc":  "The lineup they actually swam last meet — the safest guess.",
            **_m(last_week),
        })

    # 2. vs their predicted lineup (v5 coach_predictor — was self_optimal, far less accurate)
    self_opt = _predict_opp_lineup_or_fallback(opp_team_name, year, week, opp_profiles, events)
    scenarios.append({
        "label": "Their predicted lineup",
        "desc":  "What we think their coach will run this week.",
        **_m(self_opt),
    })

    # 3. vs their best counter (they know our exact lineup)
    # Use greedy counter (LP-free) — this is a display scenario, not the lineup answer.
    our_entries      = lineup_to_entries(your_profiles, your_lineup, events)
    their_counter, _ = _greedy_counter_lineup(opp_profiles, our_entries, events)
    scenarios.append({
        "label": "Their best counter to us",
        "desc":  "Their perfect counter if they somehow knew our exact lineup — worst case for us.",
        **_m(their_counter),
    })

    return scenarios


def win_class(pct):
    if pct >= 99: return "win100"
    if pct >= 75: return "win75"
    if pct >= 50: return "win50"
    if pct >= 25: return "win25"
    return ""


def _team_data(results, team_name):
    """Look up a team's event dict from scraper results, case-insensitive."""
    if team_name in results:
        return results[team_name]
    for k, v in results.items():
        if k.lower() == team_name.lower():
            return v
    raise KeyError(
        f"'{team_name}' not found in scraped results. "
        f"Available: {list(results.keys())}. "
        "Delete results.json and retry to force a fresh scrape."
    )


def _run_and_cache(your_team, opp_team, year,
                   use_opp_fingerprint=True, absent_swimmers=None,
                   opp_absent_swimmers=None,
                   use_imputation=True, your_excludes=None, opp_excludes=None,
                   swimup_only_if_scoring=False):
    """
    Run the optimizer and cache results. Profiles must already be loaded into
    _cache via /api/load_setup. (If not present, falls back to scrape — slow.)

    use_imputation: if True, augment profiles with prior-year z-score imputation
                    for swimmers missing current-year data in some events.
    your_excludes / opp_excludes: lists of swimmer names (normalized or raw)
                    to EXCLUDE from imputation pool (i.e., not on team this year).
    """
    absent_set     = {normalize_name(n) for n in (absent_swimmers or [])}
    opp_absent_set = {normalize_name(n) for n in (opp_absent_swimmers or [])}

    # Use pre-loaded profiles from /api/load_setup if present (fast path)
    setup = _cache.get("setup") or {}
    profiles_ready = (
        setup.get("your_team") == your_team
        and setup.get("opp_team") == opp_team
        and setup.get("year") == year
        and _cache.get("your_profiles") is not None
        and _cache.get("opp_profiles")  is not None
    )

    if profiles_ready:
        your_profiles = _cache["your_profiles"]
        opp_profiles  = _cache["opp_profiles"]
        events        = _cache["events"]
        print("[run] using pre-loaded profiles from /api/load_setup", flush=True)
    else:
        # Fallback: scrape + build from scratch (slow path, for backward compat)
        from Optimizer import build_profiles_recency_weighted
        print("[run] no pre-loaded profiles; scraping fresh (slow)", flush=True)
        results = _scrape_or_load(your_team, opp_team, year)
        your_dated = _build_dated_results(_team_data(results, your_team))
        opp_dated  = _build_dated_results(_team_data(results, opp_team))
        ladder_us = _load_ladder_for_team(your_team)
        ladder_op = _load_ladder_for_team(opp_team)
        ladder_status = {"our": None, "opp": None}
        if ladder_us:
            added, skipped, _ = _merge_ladder_into_dated(your_dated, ladder_us)
            ladder_status["our"] = {"added": added, "skipped": skipped, "total": len(ladder_us)}
        if ladder_op:
            added, skipped, _ = _merge_ladder_into_dated(opp_dated, ladder_op)
            ladder_status["opp"] = {"added": added, "skipped": skipped, "total": len(ladder_op)}
        if (mtu := _manual_times_as_ladder(your_team)):
            _merge_ladder_into_dated(your_dated, mtu, max_date=None)
        if (mto := _manual_times_as_ladder(opp_team)):
            _merge_ladder_into_dated(opp_dated, mto, max_date=None)
        your_profiles = build_profiles_recency_weighted(your_dated, decay=0.7, cur_year=year)
        opp_profiles  = build_profiles_recency_weighted(opp_dated, decay=0.7, cur_year=year)
        events        = sorted(set(your_dated.keys()) | set(opp_dated.keys()))
        _cache["ladder_status"] = ladder_status

    # ── EVENT-UNIVERSE fix (app_calibrated.py — 2026-06-13) ───────────────────
    # The event list is derived from the two teams' observed data, so in a
    # lopsided matchup whole standard events can be missing from the union and
    # silently score 0-0 — they're contested in the real meet (the deeper team
    # sweeps). Union in every standard league event (parse_event-valid keys of
    # the season baseline) so all 40 are scored. Small effect (~4% of the
    # opponent-strength residual in backtest) but it removes a real artifact and
    # cannot hurt: events where neither team has a profiled swimmer stay 0-0.
    def _is_std_event(ev):
        try:
            parse_event(ev); return True
        except Exception:
            return False
    try:
        std = {ev for ev in _build_league_baselines(year) if _is_std_event(ev)}
        events = sorted(set(events) | std)
    except Exception as e:
        print(f"[event-universe] SKIPPED ({e})", flush=True)

    # Drop swimmers marked absent for this week — they're not available
    if absent_set:
        before = len(your_profiles)
        your_profiles = {n: p for n, p in your_profiles.items()
                         if normalize_name(n) not in absent_set}
        dropped = before - len(your_profiles)
        print(f"[absences] excluded {dropped} swimmer(s) from {your_team}: "
              f"{sorted(absent_set)}", flush=True)

    # Same for known opponent absences (e.g. a star sitting out their meet) —
    # drop them so their predicted lineup is built without those swimmers.
    if opp_absent_set:
        before = len(opp_profiles)
        opp_profiles = {n: p for n, p in opp_profiles.items()
                        if normalize_name(n) not in opp_absent_set}
        dropped = before - len(opp_profiles)
        print(f"[absences] excluded {dropped} swimmer(s) from {opp_team}: "
              f"{sorted(opp_absent_set)}", flush=True)

    # Drop swimmers the user marked as no longer on the team (the returning-swimmer
    # review "unchecks"). Like absences, these must be removed from the profiles
    # OUTRIGHT — not merely skipped during imputation — because a base profile built
    # from last year's history (the W1 / no-current-data case) already contains them,
    # so skipping imputation alone leaves them in the lineup. This was the bug.
    # Default to the saved roster-excludes when a caller omits them (passes None),
    # so a swimmer the coach marked off the team is never fielded by an
    # exclude-less code path. api_run always passes an explicit list, so this only
    # backstops other callers; pass [] to deliberately field everyone.
    if your_excludes is None:
        your_excludes = sorted(_load_roster_excludes(your_team))
    if opp_excludes is None:
        opp_excludes = sorted(_load_roster_excludes(opp_team))
    your_drop_norm = {normalize_name(n) for n in (your_excludes or [])}
    opp_drop_norm  = {normalize_name(n) for n in (opp_excludes  or [])}
    if your_drop_norm:
        before = len(your_profiles)
        your_profiles = {n: p for n, p in your_profiles.items()
                         if normalize_name(n) not in your_drop_norm}
        print(f"[excludes] removed {before - len(your_profiles)} non-returner(s) "
              f"from {your_team}", flush=True)
    if opp_drop_norm:
        before = len(opp_profiles)
        opp_profiles = {n: p for n, p in opp_profiles.items()
                        if normalize_name(n) not in opp_drop_norm}
        print(f"[excludes] removed {before - len(opp_profiles)} non-returner(s) "
              f"from {opp_team}", flush=True)

    # Drop swimmers who have AGED OUT of the league (projected NVSL age >= 19). Exact
    # ages from results make this definitive — they can't compete, so no user review:
    # they never appear in the returning-swimmer list and are never fielded (fixes the
    # Sasha Reyes / Colin Bramble case — 18 last season, 19 now).
    your_aged = _aged_out_norm(your_team, year)
    if your_aged:
        before = len(your_profiles)
        your_profiles = {n: p for n, p in your_profiles.items()
                         if normalize_name(n) not in your_aged}
        print(f"[aged-out] dropped {before - len(your_profiles)} >=19 swimmer(s) from {your_team}", flush=True)
    opp_aged = _aged_out_norm(opp_team, year)
    if opp_aged:
        before = len(opp_profiles)
        opp_profiles = {n: p for n, p in opp_profiles.items()
                        if normalize_name(n) not in opp_aged}
        print(f"[aged-out] dropped {before - len(opp_profiles)} >=19 swimmer(s) from {opp_team}", flush=True)

    # Apply imputation: augment profiles with prior-year-z-score data for swimmers
    # missing current-year coverage. Excludes the user's "not on team" picks.
    imp_stats = {"used": False}
    if use_imputation:
        try:
            your_excl_norm = {normalize_name(n) for n in (your_excludes or [])}
            opp_excl_norm  = {normalize_name(n) for n in (opp_excludes  or [])}
            # Also auto-exclude absent swimmers from imputation (they're gone this week)
            your_excl_norm |= absent_set
            opp_excl_norm  |= opp_absent_set
            baselines = _all_league_baselines()
            # IMPUTE_8U (default off): for OUR lineup, don't fabricate 8U entries in
            # strokes a kid has never swum — it breaks the ladder and isn't a real
            # coaching option. The OPPONENT keeps 8U imputation (they really will field
            # that depth; thinning it would make us over-predict beating them).
            impute_8u_ours = os.environ.get("IMPUTE_8U", "0") == "1"
            your_profiles, ypa, yla = augment_profiles_with_imputation(
                your_profiles, your_team, year, your_excl_norm, baselines,
                impute_8u_cross=impute_8u_ours)
            opp_profiles,  opa, ola = augment_profiles_with_imputation(
                opp_profiles,  opp_team,  year, opp_excl_norm,  baselines,
                impute_8u_cross=True)
            imp_stats = {"used": True,
                         "your_prior": ypa, "your_league": yla,
                         "opp_prior":  opa, "opp_league":  ola}
            print(f"[imputation] {your_team}: +{ypa} prior-year + {yla} league-avg stroke profiles", flush=True)
            print(f"[imputation] {opp_team}:  +{opa} prior-year + {ola} league-avg stroke profiles", flush=True)
        except Exception as e:
            print(f"[imputation] FAILED, falling back to raw profiles: {e}", flush=True)

    # Coverage blend (W2-W5 cure, wired 2026-05-30 — see coverage_blend.py and
    # improvements.md issue #1). Adds raw prior-year (2024) times for swimmers
    # missing from current profile. Validated to drop bias from -22 → +0.5,
    # MAE 32 → 20.5 on W2-W5 batch (n=384). Applied to BOTH teams' profiles.
    # W1 is unaffected (cache profile already uses 2024 data wholesale, so the
    # blend is a no-op; W1 architecture is handled separately by w1_predictor).
    # Week number for both the coverage blend (below) and the prior-week lineup
    # fetch (further down). Computed here so the blend can see it — previously
    # the blend referenced an undefined `week`, raising NameError that the broad
    # except silently swallowed as "SKIPPED", disabling the cure on every run.
    cfg_now    = _cache.get("config", {})
    week_num_n = cfg_now.get("week_num") or 0

    try:
        from coverage_blend import apply_coverage_blend
        if week_num_n and week_num_n >= 2:
            n_before_us = len(your_profiles)
            n_before_opp = len(opp_profiles)
            # Pass absent + user-exclude sets so the blend doesn't resurrect swimmers the
            # coach marked unavailable (they were already dropped from the profile above;
            # without this, coverage_blend re-adds them from prior-year data, silently
            # undoing the absence marking).
            your_cb_excl = set(absent_set) | {normalize_name(n) for n in (your_excludes or [])}
            opp_cb_excl  = set(opp_absent_set) | {normalize_name(n) for n in (opp_excludes or [])}
            your_profiles = apply_coverage_blend(your_profiles, your_team, year, exclude_norm=your_cb_excl)
            opp_profiles  = apply_coverage_blend(opp_profiles,  opp_team,  year, exclude_norm=opp_cb_excl)
            print(f"[coverage_blend] {your_team}: {n_before_us} -> {len(your_profiles)} swimmers", flush=True)
            print(f"[coverage_blend] {opp_team}:  {n_before_opp} -> {len(opp_profiles)} swimmers", flush=True)
    except Exception:
        print(f"[coverage_blend] SKIPPED:\n{traceback.format_exc()}", flush=True)

    _cache["your_profiles"] = your_profiles
    _cache["opp_profiles"]  = opp_profiles
    _cache["events"]        = events
    _cache["absent_count"]  = len(absent_set)
    _cache["imputation"]    = imp_stats

    # Relay uses fastest eligible swimmers per age group — simulate stochastically
    relay_data    = _build_relay_results(your_profiles, opp_profiles)
    relay_exp_pts = _sum_relay_exp_pts(relay_data)          # for display only

    # Load BOTH teams' real prior-week lineups (from mynvsl results) so the robust
    # mixture's "Reuse last week" + "Counter our last week" scenarios use real data.
    # (Prior bug: opp's prior was being read from opp_lineup.json, which gets
    #  overwritten on every Run with self_optimal — so "reuse last week" was a
    #  duplicate of "self_optimal," wasting 10% of mixture weight. This drove
    #  the optimizer to over-spread opp's top swimmers across events.)
    prior_our_lineup = None
    prior_opp_lineup = None
    if week_num_n and week_num_n > 1:
        try:
            prior_our_lineup, _, _ = _fetch_team_prev_lineup(your_team, week_num_n, year)
        except Exception as e:
            print(f"[run] prev lineup fetch failed for {your_team}: {e}", flush=True)
        try:
            prior_opp_lineup, _, _ = _fetch_team_prev_lineup(opp_team, week_num_n, year)
        except Exception as e:
            print(f"[run] prev lineup fetch failed for {opp_team}: {e}", flush=True)

    robust_breakdown = None
    robust_mixture   = None
    robust_ewp       = None
    robust_norm_winp = None
    robust_their_ewp = None

    # Detect upload state for the W1 asymmetric/symmetric branch. If we have a
    # ladder for our team, our profile reflects current-season times → opp
    # should be augmented with v4Rt phantoms to compensate. If we don't, both
    # teams are on 2024 baselines and we keep raw-vs-raw for symmetry.
    ladder_status = _cache.get("ladder_status") or {}
    our_has_current  = bool((ladder_status.get("our")  or {}).get("added"))
    opp_has_current  = bool((ladder_status.get("opp")  or {}).get("added"))

    your_lineup, robust_ewp, robust_mixture, robust_breakdown = strategy_robust(
        your_profiles, opp_profiles, events, relay_data,
        prior_opp_lineup=prior_opp_lineup,
        prior_our_lineup=prior_our_lineup,
        opp_team_name=opp_team,
        use_opp_fingerprint=use_opp_fingerprint,
        year=year, week=week_num_n,
        has_our_current_data=our_has_current,
    )
    # Keep the opponent mixture the optimizer hedged over so /api/promote_lineup
    # can re-check a coach-edited lineup against the SAME uncertainty set (a swap
    # can beat the single predicted opp yet be more brittle).
    _cache["robust_mixture"] = robust_mixture
    # Headline win prob now uses a calibrated probit (issue #2) applied
    # to the simulate_match margin below — see _w25_winprob. This REPLACES the old
    # symmetric opp-perspective strategy_robust run (the "whoever runs the optimizer
    # wins" normalization), which cost ~60s and was only ever a crude de-bias. The
    # analytical form is faster and calibrated on held-out data. robust_norm_winp /
    # robust_their_ewp stay None (kept in the payload for back-compat).
    # Display opp lineup = v5 prediction (was self_optimal, now realistic predictor)
    opp_lineup = _predict_opp_lineup_or_fallback(opp_team, year, week_num_n, opp_profiles, events)
    # Hybrid 8U: keep opp's real 8U swimmers, fill empty slots with division-typical
    # depth (fixes the v5 under-fielded-8U sweep). Rebinds opp_profiles to the
    # augmented copy so all downstream scoring stays consistent. Older bands untouched.
    # W1-ONLY: the hybrid 8U fill exists because at W1 the opponent's 8U is almost
    # entirely unknown (no current-season races yet) and the predictor under-fields
    # it, so we'd phantom-sweep. From W2 on, the opponent has real 8U race data and
    # the fill (a) isn't needed and (b) injects confusing synthetic "div-avg"
    # swimmers calibrated on W1 turnout. W2-W5 bias was validated WITHOUT this fill
    # (prod_eval pre-4069d2f), so gating it here restores that validated config.
    if week_num_n == 1:
        try:
            opp_lineup, opp_profiles = _hybrid_fill_opp_8u(
                opp_team, year, opp_lineup, opp_profiles, events)
            # CRITICAL: rebind cache to the hybrid-augmented opp_profiles, otherwise
            # /api/score_lineup reads the pre-hybrid copy and can't resolve the
            # synthetic 8U swimmer names — opp_mc comes back empty, we artificially
            # sweep 8U breast/fly, and Check shows phantom +5pt deltas vs cached.
            _cache["opp_profiles"] = opp_profiles
            # Relays were built (above) against the PRE-fill opponent 8U, which at W1 is
            # artificially thin/slow — so we phantom-win the 8U relays (e.g. Girls 8U
            # Free read 100% when reality is ~5%). Rebuild them against the hybrid-filled
            # opp so the 8U relays use realistic division-average depth.
            relay_data    = _build_relay_results(your_profiles, opp_profiles)
            relay_exp_pts = _sum_relay_exp_pts(relay_data)
        except Exception as e:
            print(f"[hybrid8u] SKIPPED ({e})", flush=True)
    # Stash v5's prediction separately so the /opponent tab shows the predicted
    # current-week lineup rather than the (also-cached) historical prior-week lineup.
    _cache["opp_predicted_lineup"] = opp_lineup
    _cache["opp_predicted_for"]    = {"year": year, "week": week_num_n, "team": opp_team}

    # ── Swim-up polish pass (added 2026-06-03) ───────────────────────────────
    # strategy_robust's CRN-locked MC can miss small home-band-vs-swim-up moves
    # in sparse-opp cases (validated SHB vs SHR W1: missed +2.48 pts moving
    # Benton from 11-12 50-back back to 9-10 50-back). This polish uses the
    # headline simulate_match engine to verify each swim-up — if home-band
    # placement gains >=0.5 pts of margin, swap. Cost ~5s per meet.
    #
    # KNOWN (by design, not a bug): the polish passes optimize the MODEL's self-score,
    # so they nudge the DISPLAYED predicted total up by ~+1.8 pts on average (decompose,
    # 2026-06). The lineup we *recommend* is genuinely better; the *predicted number* is
    # the optimizer's own (slightly inflated) self-assessment. If we ever want the shown
    # score to be a calibrated prediction rather than the optimizer-max, decouple them
    # (predict from a baseline, recommend the polished lineup). See improvements.md.
    try:
        polished_lineup, polish_log = _polish_swim_ups(
            your_lineup, your_profiles, opp_profiles, events,
            opp_lineup, relay_data,
            accept_threshold=0.5, n_sim=4000,
            baselines=(_build_league_baselines(year) or _build_league_baselines(year - 1)),
            swimup_aversion=0.5,
        )
        for move in polish_log:
            tag = "ACCEPTED" if move.get("accepted") else "rejected"
            delta = f"{move.get('delta', 0):+.2f}" if "delta" in move else "n/a"
            reason = f" ({move.get('reason')})" if move.get("reason") else ""
            print(f"[polish] {tag}: {move['swimmer']} {move['from']} → "
                  f"{move['to']}  delta={delta}{reason}", flush=True)
        n_accepted = sum(1 for m in polish_log if m.get("accepted"))
        if n_accepted:
            print(f"[polish] applied {n_accepted} swim-up correction(s)", flush=True)
            your_lineup = polished_lineup
    except Exception as e:
        print(f"[polish] SKIPPED ({e})", flush=True)

    # ── Within-band cross-event swap polish (added 2026-06) ──────────────────
    # Catches same-band stroke trades the optimizer missed — e.g. Russell ↔
    # one swimmer across back/fly in 15-18 Boys (validated +0.87-1.04 on SHB W1).
    # Cost ~30-120s. Order matters: must run AFTER swim-up polish so we don't
    # have to re-validate within-band constraints around swim-up moves.
    try:
        # fresh = swimmers with a current-season time (uploaded ladder); everyone
        # else is "stale" and gets the natural-stroke preference weighted up.
        _fresh = {normalize_name(e["Name"])
                  for e in (_load_ladder_for_team(your_team) or []) if e.get("Name")}
        # PURE SCORE (2026-06-16): the natural-stroke / specialty preference was
        # DOMINATING the objective — it accepted 20 reshuffling moves with score
        # Δ≈0 (some negative), spending the optimizer's effort on stroke-realism
        # instead of points. style_weight=specialty_weight=0 makes the polish
        # accept a swap ONLY when it genuinely improves expected score (> threshold).
        your_lineup, within_log = _polish_within_band_swaps(
            your_lineup, your_profiles, opp_profiles, events,
            opp_lineup, relay_data,
            accept_threshold=0.1, n_sim=1000, prune_downgrade_secs=2.0,
            style_weight=0.0,
            baselines=(_build_league_baselines(year) or _build_league_baselines(year - 1)),
            fresh_swimmers=_fresh, specialty_weight=0.0, stale_boost=2.0,
        )
    except Exception as e:
        print(f"[within-polish] SKIPPED ({e})", flush=True)

    # Compute opt_score as expected points vs that representative opp
    opp_entries = lineup_to_entries(opp_profiles, opp_lineup, events)
    your_score  = sum(
        race_points(
            [your_profiles[s]["strokes"][parse_event(e)[2]]["mean"]
             for s in your_lineup[e]["swimmers"]
             if s in your_profiles and parse_event(e)[2] in your_profiles[s].get("strokes", {})],
            [m for m, _ in opp_entries.get(e, [])]
        )
        for e in events
    )

    # Operator rule: never field a league-average fill (no real data) in a slot an
    # available real/prior-year swimmer could take. Runs after polish so it corrects
    # the final recommended lineup; genuine fills (no real option) are left in place.
    your_lineup, _n_real_swap = _prefer_real_over_imputed_fill(your_lineup, your_profiles, events)
    if _n_real_swap:
        print(f"[real-over-impute] swapped {_n_real_swap} league-avg fill(s) for real swimmers", flush=True)

    # Coaching option (default off): don't swim a young swimmer up into an older event
    # they can't score in — use a native-age swimmer instead. See _demote_hopeless_swimups.
    if swimup_only_if_scoring:
        your_lineup, _n_demoted = _demote_hopeless_swimups(
            your_lineup, opp_lineup, your_profiles, opp_profiles, events)
        if _n_demoted:
            print(f"[swimup-rule] demoted {_n_demoted} hopeless swim-up(s) to native swimmers", flush=True)

    # Follow the ladder: swap a benched faster swimmer (with a spare event) in for a
    # slower fielded one when it doesn't cost points — fixes a hopeless time parked in
    # a throwaway slot. Runs last so it cleans up the final recommended lineup.
    your_lineup, _n_ladder = _ladder_fill_slots(your_lineup, opp_lineup, your_profiles, opp_profiles, events)
    if _n_ladder:
        print(f"[ladder-fill] swapped {_n_ladder} slower fielded swimmer(s) for faster benched ones", flush=True)

    # Coaching preference: put non-scoring swimmers in their natural (faster) strokes
    # rather than off-events, when it changes no points. Runs last.
    your_lineup, _n_nat = _prefer_natural_events(your_lineup, opp_lineup, your_profiles, opp_profiles, events)
    if _n_nat:
        print(f"[natural-events] moved {_n_nat} swimmer(s) into a more natural stroke (point-neutral)", flush=True)

    save_lineup(your_lineup, os.path.join(BASE_DIR, "your_lineup.json"))
    save_lineup(opp_lineup,  os.path.join(BASE_DIR, "opp_lineup.json"))

    N              = 10000
    rows, mc_total = _build_lineup_rows(your_lineup, opp_lineup, your_profiles, opp_profiles)

    # ── W1 presence-MC headline correction (see _presence_adjusted_rows) ──────
    # Production W1 over-predicts +29 (full 2024 roster fields ghosts + absentees).
    # Replace the all-present headline with the attendance-realistic expectation.
    # Only at W1 AND when we have NO current-season upload (an uploaded ladder
    # already removed ghosts, so applying p=0.60 there would over-deflate). The
    # recommended lineup (swimmers) is unchanged; only the expected score moves.
    mc_total_raw = mc_total
    presence_p   = None
    if week_num_n == 1 and not our_has_current and W1_PRESENCE_MC_ENABLED:
        try:
            per_ev, mc_total = _presence_adjusted_rows(
                your_lineup, opp_lineup, your_profiles, opp_profiles)
            presence_p = W1_PRESENCE_P
            for r in rows:                       # keep per-event rows consistent with headline
                if r["event"] in per_ev:
                    r["mc_pts"] = per_ev[r["event"]]
            print(f"[presence] W1 no-upload: headline {mc_total_raw:.1f} -> {mc_total:.1f} "
                  f"(p={W1_PRESENCE_P})", flush=True)
        except Exception as e:
            mc_total = mc_total_raw
            print(f"[presence] SKIPPED ({e})", flush=True)

    # ── Forfeit discount (participation fantasy, CALIBRATION_STATE §0.008/0.009) ──
    # league_avg fills are phantom lanes a thin team often can't actually field, so
    # the optimizer over-credits weak teams (concentrated in div 15-17). Blend each
    # fill event's points toward its fill-stripped score with probability = the
    # team's MEASURED prior-year no-show rate (1 − participation) — no fitted
    # constants. Applied to the rows so the conserving pass keeps it. Skipped when
    # the W1 presence adjustment already modeled attendance (would double-count).
    if FORFEIT_DISCOUNT_MODE != "off" and presence_p is None:
        try:
            if FORFEIT_DISCOUNT_MODE == "team":
                _fr = max(0.0, min(1.0, 1.0 - float(_team_participation(your_team, year))))
            else:
                _fr = max(0.0, min(1.0, float(FORFEIT_DISCOUNT_MODE)))
            if _fr > 0.0:
                rows, mc_total, _nfe = _forfeit_discount_rows(
                    rows, your_lineup, opp_lineup, your_profiles, opp_profiles, _fr)
                if _nfe:
                    print(f"[forfeit-discount] rate={_fr:.2f} over {_nfe} fill event(s): "
                          f"headline {mc_total_raw:.1f} -> {mc_total:.1f}", flush=True)
        except Exception as e:
            print(f"[forfeit-discount] SKIPPED ({e})", flush=True)

    # Match win probability: simulate against opp's predicted lineup. Reuse the
    # hybrid-filled opp_lineup (don't re-predict — that would feed synthetic 8U
    # swimmers back into v5).
    opp_self_opt   = opp_lineup
    match          = simulate_match(your_lineup, opp_self_opt, your_profiles, opp_profiles,
                                    events, relay_data, n=N)
    mstats         = match_stats(match)
    # Raw (pre-calibration) medians — these conserve (your + opp ≈ the meet pool).
    # The conserving-scoresheet pass below uses them to keep the displayed numbers
    # adding up after calibration shifts only the margin.
    mstats_raw     = dict(mstats) if mstats else None
    # Headline win prob — both weeks use an intercept probit on the simulate_match
    # mean margin, each fit on its own regime's production margins:
    #   W1   — _w1_winprob (w1_winprob_fit.py, 98 W1 sides): near coin-flip regime,
    #          no current-season data; deliberately timid, UI labels low-confidence.
    #   W2-5 — _w25_winprob (overnight cache refit 2026-06-09, 390 sides): the
    #          intercept absorbs the optimizer's ~+24 margin inflation (winner's
    #          curse). Replaces the greedy-margin shrinkage form (issue #2, resolved).
    # Capture the win% RESPONSE FN + the margin it was evaluated at, so the
    # last-week toggle can reuse the SAME mapping (shifted only by the genuine
    # raw-margin difference) — otherwise a same-strength opponent reads a different
    # win% purely from a basis mismatch.
    _winp_fn = None
    if week_num_n == 1:
        _winp_fn = _w1_winprob
        analytical_winp = _w1_winprob(mstats.get("margin", 0.0))
    elif week_num_n and week_num_n >= 2:
        _winp_fn = _w25_winprob
        analytical_winp = _w25_winprob(mstats.get("margin", 0.0))
    else:
        analytical_winp = None
    _winp_margin = mstats.get("margin", 0.0)

    # ── DISPLAY CALIBRATION LAYER (app_calibrated.py only — 2026-06-11) ────────
    # Constants measured on the data-repaired production-basis backtest
    # (mock_final_repaired.jsonl; see CALIBRATION_STATE.md). Two corrections:
    #   1. Value-add shrink: the optimizer's headline total carries winner's-curse
    #      inflation (claimed improvement ~+34..+51/wk vs the real, measured +7).
    #      total_adj_w = (1 − 7/mean_va_w)·mean_va_w, subtracted from the displayed
    #      total, medians, and margin so every number tells the same honest story.
    #   2. Win prob refit (intercept probit) on the CALIBRATED margins — replaces
    #      the raw-margin probits above. NOTE: an origin-forced fit (margin 0 →
    #      50%) is NOT statistically supportable; the intercept carries real
    #      residual opponent-side inflation (option-2 strength terms were tested
    #      and are unknowable pre-meet — see CALIBRATION_STATE.md).
    # The optimizer's LINEUP is untouched — only displayed expectations move.
    try:
        if week_num_n:
            import math
            cal = _load_calibration_constants()
            wk_key = str(min(int(week_num_n), 5))
            adj = cal.get(wk_key, {}).get("total_adj", 0.0)
            # Calibrate the FULL displayed total (individual mc_total + relay). The
            # constants were fit on the full meet score; relay (~30 pts) is essentially
            # fixed, so we carry the adjustment on mc_total and apply the full delta to
            # margin/median (simulate_match already scores the full meet).
            relay_pts = relay_exp_pts or 0.0
            full_raw = mc_total + relay_pts
            #   THEORY 2 — per-meet COACH-PREDICTOR ANCHOR (replaces the per-team
            #   value-add shrink + strength-split). The optimizer's full-pool total
            #   over-credits swimmers who won't actually be fielded (winner's curse +
            #   participation fantasy); validated as ~60% of the over-prediction and
            #   the single largest bias lever. Instead, anchor the displayed total on
            #   the model's score of the COACH-PREDICTED our lineup (what a coach
            #   actually fields — present swimmers, no over-optimization), then add the
            #   earned +7 via the fitted per-week reanchor, and flatten the residual
            #   opponent-strength staircase with ONE linear opp-division term.
            #   Constants: calibration_constants_theory2.json (full-league backtest,
            #   488 sides, 5-fold CV stable). Validated end state: league +7, weeks
            #   flat, ALL 17 divisions within ±10. Predict-time-safe (no truth used).
            t2 = _load_theory2_constants()
            our_pred_raw = _predict_opp_lineup_or_fallback(
                your_team, year, week_num_n, your_profiles, events)
            our_pred = {}
            for _ev in events:
                try:
                    _a, _g, _sf = parse_event(_ev)
                except Exception:
                    continue
                _sws = [s for s in (our_pred_raw.get(_ev, {}).get("swimmers") or [])
                        if s in your_profiles and _sf in your_profiles[s].get("strokes", {})]
                our_pred[_ev] = {"swimmers": _sws, "expected_points": 0}
            _cr, mc_coach = _build_lineup_rows(our_pred, opp_lineup, your_profiles, opp_profiles)
            pred_coach = mc_coach + relay_pts
            reanc = t2["reanchor"].get(wk_key, 0.0)
            division = _team_division(your_team, year)
            div_term = (t2["div_slope_b"] * (division - t2["divmean"])) if division else 0.0
            full = pred_coach + reanc - div_term
            delta = full - full_raw
            mc_total_precal = mc_total
            mc_total = mc_total + delta
            if mstats:
                mstats = dict(mstats)
                mstats["margin"] = mstats.get("margin", 0.0) + delta
                if mstats.get("our_median") is not None:
                    mstats["our_median"] += delta
            # NOTE: win % is no longer computed here. It is computed AFTER the
            # conserving pass below, from the SAME honest displayed margin the
            # points show (mstats["margin"] = your_full − opp_full), so the points
            # and the win % finally tell one story. See the conserve block.
            print(f"[calibration-theory2] W{week_num_n} D{division}: optim {mc_total_precal:.1f} "
                  f"-> coach-anchored {mc_total:.1f} (pred_coach {pred_coach:.1f}, reanc {reanc:+.1f}, "
                  f"div_term {-div_term:+.1f})", flush=True)
    except Exception as e:
        print(f"[calibration] SKIPPED ({e})", flush=True)

    # ── CONSERVING SCORESHEET (2026-06) ───────────────────────────────────────
    # Make the displayed numbers add up like a real meet sheet WITHOUT distorting
    # the per-event points: each event keeps its GENUINE expected points (a sweep
    # ≈ 8, a blowout loss ≈ 1, a toss-up ≈ 4.5), the team totals are their sums,
    # and your + opp totals conserve to the meet pool (~420). The honesty
    # de-inflation lives ONLY in the calibrated win % (analytical_winp above) — NOT
    # smeared across the points, which previously pulled every event toward 4.5 and
    # made guaranteed sweeps read like toss-ups.
    try:
        if mstats and mstats_raw and mstats_raw.get("opp_median") is not None:
            pool = float(mstats_raw["our_median"]) + float(mstats_raw["opp_median"])

            def _event_alloc(ev):
                """Points awarded in an event = top min(finishers, 3) places (5-3-1)."""
                st = parse_event(ev)[2]
                ny = sum(1 for s in your_lineup.get(ev, {}).get("swimmers", [])
                         if st in your_profiles.get(s, {}).get("strokes", {}))
                no = sum(1 for s in opp_lineup.get(ev, {}).get("swimmers", [])
                         if st in opp_profiles.get(s, {}).get("strokes", {}))
                return sum((5, 3, 1)[i] for i in range(min(ny + no, 3)))

            for r in rows:
                # Genuine per-event your points stay as-is; opp = points awarded − yours.
                ye = float(r.get("mc_pts", 0.0))
                r["opp_pts"] = max(0.0, _event_alloc(r["event"]) - ye)

            mc_total  = sum(float(r.get("mc_pts", 0.0)) for r in rows)   # rows sum to this
            your_full = mc_total + float(relay_exp_pts or 0.0)
            # ── League reanchor (OVERNIGHT_SYSTEMATIC's surgical fix, wired 2026-08) ──
            # One measured league-level constant (calibration_constants.json key
            # "_league_reanchor", absent/0 = off), applied MULTIPLICATIVELY to the
            # rows so the scoresheet stays coherent (a sweep stays a sweep; rows
            # still sum to the headline; opp gets the remainder via opp_pts below).
            # This is the flat league-wide layer the D17 addendum recommended —
            # it removes the uniform raw-total elevation, NOT per-division fits.
            try:
                _reanc_c = float((_load_calibration_constants() or {}).get("_league_reanchor", 0.0))
            except Exception:
                _reanc_c = 0.0
            if _reanc_c and mc_total > max(1.0, _reanc_c):
                _scale = (mc_total - _reanc_c) / mc_total   # relay rows stay honest; the
                for r in rows:                              # individual rows absorb all of C
                    r["mc_pts"] = float(r.get("mc_pts", 0.0)) * _scale
                mc_total  = sum(float(r.get("mc_pts", 0.0)) for r in rows)
                your_full = mc_total + float(relay_exp_pts or 0.0)
                print(f"[league-reanchor] -{_reanc_c:.1f} via x{_scale:.3f} on individual rows", flush=True)
            mstats = dict(mstats)
            mstats["our_median"] = your_full
            mstats["opp_median"] = pool - your_full
            mstats["margin"]     = your_full - (pool - your_full)
            # Win % reads the SAME honest margin the points show — origin-forced
            # Φ(margin/σ), σ=50 (validated on 488 bias-corrected backtest sides:
            # MLE σ≈49.06→50, intercept fits ~0, per-week overfits; see
            # WIN_PCT_HANDOFF.md + the fit panel). Displayed tie → 50%, +40 → ~79%.
            # Replaces the σ=30 / theory-2-margin win % removed above.
            import math
            _winp_sigma = float((_load_calibration_constants() or {}).get("_winprob_sigma_margin", 50.0))
            _winp_fn = lambda m, _s=_winp_sigma: 0.5 * (1.0 + math.erf((float(m) / _s) / math.sqrt(2.0)))
            _winp_margin = float(mstats["margin"])
            analytical_winp = _winp_fn(_winp_margin)
            print(f"[conserve] your {your_full:.1f} + opp {pool - your_full:.1f} = {pool:.1f} "
                  f"(margin {_winp_margin:+.1f} -> win% {100*analytical_winp:.0f}%, sigma {_winp_sigma:.0f})",
                  flush=True)
            _rawo = float(mstats_raw['our_median']); _rawp = float(mstats_raw['opp_median'])
            print(f"[debug-basis] raw_our {_rawo:.1f} raw_opp {_rawp:.1f} raw_margin {_rawo-_rawp:+.1f} | "
                  f"conserve your_full {your_full:.1f} margin {_winp_margin:+.1f} | "
                  f"theory2_anchored_mc_total {mc_total:.1f}", flush=True)
    except Exception as e:
        print(f"[conserve] SKIPPED ({e})", flush=True)

    # Legacy pct-to-210 (kept for secondary display)
    our_totals     = simulate_our_total(your_lineup, opp_lineup, your_profiles, opp_profiles, events, n=N)
    relay_trials   = simulate_relay_trials(relay_data, n=N)
    stats          = score_stats(our_totals + relay_trials, target=TARGET_SCORE)

    scenarios      = _compute_scenarios(your_lineup, your_profiles, opp_profiles, events,
                                        relay_data=relay_data,
                                        opp_team_name=opp_team, year=year, week=week_num_n)

    # Opp last-week (actual) lineup for the lineup-page toggle.
    # api_run already called _fetch_opp_prev_lineup, which stashes the result in
    # _cache["opp_actual"]. Reuse that — no need to re-fetch (and re-fetching here
    # would overwrite opp_lineup.json after _compute_scenarios already read it).
    opp_lastweek = _cache.get("opp_actual")

    your_is_home = bool(_cache.get("config", {}).get("your_is_home", True))
    your_fillers = _compute_event_fillers(your_lineup, your_profiles)
    # Per-event scores vs the opp's LAST-WEEK lineup, so the lineup-page toggle can
    # swap the numbers (not just the names) between predicted and last week.
    rows_lastweek = None
    if opp_lastweek:
        try:
            rows_lastweek, _ = _build_lineup_rows(your_lineup, opp_lastweek,
                                                  your_profiles, opp_profiles)
        except Exception:
            rows_lastweek = None
    event_payload = _build_event_payload(
        your_lineup, your_profiles,
        opp_lineup, opp_lastweek, opp_profiles,
        rows, your_is_home, your_fillers,
        rows_lastweek=rows_lastweek,
    )

    # Whole-meet headline win% + median vs the last-week lineup, on the SAME
    # calibrated basis as predicted: our displayed total is an opp-independent
    # forecast, so only the opponent's total — hence the margin and win% — move.
    # Reuse the raw last-week medians already in the "last week" scenario.
    headline_wp_lastweek = opp_median_lastweek = None
    if opp_lastweek and mstats and mstats.get("our_median") is not None and _winp_fn is not None:
        try:
            _lw = next((s for s in scenarios if s.get("label") == "Their last week lineup"), None)
            if _lw and _lw.get("our_median") is not None:
                # Median: our displayed total is opp-independent, so only the opp
                # side moves (raw last-week medians around the same pool).
                pool_lw  = float(_lw["our_median"]) + float(_lw["opp_median"])
                opp_median_lastweek = pool_lw - float(mstats["our_median"])
                # Win%: reuse the SAME response fn the predicted headline used,
                # shifted by the genuine raw-margin difference (last-week − predicted)
                # so a same-strength opponent reads the same win%.
                raw_pred_margin = float((mstats_raw or {}).get("margin", 0.0))
                raw_lw_margin   = float(_lw.get("margin", raw_pred_margin))
                wlw = _winp_fn(_winp_margin + (raw_lw_margin - raw_pred_margin))
                headline_wp_lastweek = (wlw * 100.0) if wlw is not None else None
        except Exception:
            pass

    _cache["lineup_data"] = {
        "rows": rows, "opt_score": your_score, "mc_total": mc_total,
        "mc_total_raw": mc_total_raw, "presence_p": presence_p,
        "score_stats": stats, "match_stats": mstats, "scenarios": scenarios,
        "relay_exp_pts": relay_exp_pts,
        "robust_ewp": robust_ewp,
        "robust_norm_winp": robust_norm_winp,
        "analytical_winp": analytical_winp,
        "week_num": week_num_n,
        "robust_their_ewp": robust_their_ewp,
        "robust_breakdown": robust_breakdown,
        # New consolidated payload — used by the unified /lineup view
        "events":             event_payload,
        "your_is_home":       your_is_home,
        "your_team":          your_team,
        "opp_team":           opp_team,
        "has_lastweek_opp":   opp_lastweek is not None,
        "headline_wp_lastweek": headline_wp_lastweek,
        "opp_median_lastweek":  opp_median_lastweek,
    }
    _cache["relay_data"]            = relay_data
    _cache["opp_lastweek_lineup"]   = opp_lastweek
    _cache["your_lineup_full"]      = your_lineup
    _cache["opp_predicted_full"]    = opp_lineup
    _save_lineup_snapshot()


def _scrape_or_load(your_team, opp_team, year):
    cache_file = os.path.join(BASE_DIR, "results.json")
    if os.path.exists(cache_file):
        with open(cache_file) as f:
            cached = json.load(f)
        cached_teams = {t.lower() for t in cached.keys()}
        if cached_teams == {your_team.lower(), opp_team.lower()}:
            keys = list(cached.keys())
            if keys[0].lower() != your_team.lower():
                cached = {keys[1]: cached[keys[1]], keys[0]: cached[keys[0]]}
            return cached

    from Scraper import scrape_teams
    results = scrape_teams([your_team, opp_team], year=year)
    with open(cache_file, "w") as f:
        json.dump(results, f, indent=2)
    return results


def _sum_relay_exp_pts(relay_data):
    """Sum expected relay points across both genders (for display only)."""
    total = 0.0
    for gdata in relay_data.values():
        for relay in gdata.get("age_relays", []):
            total += relay.get("exp_pts", 0.0)
        total += gdata.get("mixed", {}).get("exp_pts", 0.0)
    return total


def _sim_one_relay(yt, ot, n):
    """Simulate one relay's per-trial outcome. Returns (our_pts, opp_pts) arrays of shape (n,)."""
    import numpy as np
    if not yt and not ot:
        return np.zeros(n), np.zeros(n)
    if not yt:
        return np.zeros(n), np.full(n, 5.0)
    if not ot:
        return np.full(n, 5.0), np.zeros(n)
    your_t = np.random.normal([m for m, _ in yt], [s for _, s in yt], (n, len(yt))).sum(axis=1)
    opp_t  = np.random.normal([m for m, _ in ot], [s for _, s in ot], (n, len(ot))).sum(axis=1)
    our_wins = your_t < opp_t
    return our_wins * 5.0, (~our_wins) * 5.0


def simulate_relay_match(relay_data, n):
    """Per-trial relay totals for both teams. Returns (our_totals, opp_totals) shape (n,)."""
    import numpy as np
    our_total = np.zeros(n)
    opp_total = np.zeros(n)
    for gdata in relay_data.values():
        for relay in gdata.get("age_relays", []):
            our_w, opp_w = _sim_one_relay(relay.get("your_tups", []), relay.get("opp_tups", []), n)
            our_total += our_w
            opp_total += opp_w
        mixed = gdata.get("mixed", {})
        our_w, opp_w = _sim_one_relay(mixed.get("your_tups", []), mixed.get("opp_tups", []), n)
        our_total += our_w
        opp_total += opp_w
    return our_total, opp_total


def simulate_relay_trials(relay_data, n):
    """Per-trial OUR relay total (back-compat). Calls simulate_relay_match and discards opp."""
    our_total, _ = simulate_relay_match(relay_data, n)
    return our_total


def _swimmer_best_z(profile, home_band, gender, baselines):
    """Lowest (best) z-score across a swimmer's strokes vs the league baseline — a
    one-number measure of how good the swimmer is. Negative = faster than the
    field; a team star runs ≈ −1.5 to −2.5. Returns None if unscorable."""
    if not baselines:
        return None
    best = None
    for stroke_key, sd in (profile.get("strokes") or {}).items():
        m = sd.get("mean")
        if m is None:
            continue
        base = baselines.get(f"{home_band} {gender} {stroke_key}")
        if not base or base.get("std", 0) <= 0:
            continue
        z = (m - base["mean"]) / base["std"]
        if best is None or z < best:
            best = z
    return best


def _polish_swim_ups(your_lineup, your_profiles, opp_profiles, events,
                     opp_lineup, relay_data,
                     accept_threshold=0.5, n_sim=4000,
                     baselines=None, swimup_aversion=0.5):
    """Post-strategy_robust polish for swim-up placements.

    strategy_robust's MC search uses CRN-locked seed + per-event cache for speed,
    which trades a small amount of accuracy. In sparse-opp situations (W1 W1
    against teams whose roster has gaps in some events because of age-up
    correction), this can leave swimmers placed in swim-up events when home-band
    placement would actually score better per the headline simulate_match engine.

    For each swim-up placement in your_lineup, this pass:
      1. Computes the swimmer's home-band equivalent event (same gender + stroke,
         age = home band, distance adjusted per NVSL rules — 8U=25, 9-10=25 fly /
         50 other, 11+=50).
      2. If the swimmer has a time for the home-band event, builds a candidate
         lineup (moves the swimmer, drops the slowest at the destination if at
         capacity, leaves the source short rather than backfilling — keeps the
         move atomic).
      3. Scores baseline vs candidate using simulate_match (same engine as the
         headline number and the /api/score_lineup check endpoint).
      4. Accepts the move if margin gain >= accept_threshold pts.

    Validated 2026-06-03 on SHB vs SHR W1: caught Benton's swim-up (+2.48 pts
    gain when moved 11-12 50-back → 9-10 50-back), correctly left Milan's
    swim-up alone (-0.04 pts when tried 13-14 → 11-12 50-back).

    Cost: ~1.8s per candidate at n_sim=4000. Typically 2-3 swim-ups per W1 meet
    → ~5s total overhead on top of strategy_robust's ~30-60s.

    Returns: (polished_lineup, moves_log) where moves_log is a list of dicts
    describing accepted/rejected moves for diagnostics.
    """
    import numpy as np
    from copy import deepcopy
    from Optimizer import AGE_GROUP_ORDER, MAX_PER_EVENT

    polished = deepcopy(your_lineup)
    moves_log = []

    # Find all swim-up placements.
    swim_ups = []  # list of (event_label, swimmer_name, swimmer_home_band)
    for ev_label in list(polished.keys()):
        try:
            ev_age, ev_gender, ev_stroke = parse_event(ev_label)
        except Exception:
            continue
        ev_idx = AGE_GROUP_ORDER.index(ev_age) if ev_age in AGE_GROUP_ORDER else None
        if ev_idx is None: continue
        for sw in list((polished.get(ev_label, {}) or {}).get("swimmers", [])):
            home = (your_profiles.get(sw, {}) or {}).get("home_age_group")
            if not home or home not in AGE_GROUP_ORDER: continue
            home_idx = AGE_GROUP_ORDER.index(home)
            if ev_idx > home_idx:
                swim_ups.append((ev_label, sw, home))

    if not swim_ups:
        return polished, moves_log

    # Score baseline (current lineup) once.
    def _score(lineup):
        match = simulate_match(lineup, opp_lineup, your_profiles, opp_profiles,
                                events, relay_data, n=n_sim)
        return match_stats(match)

    base = _score(polished)
    base_margin = base["margin"]

    # NVSL band/distance helper: returns the distance the swimmer's home band
    # uses for `stroke`.
    def _home_dist(home_band, stroke):
        if home_band == "8U": return 25
        if home_band == "9-10": return 25 if stroke == "fly" else 50
        return 50  # 11-18

    for ev_swimup, swimmer, home_band in swim_ups:
        try:
            _, ev_gender, ev_stroke_full = parse_event(ev_swimup)
        except Exception:
            continue
        stroke = ev_stroke_full.split("-")[-1] if "-" in ev_stroke_full else ev_stroke_full
        home_dist = _home_dist(home_band, stroke)
        target_ev = f"{home_band} {ev_gender} {home_dist}-{stroke}"
        target_stroke_key = f"{home_dist}-{stroke}"

        # Skip if swimmer doesn't have a time for the home-band event.
        sw_profile = your_profiles.get(swimmer, {}) or {}
        if target_stroke_key not in (sw_profile.get("strokes") or {}):
            moves_log.append({
                "swimmer": swimmer, "from": ev_swimup, "to": target_ev,
                "accepted": False, "reason": "no time at home band",
            })
            continue

        # Skip if target event isn't in our event list (shouldn't happen for
        # standard NVSL meets, but defensive).
        if target_ev not in polished:
            moves_log.append({
                "swimmer": swimmer, "from": ev_swimup, "to": target_ev,
                "accepted": False, "reason": "target event not in lineup",
            })
            continue

        # Build candidate: remove swimmer from source, add to target (drop
        # slowest if at capacity). No backfill at source — strategy_robust
        # already considered backfills, and the polish is intentionally minimal.
        cand = {k: dict(v) for k, v in polished.items()}
        cand[ev_swimup] = {**cand[ev_swimup],
                           "swimmers": [s for s in cand[ev_swimup].get("swimmers", [])
                                         if s != swimmer]}
        target_swimmers = list(cand.get(target_ev, {}).get("swimmers", []))
        if swimmer in target_swimmers:
            # Already there (shouldn't normally happen for a swim-up case)
            continue
        if len(target_swimmers) >= MAX_PER_EVENT:
            # Drop slowest by mean time
            def _t(n): return (your_profiles.get(n, {}) or {}).get(
                "strokes", {}).get(target_stroke_key, {}).get("mean", 1e9)
            target_swimmers.sort(key=_t)
            target_swimmers = target_swimmers[:MAX_PER_EVENT - 1]
        target_swimmers.append(swimmer)
        cand[target_ev] = {**cand.get(target_ev, {}), "swimmers": target_swimmers}

        # Score candidate.
        cand_stats = _score(cand)
        delta = cand_stats["margin"] - base_margin

        # Swim-up aversion (coaching preference): a coach won't race a FAST swimmer
        # up a band for nothing — if a swim-up must happen, a slower swimmer fills
        # the lane. So credit bringing a swimmer home by how good they are
        # (relief = swimup_aversion × max(0, −best_z)); a star comes home from a
        # point-neutral swim-up, an average swimmer still needs a real point gain.
        relief = 0.0
        if baselines and swimup_aversion:
            bz = _swimmer_best_z(sw_profile, home_band, ev_gender, baselines)
            if bz is not None:
                relief = swimup_aversion * max(0.0, -bz)

        if delta + relief >= accept_threshold:
            polished = cand
            base_margin = cand_stats["margin"]
            moves_log.append({
                "swimmer": swimmer, "from": ev_swimup, "to": target_ev,
                "delta": delta, "relief": round(relief, 2), "accepted": True,
            })
        else:
            moves_log.append({
                "swimmer": swimmer, "from": ev_swimup, "to": target_ev,
                "delta": delta, "relief": round(relief, 2), "accepted": False,
                "reason": f"delta {delta:+.2f} + aversion-relief {relief:+.2f} "
                          f"< threshold {accept_threshold}",
            })

    return polished, moves_log


def _polish_within_band_swaps(your_lineup, your_profiles, opp_profiles, events,
                              opp_lineup, relay_data,
                              accept_threshold=0.1, n_sim=1000,
                              prune_downgrade_secs=2.0,
                              style_weight=0.5,
                              baselines=None, fresh_swimmers=None,
                              specialty_weight=0.5, stale_boost=2.0):
    """Catch SAME-band CROSS-stroke swaps the optimizer missed AND prefer
    swimmers be placed in their REAL strokes (not league-avg imputation).

    OBJECTIVE — combined value function:
        value(swap) = score_delta + style_weight × real_delta
    where real_delta is the NET CHANGE in count of swimmers placed in events
    where their profile time came from actual race data (vs `league_avg_*`
    imputation). For style_weight=0.5, gaining one real-stroke deployment is
    worth ~0.5pt of expected score — enough to break ties / sub-noise gaps
    in favor of coaching-realistic placements, not enough to override real
    score improvements.

    SEARCH — best-improvement (NOT first-improvement) greedy:
    Each outer iteration enumerates ALL valid swap candidates, scores them
    all with CRN-locked MC, computes their combined value, and accepts the
    SINGLE HIGHEST-VALUE swap (if value > threshold). This fixes two bugs:
      1. Order independence — Bram↔Nico (value +0.55) correctly wins
         over Flynn↔Nico (value +0.10) regardless of iteration order.
      2. No oscillation — since the same metric is used both directions,
         once a swap is accepted its reverse has value=-original < threshold.

    COMMON RANDOM NUMBERS (CRN):
    Every per-event MC call uses a swimmer-indexed bank of pre-drawn N(0,1)
    samples (seed=42). Same swimmer → same simulated times across both
    lineups, so noise on the delta cancels almost completely. Effective
    SD on the combined value is near-zero, letting us reliably detect the
    +0.1pt real-signal swaps that pre-CRN was missing or flickering on.

    Relays are EXCLUDED from the delta calc — within-band individual-event
    swaps don't change relay leg picks (relay uses each swimmer's BEST
    stroke regardless of individual slot). So relay totals cancel exactly
    and we save the time.

    Tunable knobs:
      style_weight=0.5: pts of expected score traded per +1 real-stroke
                        deployment. 0 = pure score (no style preference);
                        1.0 = aggressive (will trade up to 1pt for real
                        deployment). Default 0.5 validated on SHB W1.
      n_sim=1000:       MC samples per per-event call.
      prune_downgrade_secs=2.0: skip swap attempts where the candidate's
                                time at destination is more than this many
                                seconds slower than the slowest currently-
                                picked swimmer in that event.

    Returns (polished_lineup, moves_log). Idempotent (running twice = no-op).
    Wall time on a typical 40-event meet: ~20-30s with best-improvement.
    """
    from copy import deepcopy
    from collections import defaultdict
    import numpy as np
    from Optimizer import MAX_EVENTS, MAX_PER_EVENT, is_eligible

    polished   = deepcopy(your_lineup)
    moves_log  = []

    # ── Build CRN bank ───────────────────────────────────────────────────────
    crn_rng = np.random.default_rng(seed=42)
    our_names_all = sorted(your_profiles.keys())
    opp_names_all = sorted(opp_profiles.keys())
    crn = {
        "our_bank": crn_rng.standard_normal((len(our_names_all), n_sim)),
        "opp_bank": crn_rng.standard_normal((len(opp_names_all), n_sim)),
        "our_idx":  {n: i for i, n in enumerate(our_names_all)},
        "opp_idx":  {n: i for i, n in enumerate(opp_names_all)},
    }

    def _score(lineup):
        """CRN-locked per-event MC. Returns mean of our individual-event
        points (relay excluded — constant across within-band swaps)."""
        total = np.zeros(n_sim)
        for ev in events:
            try:
                age, _, stroke = parse_event(ev)
            except Exception:
                continue
            our_combo = tuple(lineup.get(ev, {}).get("swimmers", []))
            opp_combo = tuple(opp_lineup.get(ev, {}).get("swimmers", []))
            our_pts, _ = _sim_event_both_pts(
                our_combo, opp_combo, stroke,
                your_profiles, opp_profiles, n_sim,
                age_group=age, crn=crn,
            )
            total += our_pts
        return float(total.mean())

    def _time_at(sw, stroke_key):
        return (your_profiles.get(sw, {}) or {}).get("strokes", {}).get(stroke_key, {}).get("mean")

    def _is_real(sw, ev):
        """True if swimmer's profile time for this event's stroke comes from
        actual race data (any source EXCEPT league_avg_* imputation).
        Fix #2's carry-forward times count as real."""
        try:
            _, _, stroke = parse_event(ev)
        except Exception:
            return False
        stats = (your_profiles.get(sw, {}) or {}).get("strokes", {}).get(stroke, {})
        src = stats.get("source", "") or ""
        return not src.startswith("league_avg")

    base = _score(polished)

    # Natural-stroke preference: reward placing swimmers in events where they're
    # most DOMINANT relative to the league field (lowest z = highest specialty),
    # and weight that preference UP for swimmers with no current-season data — a
    # stale freestyler belongs in free even when a marginal breast time looks fine,
    # because the breast number is a year-old guess. A point-NEUTRAL swap that puts
    # such a swimmer back in their event then clears the threshold; it never
    # overrides a real points gain.
    _fresh = fresh_swimmers if fresh_swimmers is not None else set()
    def _specialty_score(lineup):
        if not baselines:
            return 0.0
        total = 0.0
        for ev, info in lineup.items():
            try:
                _, _, stroke = parse_event(ev)
            except Exception:
                continue
            base_ev = baselines.get(ev)
            if not base_ev or base_ev.get("std", 0) <= 0:
                continue
            for sw in info.get("swimmers", []):
                m = (your_profiles.get(sw, {}) or {}).get("strokes", {}).get(stroke, {}).get("mean")
                if m is None:
                    continue
                z = (m - base_ev["mean"]) / base_ev["std"]
                boost = stale_boost if (fresh_swimmers is not None and sw not in _fresh) else 1.0
                total += (-z) * boost
        return total
    base_spec = _specialty_score(polished)

    bands = defaultdict(list)
    for ev in events:
        a, g, _ = parse_event(ev)
        bands[(a, g)].append(ev)

    def _violators(lineup):
        """Set of swimmers violating a constraint (>MAX_EVENTS events, or the same
        stroke twice). A swap is allowed only if it introduces NO NEW violator vs
        the base lineup — a PRE-EXISTING violation elsewhere (e.g. strategy_robust
        assigning a swimmer the same stroke twice in another band) must NOT block
        valid swaps in this band. Previously _constraints_ok scanned the whole
        lineup and returned False on ANY violation, so one bad assignment disabled
        the entire polish — leaving real point gains (Ryker → back) on the table."""
        ec = defaultdict(int)
        sc = defaultdict(set)
        bad = set()
        for ev, info in lineup.items():
            try:
                _, _, sf = parse_event(ev)
            except Exception:
                continue
            stk = sf.split("-")[-1]
            for sw in info.get("swimmers", []):
                ec[sw] += 1
                if ec[sw] > MAX_EVENTS:
                    bad.add(sw)
                if stk in sc[sw]:
                    bad.add(sw)
                sc[sw].add(stk)
        return bad

    def _eval(trial, real_delta):
        """Score a candidate trial; return (value, score_delta, new_score).
        Reads the current `base`/`base_spec` at call time. Value blends the MC
        point delta with the real-stroke and natural-stroke style preferences."""
        new = _score(trial)
        sd = new - base
        spec_d = (_specialty_score(trial) - base_spec) if baselines else 0.0
        return sd + style_weight * real_delta + specialty_weight * spec_d, sd, new

    # Run to convergence: the loop breaks as soon as no candidate beats the
    # threshold, so this cap is just a runaway guard. (Was 4 — too low for big
    # rosters: it stopped after the 4 best swaps and left smaller-but-real ones,
    # e.g. a +0.2 natural-stroke trade, on the table.)
    base_viol = _violators(polished)   # pre-existing violations don't block valid swaps
    for outer in range(20):
        # Best-improvement: enumerate ALL valid candidates, pick highest-value.
        best = None  # (value, score_delta, real_delta, trial, desc, new_score)
        for (age, gen), bes in bands.items():
            if len(bes) < 2: continue

            # ── Reciprocal pairwise swaps: s1 and s2 trade events ──────────────
            for i, ev1 in enumerate(bes):
                for ev2 in bes[i+1:]:
                    _, _, sf1 = parse_event(ev1)
                    _, _, sf2 = parse_event(ev2)
                    if sf1.split("-")[-1] == sf2.split("-")[-1]: continue
                    sw1 = list(polished.get(ev1, {}).get("swimmers", []))
                    sw2 = list(polished.get(ev2, {}).get("swimmers", []))
                    slowest_in_ev1 = max((_time_at(s, sf1) or 0) for s in sw1) if sw1 else 0
                    slowest_in_ev2 = max((_time_at(s, sf2) or 0) for s in sw2) if sw2 else 0

                    for s1 in sw1:
                        if s1 in sw2: continue
                        if not is_eligible(your_profiles.get(s1, {}), age, gen, sf2): continue
                        if prune_downgrade_secs is not None:
                            t = _time_at(s1, sf2)
                            if t is None or t > slowest_in_ev2 + prune_downgrade_secs:
                                continue
                        for s2 in sw2:
                            if s2 in sw1: continue
                            if not is_eligible(your_profiles.get(s2, {}), age, gen, sf1): continue
                            if prune_downgrade_secs is not None:
                                t = _time_at(s2, sf1)
                                if t is None or t > slowest_in_ev1 + prune_downgrade_secs:
                                    continue
                            trial = deepcopy(polished)
                            trial[ev1]["swimmers"] = [s2 if x == s1 else x for x in sw1]
                            trial[ev2]["swimmers"] = [s1 if x == s2 else x for x in sw2]
                            if _violators(trial) - base_viol: continue   # only block NEW violations
                            real_before = int(_is_real(s1, ev1)) + int(_is_real(s2, ev2))
                            real_after  = int(_is_real(s1, ev2)) + int(_is_real(s2, ev1))
                            real_delta  = real_after - real_before
                            value, score_delta, new = _eval(trial, real_delta)
                            if best is None or value > best[0]:
                                best = (value, score_delta, real_delta, trial,
                                        f"{s1} ({ev1}) ↔ {s2} ({ev2})", new)

            # ── Single-swimmer relocations: move s1 to a better event in its band,
            #    displacing that event's slowest if it's full. Catches moves the
            #    reciprocal swap can't — e.g. the team's fastest free swimmer stuck
            #    in fly because the free swimmer he'd bump can't swim fly. ─────────
            for ev1 in bes:
                _, _, sf1 = parse_event(ev1)
                sw1 = list(polished.get(ev1, {}).get("swimmers", []))
                for ev2 in bes:
                    if ev2 == ev1: continue
                    _, _, sf2 = parse_event(ev2)
                    if sf1.split("-")[-1] == sf2.split("-")[-1]: continue
                    sw2 = list(polished.get(ev2, {}).get("swimmers", []))
                    for s1 in sw1:
                        if s1 in sw2: continue
                        if not is_eligible(your_profiles.get(s1, {}), age, gen, sf2): continue
                        if _time_at(s1, sf2) is None: continue
                        new2 = sw2 + [s1]
                        dropped = None
                        if len(new2) > MAX_PER_EVENT:
                            dropped = max(new2, key=lambda s: _time_at(s, sf2) or 9999)
                            new2 = [x for x in new2 if x != dropped]
                        trial = deepcopy(polished)
                        trial[ev1]["swimmers"] = [x for x in sw1 if x != s1]
                        trial[ev2]["swimmers"] = new2
                        if _violators(trial) - base_viol: continue   # only block NEW violations
                        real_delta = int(_is_real(s1, ev2)) - int(_is_real(s1, ev1))
                        value, score_delta, new = _eval(trial, real_delta)
                        desc = (f"{s1} ({ev1} → {ev2}"
                                + (f", drop {dropped}" if dropped else "") + ")")
                        if best is None or value > best[0]:
                            best = (value, score_delta, real_delta, trial, desc, new)

        if best is None or best[0] <= accept_threshold:
            break
        value, score_d, real_d, trial, desc, new_score = best
        kind = "score" if real_d == 0 else f"style+{real_d}real"
        print(f"[within-polish] ACCEPT ({kind}): {desc}  "
              f"score Δ={score_d:+.2f}  value={value:+.2f}")
        moves_log.append({"desc": desc, "score_delta": score_d, "real_delta": real_d,
                          "value": value, "accepted": True})
        polished = trial
        base = new_score
        base_spec = _specialty_score(polished)
        base_viol = _violators(polished)   # refresh after the accepted move

    if not moves_log:
        print(f"[within-polish] no swaps accepted (threshold {accept_threshold})")
    else:
        print(f"[within-polish] applied {len(moves_log)} move(s)")
    return polished, moves_log


def simulate_match(our_lineup, opp_lineup, our_profiles, opp_profiles, events, relay_data, n=10000):
    """
    Simulate the full meet for both sides. Returns dict with per-trial totals.
    """
    our_indiv = simulate_our_total(our_lineup, opp_lineup, our_profiles, opp_profiles, events, n=n)
    opp_indiv = simulate_our_total(opp_lineup, our_lineup, opp_profiles, our_profiles, events, n=n)
    our_relay, opp_relay = simulate_relay_match(relay_data, n=n)
    return {
        "our_totals": our_indiv + our_relay,
        "opp_totals": opp_indiv + opp_relay,
    }


def match_stats(match):
    """Summary stats from simulate_match output."""
    import numpy as np
    our = match["our_totals"]
    opp = match["opp_totals"]
    return {
        "win_prob":   float((our > opp).mean()),
        "our_median": float(np.median(our)),
        "opp_median": float(np.median(opp)),
        "margin":     float(np.mean(our - opp)),
    }


# Issue #2 RESOLVED (2026-06-09) — W2-5 headline win prob, refit on PRODUCTION
# margins. The old form Φ(k·m/σ) (k=0.455, σ=64) was fit on race_points(pure_greedy)
# margins, not the simulate_match MC margin this function actually receives. Refit
# on the overnight cache (390 deduped W2-5 sides, current pipeline): a probit WITH
# an intercept — same family as W1 — beats every no-intercept candidate because the
# raw predicted margin carries the optimizer's ~+24 value-add inflation (winner's
# curse), which an origin-forced curve cannot absorb. 5-fold CV: Brier 0.164 vs
# 0.184 shipped, bias -0.006 vs +0.060; per-week bias within ±0.01 (no per-week
# constants needed). Note: margin 0 → ~39% is CORRECT, not a bug — an optimized
# lineup that only ties the predicted opponent is genuinely a slight underdog.
# IMPORTANT: fit on the RAW mstats["margin"]; if a value-add shrink is ever applied
# to the margin fed in here, these constants must be refit on the shrunken basis.
_CALIBRATION_CONSTANTS = None
def _load_calibration_constants():
    """calibration_constants.json — measured on the repaired production-basis
    backtest. Keys '1'..'5' -> {total_adj, blend_w, reanchor}; 'w1'/'w25' -> {b0,b1};
    '_strength_split' -> {enabled}."""
    global _CALIBRATION_CONSTANTS
    if _CALIBRATION_CONSTANTS is None:
        with open(os.path.join(BASE_DIR, "calibration_constants.json")) as f:
            _CALIBRATION_CONSTANTS = json.load(f)
    return _CALIBRATION_CONSTANTS


_THEORY2_CONSTANTS = None
def _load_theory2_constants():
    """calibration_constants_theory2.json — the coach-predictor-anchor calibration.
    {divmean, div_slope_b, reanchor:{week->pts}}. Full-league backtest, 5-fold CV
    stable; displayed = pred_coach + reanchor[week] - div_slope_b*(division-divmean)."""
    global _THEORY2_CONSTANTS
    if _THEORY2_CONSTANTS is None:
        with open(os.path.join(BASE_DIR, "calibration_constants_theory2.json")) as f:
            _THEORY2_CONSTANTS = json.load(f)
    return _THEORY2_CONSTANTS


_DIV_INDEX = None
def _team_division(team, year):
    """Team -> division number for `year` (nvsl_divisions_by_year.json). 0 if unknown
    (the div-term then no-ops, leaving the coach-anchor + reanchor)."""
    global _DIV_INDEX
    if _DIV_INDEX is None:
        _DIV_INDEX = {}
        try:
            dby = json.load(open(os.path.join(BASE_DIR, "nvsl_divisions_by_year.json")))
            for ys, divs in dby.items():
                if not isinstance(divs, dict):      # skip _source/_fetched/_note string keys
                    continue
                m = {}
                for d, ts in divs.items():
                    try:
                        di = int(d)
                    except Exception:
                        continue
                    if isinstance(ts, list):
                        for t in ts:
                            m[t] = di
                _DIV_INDEX[str(ys)] = m
        except Exception:
            pass
    return (_DIV_INDEX.get(str(year)) or {}).get(team, 0)


_TEAM_SCORE_INDEX = None
def _team_score_index():
    """{year:int -> {team -> {week:int -> final_score}}} from meet history,
    for the strength-split (demonstrated team strength)."""
    global _TEAM_SCORE_INDEX
    if _TEAM_SCORE_INDEX is None:
        idx = {}
        for ys, weeks in (_load_history() or {}).items():
            try: y = int(ys)
            except Exception: continue
            for wl, meets in (weeks or {}).items():
                try: wk = int(str(wl).split()[1])
                except Exception: continue
                for m in (meets or {}).values():
                    for side in ("team_a", "team_b"):
                        t = m.get(side) or {}
                        nm, sc = t.get("name"), t.get("score")
                        if nm and sc is not None:
                            idx.setdefault(y, {}).setdefault(nm, {})[wk] = float(sc)
        _TEAM_SCORE_INDEX = idx
    return _TEAM_SCORE_INDEX


def _team_strength(team, year, week):
    """Demonstrated strength = mean of this team's PRIOR-week actual scores this
    season (causal, no leakage); at W1, mean of last season's scores. Falls back to
    the ~210 league-typical half-meet score when a team has no history."""
    import numpy as np
    try:
        idx, y = _team_score_index(), int(year)
        if week and int(week) > 1:
            prior = [s for w, s in idx.get(y, {}).get(team, {}).items() if w < int(week)]
            if prior:
                return float(np.mean(prior))
        py = list(idx.get(y - 1, {}).get(team, {}).values())
        if py:
            return float(np.mean(py))
    except Exception:
        pass
    return 210.0


_TEAM_PART_INDEX = None
def _team_participation(team, year):
    """Prior-year mean participation = fielded individual swims / 120 per meet.
    Low participation => the optimizer fills more phantom lanes (no-show swimmers)
    => bigger value-add fantasy, concentrated in div 16-17 (CALIBRATION_STATE §0.009).
    Stable year-to-year (r=0.96). Falls back to the league-typical 0.89."""
    import numpy as np
    global _TEAM_PART_INDEX
    if _TEAM_PART_INDEX is None:
        idx = {}
        for ys, weeks in (_load_history() or {}).items():
            try: y = int(ys)
            except Exception: continue
            for meets in (weeks or {}).values():
                for m in (meets or {}).values():
                    for side in ("team_a", "team_b"):
                        t = m.get(side) or {}
                        nm = t.get("name")
                        if not nm:
                            continue
                        n = sum(len(d.get("swimmers") or [])
                                for d in (t.get("lineup") or {}).values())
                        idx.setdefault(y, {}).setdefault(nm, []).append(min(n, 120) / 120.0)
        _TEAM_PART_INDEX = idx
    try:
        rr = _TEAM_PART_INDEX.get(int(year) - 1, {}).get(team)
        if rr:
            return float(np.mean(rr))
    except Exception:
        pass
    return 0.89


WINP_W25_B0 = -0.2737
WINP_W25_B1 = 0.01122
def _w25_winprob(pred_margin):
    """W2-5 headline win prob: intercept probit on the simulate_match mean margin
    (our - opp). Replaces the greedy-margin shrinkage form (k=0.455/σ=64)."""
    import math
    z = WINP_W25_B0 + WINP_W25_B1 * float(pred_margin)
    return 0.5 * (1.0 + math.erf(z / math.sqrt(2.0)))


# W1-specific headline win prob. W1 margins are their own regime: mean predicted
# margin ~+49 while the actual win rate is ~50%, std ~62. The W2-5 form Φ(k·m/σ) has
# NO intercept (forces 50% at margin=0), so on W1 it's badly overconfident (+13.7%
# win-prob bias, Brier 0.258 — worse than a coin flip). This is a probit WITH an
# intercept (absorbs the additive margin bias), fit by MLE on the 98 W1 sides
# (w1_winprob_fit.py, 2026-06-07): ~0 bias, Brier 0.240. Deliberately timid — a +56
# predicted margin is a coin flip, +150 only ~65% — reflecting real W1 uncertainty
# with no current-season race data. The UI flags it as low-confidence.
WINP_W1_B0 = -0.2262
WINP_W1_B1 = 0.00404
def _w1_winprob(pred_margin):
    import math
    z = WINP_W1_B0 + WINP_W1_B1 * float(pred_margin)
    return 0.5 * (1.0 + math.erf(z / math.sqrt(2.0)))


# ── Robust optimizer (Phase 1: maximize E[win prob] across opp mixture) ────────

def _load_team_styles():
    """Load per-team consistency fingerprints from team_styles.json. Returns {} on miss."""
    path = os.path.join(BASE_DIR, "team_styles.json")
    if not os.path.exists(path):
        return None
    try:
        with open(path) as f:
            return json.load(f)
    except Exception:
        return None


def _greedy_counter_lineup(profiles, opp_entries, events):
    """Fast greedy counter (LP-free). Per event, picks top-3 eligible swimmers by
    mean time, respecting cross-event constraints (MAX_EVENTS, max-1-per-stroke).

    Used only for _build_opp_mixture scenarios — these are 'plausible opp lineups,'
    not the final answer, so greedy is fine. ~1s vs ~38s for the LP version.

    Event order matters for greedy; we process strongest events first (largest
    expected point swings vs opponent) so important constraints bind sensibly.
    """
    from Optimizer import (MAX_PER_EVENT, MAX_EVENTS, is_eligible, race_points)
    from collections import defaultdict
    lineup = {ev: {"swimmers": [], "expected_points": 0.0} for ev in events}
    events_used  = defaultdict(int)        # swimmer -> # of events assigned
    strokes_used = defaultdict(set)        # swimmer -> set of strokes assigned

    # Sort events by "stakes" — events where opp has stronger options first.
    # Use the best (lowest-mean) eligible swimmer as a proxy for opp strength.
    def _event_stakes(ev):
        age, gender, stroke = parse_event(ev)
        means = [p["strokes"][stroke]["mean"]
                 for p in profiles.values()
                 if is_eligible(p, age, gender, stroke)]
        return min(means) if means else 9999.0
    ordered_events = sorted(events, key=_event_stakes)

    for ev in ordered_events:
        age, gender, stroke = parse_event(ev)
        opp_means = [m for m, _ in opp_entries.get(ev, [])]
        eligible = sorted(
            [(name, p["strokes"][stroke]["mean"])
             for name, p in profiles.items()
             if is_eligible(p, age, gender, stroke)
                and events_used[name] < MAX_EVENTS
                and stroke not in strokes_used[name]],
            key=lambda x: x[1],
        )
        sel = [n for n, _ in eligible[:MAX_PER_EVENT]]
        if sel:
            means = [profiles[s]["strokes"][stroke]["mean"] for s in sel]
            pts = race_points(means, opp_means)
        else:
            pts = 0.0
        lineup[ev] = {"swimmers": sel, "expected_points": pts}
        for s in sel:
            events_used[s] += 1
            strokes_used[s].add(stroke)
    total = sum(d["expected_points"] for d in lineup.values())
    return lineup, total


def _build_opp_mixture(your_profiles, opp_profiles, events,
                       prior_opp_lineup=None, prior_our_lineup=None,
                       opp_team_name=None, use_opp_fingerprint=True,
                       year=None, week=None, has_our_current_data=True):
    """
    Probability-weighted set of plausible opp lineups for robust optimization.
    Returns list of (weight, opp_lineup, label, description).

    Base weights — recalibrated for v5 coach_predictor as the "best independent"
    scenario (which is now 53% Jaccard accurate vs. 38% for the old self_optimal):
       0% Reuse last week        (now redundant; v5 already encodes recency)
      70% Their predicted lineup (v5 coach_predictor — best single guess)
      15% Counter our last week  (some coaches do counter)
      15% Counter our self-opt   (sophisticated coaches anticipate)

    If use_opp_fingerprint=True and opp_team_name has a record in team_styles.json,
    weights shift based on opp's historical consistency (Jaccard):
      - HIGH consistency (z >= +1): they don't adapt → 0/85/8/7
      - LOW  consistency (z <= -1): they adapt heavily → 0/50/25/25
      - In between: linear interpolation

    The v5 predictor falls back to self_optimal when year/week are not provided
    or when the model is unavailable.

    W1 OVERRIDE (added 2026-05-30, upgraded to v4Rt 2026-06-02): when week==1,
    replace the v5 mixture with the validated tier-aware W1 architecture.

    v4Rt is tier-conditional:
      - Top-tier opps (D1-D6): use raw 2024 profile unchanged. Top-tier teams have
        large complete rosters in 2024 history (~10+ above-p50 swimmers per team);
        adding phantoms over-inflates them via LP redistribution.
      - Mid/bot-tier opps (D7-D17): add per-tier phantom rookies (v4R), calibrated
        from empirical mean_n_per_team and star_rate over 2022-2024 transitions.

    Cross-year validation:
      W1 2025: bias -0.8, MAE 31.4 (vs v3: bias -14, MAE 33)
      W1 2024 (held-out params): bias -0.06, MAE 38.2 (vs v3: bias -14)
    See improvements.md issue #1 and w1_predictor.py for full empirical basis.
    """
    if week == 1:
        try:
            import w1_predictor as W1
            if has_our_current_data:
                # ASYMMETRIC W1: we have current roster (SwimTopia/ladder uploaded
                # for our side), opp doesn't. Augment opp with v4Rt phantoms to
                # compensate for their stale 2024-only data.
                #
                # 4-stroke rotation mixture: 4 equal-weight components varying which
                # stroke each phantom is fast in. Removes the brittle deterministic
                # stroke assignment of plain v4Rt. For top-tier opps, returns a
                # single component (raw 2024 — no augmentation). Downstream
                # strategy_robust picks our lineup against the mixture, eliminating
                # over-defense of arbitrarily-cycled phantom strokes.
                #
                # Validated 2026-06-02: MAE -0.12 to -0.34 vs single-offset v4Rt
                # across both W1 2024 and W1 2025. Bot-tier bias tightens (+0.91
                # → -0.46 on 2025). See improvements.md issue #1.
                mixture = W1.predict_w1_lineup_mixture(
                    opp_profiles, opp_team_name, events, year=year,
                    picker="self_optimal", n_rotations=4,
                )
                return mixture
            else:
                # SYMMETRIC W1 FALLBACK: no upload on our side, so we're using
                # last year's data for us too. Use raw 2024 opp prediction (no
                # augmentation) to keep the comparison symmetric — both teams
                # are predicted as their 2024 final-state versions. Empirical
                # bias +1.5 to +1.9 on this fallback mode (vs -7 on the
                # asymmetric/broken version). Tell the UI so it can label the
                # prediction as "2024 baseline only — upload current roster
                # for accuracy."
                from Optimizer import self_optimal as _self_opt
                raw_lineup = _self_opt(opp_profiles, events)
                return [(1.0, raw_lineup,
                         "W1 raw 2024 (symmetric, no roster uploaded)",
                         "No current-season ladder/roster for your side, so the "
                         "comparison is kept symmetric: both teams predicted from "
                         "their 2024 final state. Upload SwimTopia roster + ladder "
                         "for full v4Rt accuracy (bias -0.8 vs +1.5 on this mode).")]
        except Exception as e:
            print(f"[w1_predictor] fallback to v5 mixture: {e}", flush=True)
            # fall through to existing logic

    # v5 prediction replaces the old self_optimal call for "their best independent"
    opp_self = _predict_opp_lineup_or_fallback(
        opp_team_name, year, week, opp_profiles, events
    )
    our_blind = self_optimal(your_profiles, events)
    our_blind_ents = lineup_to_entries(your_profiles, our_blind, events)
    opp_counter_blind, _ = _greedy_counter_lineup(opp_profiles, our_blind_ents, events)

    has_opp_prior = bool(prior_opp_lineup) and any(
        prior_opp_lineup.get(e, {}).get("swimmers") for e in events
    )
    has_our_prior = bool(prior_our_lineup) and any(
        prior_our_lineup.get(e, {}).get("swimmers") for e in events
    )

    opp_counter_lw = None
    if has_our_prior:
        our_lw_ents = lineup_to_entries(your_profiles, prior_our_lineup, events)
        opp_counter_lw, _ = _greedy_counter_lineup(opp_profiles, our_lw_ents, events)

    # Base weights for v5 predictor: 0/70/15/15 (reuse dropped — redundant with v5)
    w_reuse, w_selfopt, w_clw, w_cso = 0.0, 0.70, 0.15, 0.15

    if use_opp_fingerprint and opp_team_name:
        styles = _load_team_styles()
        if styles and styles.get("teams", {}).get(opp_team_name):
            team_data = styles["teams"][opp_team_name]
            opp_consistency = team_data.get("consistency_weighted")
            league_avg = styles.get("metadata", {}).get("league_avg", 0.46)
            league_stdev = styles.get("metadata", {}).get("league_stdev", 0.13)

            z = (opp_consistency - league_avg) / league_stdev if league_stdev > 0 else 0
            z = max(-1.0, min(1.0, z))

            # Endpoints recalibrated for v5 base:
            # HIGH (z=+1): 0/85/8/7   — predictable team; trust v5 even more
            # LOW  (z=-1): 0/50/25/25 — adaptive team; weight counter scenarios more
            if z >= 0:
                w_reuse   = 0.0
                w_selfopt = 0.70 + z * (0.85 - 0.70)
                w_clw     = 0.15 + z * (0.08 - 0.15)
                w_cso     = 0.15 + z * (0.07 - 0.15)
            else:
                w_reuse   = 0.0
                w_selfopt = 0.70 + (-z) * (0.50 - 0.70)
                w_clw     = 0.15 + (-z) * (0.25 - 0.15)
                w_cso     = 0.15 + (-z) * (0.25 - 0.15)

    raw = []
    if has_opp_prior and w_reuse > 0:
        raw.append((w_reuse, prior_opp_lineup, "Reuse last week",
                    "Opp reuses last week's lineup unchanged"))
    # v5 component: a single MAP lineup, or — when N_V5_SAMPLES>1 — several plausible
    # lineups sampled from the predictor's uncertainty, splitting the v5 weight across
    # them so the optimizer hedges over what the opponent might actually field (not
    # just the single best guess). Gated; default 1 = current behaviour.
    n_v5 = int(os.environ.get("N_V5_SAMPLES", "1"))
    # V5_SAMPLER selects HOW the extra v5 lineups are drawn:
    #   "strategic" (default when sampling) — redeploy the opponent's stars across the
    #       strokes they've raced (the dimension our response is actually sensitive to;
    #       the variation-probe found prob-jitter is structurally blind to it).
    #   "jitter" — the original log-normal probability perturbation (kept for A/B).
    v5_sampler = os.environ.get("V5_SAMPLER", "strategic").lower()
    v5_samples = None
    if n_v5 > 1:
        try:
            import coach_predictor as _cp
            if v5_sampler == "jitter":
                v5_samples = _cp.sample_opp_lineups(opp_team_name, year, week,
                                                    opp_profiles, events, k=n_v5)
            else:
                v5_samples = _cp.sample_opp_lineups_strategic(opp_team_name, year, week,
                                                              opp_profiles, events, k=n_v5)
        except Exception as _e:
            print(f"[mixture] v5 sampling failed ({_e}); single v5", flush=True)
            v5_samples = None
    if v5_samples and len(v5_samples) > 1:
        wv = w_selfopt / len(v5_samples)
        for i, lu in enumerate(v5_samples):
            raw.append((wv, lu, f"Their predicted lineup (v5 sample {i+1}/{len(v5_samples)})",
                        "One plausible opp lineup sampled from the coach-predictor's uncertainty"))
    else:
        raw.append((w_selfopt, opp_self, "Their predicted lineup",
                    "Opp plays their most likely lineup (trained coach predictor)"))
    if has_our_prior:
        raw.append((w_clw, opp_counter_lw, "Counter our last week",
                    "Opp reads our last-week results and counters that specific lineup"))
    raw.append((w_cso, opp_counter_blind, "Counter our self-optimal play",
                "Opp predicts we'll play self-optimal (best independent) and counters"))

    # Normalize weights (some scenarios may be missing for early weeks)
    total_w = sum(w for w, *_ in raw)
    return [(w / total_w, l, lab, desc) for w, l, lab, desc in raw]



def _sim_event_both_pts(our_combo, opp_combo, stroke, our_profiles, opp_profiles, n,
                         age_group=None, crn=None):
    """
    Joint per-trial sim for one event. Returns (our_pts, opp_pts) arrays of shape (n,).
    Both teams' swimmers race in the same simulation so the points are paired.
    age_group: if provided, uses log-normal sampling for 8U + 9-10 25-fly events.

    crn: optional dict {our_bank, our_idx, opp_bank, opp_idx} for Common Random
    Numbers. our_bank is (n_total_our_swimmers, n) of N(0,1); our_idx maps
    swimmer name -> row index. Same idea for opp. With CRN, every paired
    lineup comparison uses identical noise draws for each swimmer, collapsing
    the noise on the *difference* between two lineups.
    """
    import numpy as np
    from Optimizer import _pts_map, _sample_times, _is_lognormal_event
    our_names = [
        s for s in our_combo
        if s in our_profiles and stroke in our_profiles[s].get("strokes", {})
    ]
    opp_names = [
        s for s in opp_combo
        if s in opp_profiles and stroke in opp_profiles[s].get("strokes", {})
    ]
    our_ents = [(our_profiles[s]["strokes"][stroke]["mean"],
                 our_profiles[s]["strokes"][stroke]["std"]) for s in our_names]
    opp_ents = [(opp_profiles[s]["strokes"][stroke]["mean"],
                 opp_profiles[s]["strokes"][stroke]["std"]) for s in opp_names]
    n_yours = len(our_ents)
    n_opp   = len(opp_ents)
    total   = n_yours + n_opp
    if total == 0:
        return np.zeros(n), np.zeros(n)
    all_means = np.array([m for m, _ in our_ents] + [m for m, _ in opp_ents])
    all_stds  = np.array([s for _, s in our_ents] + [s for _, s in opp_ents])
    use_log   = (age_group is not None) and _is_lognormal_event(age_group, stroke)

    crn_norms = None
    if crn is not None:
        # Stack N(0,1) rows for each swimmer in the race, in order
        rows = ([crn["our_bank"][crn["our_idx"][s], :n] for s in our_names]
              + [crn["opp_bank"][crn["opp_idx"][s], :n] for s in opp_names])
        crn_norms = np.column_stack(rows) if rows else None

    times    = _sample_times(all_means, all_stds, n, use_lognormal=use_log,
                              crn_norms=crn_norms)
    ranks    = np.argsort(np.argsort(times, axis=1), axis=1)
    pts_grid = _pts_map(total)[ranks]
    our_pts  = pts_grid[:, :n_yours].sum(axis=1) if n_yours else np.zeros(n)
    opp_pts  = pts_grid[:, n_yours:].sum(axis=1) if n_opp   else np.zeros(n)
    return our_pts, opp_pts


def strategy_robust(your_profiles, opp_profiles, events, relay_data,
                    prior_opp_lineup=None, prior_our_lineup=None,
                    opp_team_name=None, use_opp_fingerprint=True,
                    n_sim=3000, max_iter=2, progress_cb=None,
                    use_margin_tiebreaker=True,
                    year=None, week=None, has_our_current_data=True):
    """
    Local search to maximize E[win prob] across the opp mixture.
    Mixture entries are all static (plausible human-coach lineups); we cache
    per-event-per-scenario point arrays so one event swap only re-simulates
    that single event.

    year/week pass through to _build_opp_mixture so the coach_predictor can be
    used (falls back to self_optimal when year/week are missing).

    Returns (lineup, e_winprob, mixture, breakdown).
    """
    import numpy as np
    from copy import deepcopy
    from itertools import combinations as iter_combos
    from Optimizer import MAX_PER_EVENT, MAX_EVENTS, is_eligible, _feasible_combo
    MC_TOP_N = 6

    mixture = _build_opp_mixture(your_profiles, opp_profiles, events,
                                 prior_opp_lineup=prior_opp_lineup,
                                 prior_our_lineup=prior_our_lineup,
                                 opp_team_name=opp_team_name,
                                 use_opp_fingerprint=use_opp_fingerprint,
                                 year=year, week=week,
                                 has_our_current_data=has_our_current_data)
    mixture_lineups = [opp_l for _, opp_l, _, _ in mixture]
    weights         = np.array([w for w, _, _, _ in mixture])
    K               = len(mixture)

    # ── Common Random Numbers (CRN) ─────────────────────────────────────────
    # Pre-draw a fixed bank of N(0,1) samples indexed by swimmer name. Every
    # call to _sim_event_both_pts uses the same draws for the same swimmer,
    # so paired comparisons of two lineups have ~zero noise on the difference.
    # Critical for resolving small margin differences (~0.1-0.5 pts) that get
    # buried in fresh-MC noise.
    crn_rng = np.random.default_rng(seed=42)
    our_names_all = sorted(your_profiles.keys())
    opp_names_all = sorted(opp_profiles.keys())
    crn = {
        "our_bank": crn_rng.standard_normal((len(our_names_all), n_sim)),
        "opp_bank": crn_rng.standard_normal((len(opp_names_all), n_sim)),
        "our_idx":  {n: i for i, n in enumerate(our_names_all)},
        "opp_idx":  {n: i for i, n in enumerate(opp_names_all)},
    }

    # ── Per-event simulation cache (memoization) ────────────────────────────
    # With CRN locked, _sim_event_both_pts is deterministic: same combos +
    # stroke + age_group → same per-trial arrays, every time. The optimizer
    # asks the same per-event question over and over during local search
    # (~97% repeat rate measured), so caching gives ~10× speedup. The cache
    # is local to this run; freed when strategy_robust returns.
    _sim_cache = {}
    def _sim_cached(our_combo, opp_combo, stroke, age_group):
        key = (tuple(sorted(our_combo)), tuple(sorted(opp_combo)), stroke, age_group)
        cached = _sim_cache.get(key)
        if cached is not None:
            return cached
        result = _sim_event_both_pts(our_combo, opp_combo, stroke,
                                      your_profiles, opp_profiles, n_sim,
                                      age_group=age_group, crn=crn)
        _sim_cache[key] = result
        return result

    # Relay trials per scenario — relay doesn't depend on individual lineup choice
    relay_our, relay_opp = simulate_relay_match(relay_data, n=n_sim)

    # Starting lineup: self-optimal (truly blind baseline)
    lineup = deepcopy(self_optimal(your_profiles, events))

    # Build initial per-event-per-scenario point caches and per-scenario totals
    our_pts_cache = {}
    opp_pts_cache = {}
    our_totals    = [np.zeros(n_sim) for _ in range(K)]
    opp_totals    = [np.zeros(n_sim) for _ in range(K)]

    for event in events:
        age_e, _, stroke = parse_event(event)
        our_combo    = tuple(lineup[event]["swimmers"])
        our_arr_list = []
        opp_arr_list = []
        for k, opp_l in enumerate(mixture_lineups):
            opp_combo = tuple(opp_l.get(event, {}).get("swimmers", []))
            our_e, opp_e = _sim_cached(our_combo, opp_combo, stroke, age_e)
            our_arr_list.append(our_e)
            opp_arr_list.append(opp_e)
            our_totals[k] += our_e
            opp_totals[k] += opp_e
        our_pts_cache[event] = our_arr_list
        opp_pts_cache[event] = opp_arr_list

    for k in range(K):
        our_totals[k] += relay_our
        opp_totals[k] += relay_opp

    win_prob_per_k = np.array([float((our_totals[k] > opp_totals[k]).mean()) for k in range(K)])
    margin_per_k   = np.array([float((our_totals[k] - opp_totals[k]).mean()) for k in range(K)])
    best_ewp = float(np.sum(weights * win_prob_per_k))
    best_mgn = float(np.sum(weights * margin_per_k))

    # Margin tiebreaker: when two lineups are essentially tied on win prob,
    # break the tie by preferring the one with higher expected margin. Critical
    # for lopsided matchups where win prob saturates at ~100%.
    EWP_TIE_EPS = 0.005   # 0.5pp = consider tied
    MGN_GAIN_EPS = 0.05   # need >0.05 pts to flip the tie

    def _accept(cand_ewp, cand_mgn, ref_ewp, ref_mgn):
        if cand_ewp > ref_ewp + EWP_TIE_EPS:
            return True
        if use_margin_tiebreaker and abs(cand_ewp - ref_ewp) <= EWP_TIE_EPS \
                and cand_mgn > ref_mgn + MGN_GAIN_EPS:
            return True
        # Standard strict-improvement fallback (when tiebreaker is off)
        if not use_margin_tiebreaker and cand_ewp > ref_ewp + 1e-9:
            return True
        return False

    if progress_cb:
        progress_cb(f"Robust start E[win]={best_ewp*100:.1f}% E[mgn]={best_mgn:+.1f}")

    for iteration in range(max_iter):
        improved = False
        for event in events:
            _, _, stroke = parse_event(event)
            age    = event.split()[0]
            gender = event.split()[1]
            eligible = [
                n for n, _ in sorted(
                    [(name, p["strokes"][stroke]["mean"])
                     for name, p in your_profiles.items()
                     if is_eligible(p, age, gender, stroke)],
                    key=lambda x: x[1],
                )[:MC_TOP_N]
            ]

            current_combo = tuple(lineup[event]["swimmers"])

            # Per-scenario totals EXCLUDING this event's current contribution
            excl_our = [our_totals[k] - our_pts_cache[event][k] for k in range(K)]
            excl_opp = [opp_totals[k] - opp_pts_cache[event][k] for k in range(K)]

            local_best_combo   = current_combo
            local_best_our_arr = our_pts_cache[event]
            local_best_opp_arr = opp_pts_cache[event]
            local_best_ewp     = best_ewp
            local_best_mgn     = best_mgn

            for r in range(0, MAX_PER_EVENT + 1):
                for combo in iter_combos(eligible, r):
                    if combo == current_combo:
                        continue
                    if not _feasible_combo(combo, event, lineup, events):
                        continue

                    # Re-sim ONLY this event, for each scenario
                    cand_our_arrs = []
                    cand_opp_arrs = []
                    cand_ewp = 0.0
                    cand_mgn = 0.0
                    for k, opp_l in enumerate(mixture_lineups):
                        opp_combo = tuple(opp_l.get(event, {}).get("swimmers", []))
                        our_e, opp_e = _sim_cached(combo, opp_combo, stroke, age)
                        cand_our_arrs.append(our_e)
                        cand_opp_arrs.append(opp_e)
                        new_our_total_k = excl_our[k] + our_e
                        new_opp_total_k = excl_opp[k] + opp_e
                        cand_ewp += weights[k] * float((new_our_total_k > new_opp_total_k).mean())
                        cand_mgn += weights[k] * float((new_our_total_k - new_opp_total_k).mean())

                    if _accept(cand_ewp, cand_mgn, local_best_ewp, local_best_mgn):
                        local_best_ewp     = cand_ewp
                        local_best_mgn     = cand_mgn
                        local_best_combo   = combo
                        local_best_our_arr = cand_our_arrs
                        local_best_opp_arr = cand_opp_arrs

            if local_best_combo != current_combo:
                lineup[event]["swimmers"] = list(local_best_combo)
                our_pts_cache[event]      = local_best_our_arr
                opp_pts_cache[event]      = local_best_opp_arr
                for k in range(K):
                    our_totals[k] = excl_our[k] + local_best_our_arr[k]
                    opp_totals[k] = excl_opp[k] + local_best_opp_arr[k]
                best_ewp = local_best_ewp
                best_mgn = local_best_mgn
                improved = True

        if progress_cb:
            progress_cb(f"Robust iter {iteration+1} E[win]={best_ewp*100:.1f}%")
        if not improved:
            break

    # ── 2-event swap pass ────────────────────────────────────────────────────
    # The 1-event sweep can't see coordinated moves like "Luke breast→free,
    # Liam backfills breast" because each step alone looks neutral-or-worse.
    # This pass proposes (swimmer S, source event A, destination event B) moves
    # with backfill of A, and re-evaluates A and B jointly. Best-improvement
    # within each sweep; repeats until no swap improves the objective.

    def _eval_two_events(event_A, combo_A, event_B, combo_B):
        """Re-simulate A and B. Returns (cand_ewp, cand_mgn, A_our, A_opp, B_our, B_opp)."""
        age_A, _, stroke_A = parse_event(event_A)
        age_B, _, stroke_B = parse_event(event_B)
        a_our_arrs, a_opp_arrs = [], []
        b_our_arrs, b_opp_arrs = [], []
        cand_ewp = 0.0
        cand_mgn = 0.0
        for k, opp_l in enumerate(mixture_lineups):
            opp_A = tuple(opp_l.get(event_A, {}).get("swimmers", []))
            opp_B = tuple(opp_l.get(event_B, {}).get("swimmers", []))
            a_our, a_opp = _sim_cached(combo_A, opp_A, stroke_A, age_A)
            b_our, b_opp = _sim_cached(combo_B, opp_B, stroke_B, age_B)
            a_our_arrs.append(a_our); a_opp_arrs.append(a_opp)
            b_our_arrs.append(b_our); b_opp_arrs.append(b_opp)
            new_our_total = (our_totals[k]
                             - our_pts_cache[event_A][k] - our_pts_cache[event_B][k]
                             + a_our + b_our)
            new_opp_total = (opp_totals[k]
                             - opp_pts_cache[event_A][k] - opp_pts_cache[event_B][k]
                             + a_opp + b_opp)
            cand_ewp += weights[k] * float((new_our_total > new_opp_total).mean())
            cand_mgn += weights[k] * float((new_our_total - new_opp_total).mean())
        return cand_ewp, cand_mgn, a_our_arrs, a_opp_arrs, b_our_arrs, b_opp_arrs

    # Pre-parse all events once (parse_event is regex-heavy; was hot under cProfile)
    _event_stroke = {ev: parse_event(ev)[2] for ev in events}

    def _swap_creates_dup_stroke(new_A_combo, event_A, new_B_combo, event_B):
        """True iff the proposed (A, B) swap would put any swimmer in two events
        with the same stroke. Bug fix — the 1-event sweep handles this via
        _feasible_combo, but swap passes need the same check across two events.

        Only checks swimmers actually involved in the swap (touched_set) — they
        are the only ones whose stroke-set could now collide. Cuts work from
        O(all_events × all_swimmers) per call to O(touched_swimmers × events)."""
        stroke_A = _event_stroke[event_A]
        stroke_B = _event_stroke[event_B]
        # Swimmers whose assignments differ between old and new
        touched = set(new_A_combo) | set(new_B_combo) \
                  | set(lineup[event_A]["swimmers"]) | set(lineup[event_B]["swimmers"])
        for s in touched:
            seen = set()
            # Walk through lineup, substituting the proposed combos for A and B
            for ev_label in events:
                stroke = _event_stroke[ev_label]
                if ev_label == event_A:
                    in_event = s in new_A_combo
                elif ev_label == event_B:
                    in_event = s in new_B_combo
                else:
                    in_event = s in lineup[ev_label]["swimmers"]
                if in_event:
                    if stroke in seen:
                        return True
                    seen.add(stroke)
        return False

    def _swap_sweep():
        nonlocal best_ewp, best_mgn
        # Best-improvement: enumerate every (S, A, B, backfill) tuple, pick best.
        best_swap = None
        best_swap_ewp = best_ewp
        best_swap_mgn = best_mgn

        # Snapshot current event counts (for backfill capacity checks)
        ev_counts = {}
        for ev in events:
            for s in lineup[ev]["swimmers"]:
                ev_counts[s] = ev_counts.get(s, 0) + 1

        # All (swimmer, source_event) pairs
        swimmer_events = [(s, ev) for ev in events for s in lineup[ev]["swimmers"]]

        for S, event_A in swimmer_events:
            age_A, gender_A, stroke_A = parse_event(event_A)
            S_prof = your_profiles.get(S, {})

            for event_B in events:
                if event_B == event_A:
                    continue
                if S in lineup[event_B]["swimmers"]:
                    continue  # already there

                age_B, gender_B, stroke_B = parse_event(event_B)
                if not is_eligible(S_prof, age_B, gender_B, stroke_B):
                    continue

                current_A = list(lineup[event_A]["swimmers"])
                current_B = list(lineup[event_B]["swimmers"])
                new_A_base = [x for x in current_A if x != S]

                # Build new B: add S, trim slowest if over capacity (drop slowest non-S)
                if len(current_B) + 1 > MAX_PER_EVENT:
                    sorted_B = sorted(
                        current_B,
                        key=lambda n: your_profiles.get(n, {}).get("strokes", {})
                                                .get(stroke_B, {}).get("mean", 1e9),
                    )
                    kept = sorted_B[:MAX_PER_EVENT - 1]  # keep fastest (cap-1)
                    new_B_combo = kept + [S]
                    dropped_from_B = [n for n in current_B if n not in kept]
                else:
                    new_B_combo = current_B + [S]
                    dropped_from_B = []

                # Backfill candidates for A: top-3 eligible (by mean) who have
                # capacity AFTER the swap. Effective capacity uses post-swap counts.
                def _post_swap_count(name):
                    c = ev_counts.get(name, 0)
                    if name in dropped_from_B:
                        c -= 1
                    return c

                backfill_pool = []
                for name, p in your_profiles.items():
                    if name == S: continue
                    if name in new_A_base: continue
                    if name in new_B_combo: continue
                    if not is_eligible(p, age_A, gender_A, stroke_A): continue
                    if _post_swap_count(name) >= MAX_EVENTS:
                        continue
                    mean = p.get("strokes", {}).get(stroke_A, {}).get("mean", 1e9)
                    if mean >= 1e8: continue
                    backfill_pool.append((mean, name))
                backfill_pool.sort()
                top_backfills = [name for _, name in backfill_pool[:3]]

                # Try: no backfill, or each of top-3 backfills
                for bf in [None] + top_backfills:
                    new_A_combo = list(new_A_base)
                    if bf is not None:
                        new_A_combo.append(bf)
                    if len(new_A_combo) > MAX_PER_EVENT:
                        continue
                    # Identical-to-current short-circuit (no actual change)
                    if (set(new_A_combo) == set(current_A)
                            and set(new_B_combo) == set(current_B)):
                        continue
                    # Bug fix: reject swaps that would put a swimmer in 2 events with the same stroke
                    if _swap_creates_dup_stroke(new_A_combo, event_A, new_B_combo, event_B):
                        continue
                    cand_ewp, cand_mgn, a_our, a_opp, b_our, b_opp = _eval_two_events(
                        event_A, tuple(new_A_combo),
                        event_B, tuple(new_B_combo),
                    )
                    if _accept(cand_ewp, cand_mgn, best_swap_ewp, best_swap_mgn):
                        best_swap_ewp = cand_ewp
                        best_swap_mgn = cand_mgn
                        best_swap = (S, event_A, event_B, new_A_combo, new_B_combo,
                                     a_our, a_opp, b_our, b_opp)

        if best_swap is None:
            return False

        S, event_A, event_B, new_A_combo, new_B_combo, \
            a_our, a_opp, b_our, b_opp = best_swap

        # Apply: update lineup, caches, totals
        for k in range(K):
            our_totals[k] = (our_totals[k]
                             - our_pts_cache[event_A][k] - our_pts_cache[event_B][k]
                             + a_our[k] + b_our[k])
            opp_totals[k] = (opp_totals[k]
                             - opp_pts_cache[event_A][k] - opp_pts_cache[event_B][k]
                             + a_opp[k] + b_opp[k])
        lineup[event_A]["swimmers"] = list(new_A_combo)
        lineup[event_B]["swimmers"] = list(new_B_combo)
        our_pts_cache[event_A] = a_our
        opp_pts_cache[event_A] = a_opp
        our_pts_cache[event_B] = b_our
        opp_pts_cache[event_B] = b_opp
        best_ewp = best_swap_ewp
        best_mgn = best_swap_mgn
        if progress_cb:
            progress_cb(f"Swap pass: {S} {event_A}→{event_B}, "
                        f"E[win]={best_ewp*100:.1f}% E[mgn]={best_mgn:+.1f}")
        return True

    # ── Pairwise swap pass ──────────────────────────────────────────────────
    # Atomic "S1 ↔ S2 trade between events A ↔ B": catches cases like
    # "Luke goes to free while Alton takes breast". No backfill needed (the
    # swap is self-balancing — each event keeps the same swimmer count, each
    # swimmer keeps the same event count).
    def _pairwise_swap_sweep():
        nonlocal best_ewp, best_mgn
        best_swap = None
        best_swap_ewp = best_ewp
        best_swap_mgn = best_mgn

        swimmer_events = [(s, ev) for ev in events for s in lineup[ev]["swimmers"]]
        n_entries = len(swimmer_events)

        for i in range(n_entries):
            S1, event_A = swimmer_events[i]
            age_A, gender_A, stroke_A = parse_event(event_A)
            S1_prof = your_profiles.get(S1, {})
            for j in range(i + 1, n_entries):
                S2, event_B = swimmer_events[j]
                if event_A == event_B:    continue  # same event — no swap
                if S1 == S2:               continue
                if S1 in lineup[event_B]["swimmers"]: continue  # S1 already in B
                if S2 in lineup[event_A]["swimmers"]: continue  # S2 already in A

                age_B, gender_B, stroke_B = parse_event(event_B)
                S2_prof = your_profiles.get(S2, {})

                # Both must be eligible for the OTHER event
                if not is_eligible(S1_prof, age_B, gender_B, stroke_B): continue
                if not is_eligible(S2_prof, age_A, gender_A, stroke_A): continue

                new_A_combo = [x for x in lineup[event_A]["swimmers"] if x != S1] + [S2]
                new_B_combo = [x for x in lineup[event_B]["swimmers"] if x != S2] + [S1]

                # Bug fix: reject swaps that would put a swimmer in 2 events with the same stroke
                if _swap_creates_dup_stroke(new_A_combo, event_A, new_B_combo, event_B):
                    continue

                cand_ewp, cand_mgn, a_our, a_opp, b_our, b_opp = _eval_two_events(
                    event_A, tuple(new_A_combo),
                    event_B, tuple(new_B_combo),
                )
                if _accept(cand_ewp, cand_mgn, best_swap_ewp, best_swap_mgn):
                    best_swap_ewp = cand_ewp
                    best_swap_mgn = cand_mgn
                    best_swap = (S1, S2, event_A, event_B, new_A_combo, new_B_combo,
                                 a_our, a_opp, b_our, b_opp)

        if best_swap is None:
            return False

        S1, S2, event_A, event_B, new_A_combo, new_B_combo, \
            a_our, a_opp, b_our, b_opp = best_swap

        for k in range(K):
            our_totals[k] = (our_totals[k]
                             - our_pts_cache[event_A][k] - our_pts_cache[event_B][k]
                             + a_our[k] + b_our[k])
            opp_totals[k] = (opp_totals[k]
                             - opp_pts_cache[event_A][k] - opp_pts_cache[event_B][k]
                             + a_opp[k] + b_opp[k])
        lineup[event_A]["swimmers"] = list(new_A_combo)
        lineup[event_B]["swimmers"] = list(new_B_combo)
        our_pts_cache[event_A] = a_our
        opp_pts_cache[event_A] = a_opp
        our_pts_cache[event_B] = b_our
        opp_pts_cache[event_B] = b_opp
        best_ewp = best_swap_ewp
        best_mgn = best_swap_mgn
        if progress_cb:
            progress_cb(f"Pair swap: {S1}↔{S2} between {event_A}↔{event_B}, "
                        f"E[win]={best_ewp*100:.1f}% E[mgn]={best_mgn:+.1f}")
        return True

    # Interleave the two swap types: any improvement in one may unlock the other.
    # Loop until both sweep types report no improvement (true local optimum
    # under the move set). Cap at 50 as a safety net — convergence is normally
    # in under 20 iterations.
    for swap_iter in range(50):
        improved_this_iter = False
        if _swap_sweep():
            improved_this_iter = True
        if _pairwise_swap_sweep():
            improved_this_iter = True
        if not improved_this_iter:
            break

    if progress_cb:
        progress_cb(f"After swap passes: E[win]={best_ewp*100:.1f}%")

    # ── Fill-empty-slots post-process ──────────────────────────────────────
    # Coaching rule: every event should have MAX_PER_EVENT swimmers if
    # eligible swimmers are available (even if they won't score). Leaving a
    # slot empty is a forfeit — strictly worse than filling it with anyone.
    # The main optimization sometimes leaves slots open due to MC noise or
    # marginal-swimmer evaluation; this pass corrects that without disturbing
    # the optimizer's main decisions.
    n_filled = 0
    for event in events:
        cur_swimmers = list(lineup[event]["swimmers"])
        n_open = MAX_PER_EVENT - len(cur_swimmers)
        if n_open <= 0: continue

        age_e, gender_e, stroke_e = parse_event(event)

        # Current event count + assigned stroke-names per swimmer (whole lineup)
        ev_counts = {}
        swimmer_strokes = {}   # name -> set of stroke names already assigned
        for ev in events:
            try:
                stk_name = parse_event(ev)[2].split("-")[-1]
            except Exception:
                stk_name = None
            for s in lineup[ev]["swimmers"]:
                ev_counts[s] = ev_counts.get(s, 0) + 1
                if stk_name:
                    swimmer_strokes.setdefault(s, set()).add(stk_name)
        stroke_name_e = stroke_e.split("-")[-1]

        # Eligible + available swimmers: not already in this event, under the
        # 2-event cap, AND not already swimming this stroke in another band. The
        # last check was missing — the fill pass could add a swimmer (via swim-up
        # eligibility) to an event whose stroke they already swim at home, making an
        # illegal lineup (e.g. Hillary Braddock in 13-14 AND 15-18 50-back). The
        # four optimization passes all guard same-stroke-twice; only this top-up didn't.
        candidates = []
        for name, p in your_profiles.items():
            if name in cur_swimmers: continue
            if ev_counts.get(name, 0) >= MAX_EVENTS: continue
            if stroke_name_e in swimmer_strokes.get(name, set()): continue
            if not is_eligible(p, age_e, gender_e, stroke_e): continue
            mean = p["strokes"][stroke_e]["mean"]
            if mean is None or mean >= 1e8: continue
            candidates.append((mean, name))
        candidates.sort()   # fastest first

        # Add up to n_open swimmers
        to_add = [name for _, name in candidates[:n_open]]
        if not to_add: continue
        new_combo = cur_swimmers + to_add
        lineup[event]["swimmers"] = new_combo
        n_filled += len(to_add)

        # Recompute this event's contribution to per-scenario totals (cache update)
        new_arrs_our, new_arrs_opp = [], []
        for k, opp_l in enumerate(mixture_lineups):
            opp_combo = tuple(opp_l.get(event, {}).get("swimmers", []))
            our_e, opp_e = _sim_cached(tuple(new_combo), opp_combo, stroke_e, age_e)
            new_arrs_our.append(our_e); new_arrs_opp.append(opp_e)
            our_totals[k] = our_totals[k] - our_pts_cache[event][k] + our_e
            opp_totals[k] = opp_totals[k] - opp_pts_cache[event][k] + opp_e
        our_pts_cache[event] = new_arrs_our
        opp_pts_cache[event] = new_arrs_opp

    if n_filled and progress_cb:
        progress_cb(f"Fill-empty-slots: added {n_filled} swimmers to under-filled events")

    # Recompute best_ewp/best_mgn after fills (in case the totals shifted)
    if n_filled:
        win_prob_per_k = np.array([float((our_totals[k] > opp_totals[k]).mean()) for k in range(K)])
        margin_per_k   = np.array([float((our_totals[k] - opp_totals[k]).mean()) for k in range(K)])
        best_ewp = float(np.sum(weights * win_prob_per_k))
        best_mgn = float(np.sum(weights * margin_per_k))

    # Final breakdown from cached totals
    breakdown = []
    for k in range(K):
        w, opp_l, label, desc = mixture[k]
        our_t = our_totals[k]
        opp_t = opp_totals[k]
        breakdown.append({
            "label":      label, "desc": desc, "weight": w,
            "win_prob":   float((our_t > opp_t).mean()),
            "margin":     float((our_t - opp_t).mean()),
            "our_median": float(np.median(our_t)),
            "opp_median": float(np.median(opp_t)),
        })

    return lineup, best_ewp, mixture, breakdown


# ── W1 presence-MC (expected-absence) score model ───────────────────────────
# At W1 the roster is the 2024 profile, but only ~p of those swimmers actually
# swim a given meet: P(returns)*P(present) ~ 0.76*0.70 ~ 0.60 (graduations +
# meet-day absences). Fielding all of them over-predicts (+29 on the full W1 set).
# Presence-MC averages the score over attendance draws; an absent swimmer's lane
# is taken by the next eligible bench swimmer (what a coach does), so coverage
# stays complete (avoids the under-fielding artifact of hard present-restriction).
# Calibrated/validated on the full 98-side W1 set (w1_overnight.py, 2026-06-07):
#   baseline bias +29.1 / MAE 38.1 / 4 blowouts  ->  p=0.60: +2.1 / 30.4 / 3.
# Stable at NDRAW=60 / MC_N=4000. Applied ONLY at W1 with no current-season upload
# (a ladder upload already strips ghosts, so p=0.60 would over-deflate that case).
W1_PRESENCE_P     = 0.60
W1_PRESENCE_NDRAW = 60
W1_PRESENCE_MC_N  = 4000
W1_PRESENCE_MAX_EVENTS = 3   # loose cap so one fast bench swimmer can't backfill everywhere
# DISABLED by default 2026-06-23. This deflation cancelled a +29 W1 over-prediction
# from roster ghosts/absentees, but the June-17 fixes (excludes bc5feaa, W1 opp-pred
# e97c6fe, imputation 101c86b) removed those at the SOURCE. Pre-presence W1 is now ~+6
# = the W2-5 level (w1_diag.py, 30 sides), so the p=0.60 knockdown double-corrected the
# displayed W1 total to -25 and left the win% inconsistent (deflated total vs raw
# simulate_match pool). Set W1_PRESENCE_MC=1 to restore. See OVERNIGHT_SYSTEMATIC.md.
W1_PRESENCE_MC_ENABLED = os.environ.get("W1_PRESENCE_MC", "0") == "1"

# Forfeit discount (participation-fantasy fix): "team" = per-team measured
# prior-year no-show rate, a float = fixed league rate, "off" = disabled.
# Default "team" since 2026-08-28: isolated effect on the 2025 backtest was
# D17 −9.1 / D16 −2.7 / D15 −1.0 bias with +1.4 (noise) on the D5 control.
FORFEIT_DISCOUNT_MODE = os.environ.get("FORFEIT_DISCOUNT", "team").strip().lower()


def _presence_bench_pick(ev, prof, present, used, already):
    """Best eligible benched swimmer for event `ev`: present, has the stroke, gender
    match, age-eligible (own band or swimming up), under the per-swimmer event cap."""
    age, gender, stroke = parse_event(ev)
    try:
        ai = AGE_GROUP_ORDER.index(age)
    except ValueError:
        ai = len(AGE_GROUP_ORDER)
    best = None; best_t = None
    for nm, p in prof.items():
        if nm in already or not present.get(nm, False):
            continue
        if used.get(nm, 0) >= W1_PRESENCE_MAX_EVENTS:
            continue
        st = p.get("strokes", {}).get(stroke)
        if not st or st.get("mean") is None:
            continue
        if p.get("gender") and p["gender"] != gender:
            continue
        ha = p.get("home_age_group")
        if ha:
            try:
                if AGE_GROUP_ORDER.index(ha) > ai:   # can't swim DOWN to a younger band
                    continue
            except ValueError:
                pass
        if best_t is None or st["mean"] < best_t:
            best, best_t = nm, st["mean"]
    return best


def _presence_adjusted_rows(your_lineup, opp_lineup, your_profiles, opp_profiles,
                            p=W1_PRESENCE_P, ndraw=W1_PRESENCE_NDRAW,
                            n=W1_PRESENCE_MC_N, seed=7):
    """Expected per-event points + total under realistic W1 attendance.
    Returns {event: expected_mc_pts} and the total. The recommended lineup itself
    is NOT changed — only the expected score is made attendance-realistic."""
    import random as _r
    rng = _r.Random(seed)
    pool = list(your_profiles.keys())
    acc = {ev: 0.0 for ev in your_lineup}
    for _ in range(ndraw):
        present = {nm: (rng.random() < p) for nm in pool}
        # re-field our lineup under this draw, backfilling absent lanes
        used = {}
        for ev, data in your_lineup.items():
            names = data.get("swimmers", [])
            kept = []
            for s in names:
                if present.get(s, False):
                    kept.append(s); used[s] = used.get(s, 0) + 1
            while len(kept) < len(names):
                sub = _presence_bench_pick(ev, your_profiles, present, used, set(kept))
                if sub is None:
                    break
                kept.append(sub); used[sub] = used.get(sub, 0) + 1
            _a, _g, stroke = parse_event(ev)
            ymc = [(your_profiles[s]["strokes"][stroke]["mean"],
                    your_profiles[s]["strokes"][stroke]["std"])
                   for s in kept if stroke in your_profiles.get(s, {}).get("strokes", {})]
            omc = [(opp_profiles[s]["strokes"][stroke]["mean"],
                    opp_profiles[s]["strokes"][stroke]["std"])
                   for s in opp_lineup.get(ev, {}).get("swimmers", [])
                   if stroke in opp_profiles.get(s, {}).get("strokes", {})]
            if ymc:
                pts, _ = mc_event(ymc, omc, n=n)
                acc[ev] += pts
    per_event = {ev: v / ndraw for ev, v in acc.items()}
    return per_event, sum(per_event.values())


def _build_lineup_rows(your_lineup, opp_lineup, your_profiles, opp_profiles):
    rows     = []
    mc_total = 0.0
    for event_label, data in your_lineup.items():
        age_group, _, stroke = parse_event(event_label)
        your_mc = [
            (your_profiles[s]["strokes"][stroke]["mean"],
             your_profiles[s]["strokes"][stroke]["std"])
            for s in data["swimmers"]
        ]
        opp_mc = [
            (opp_profiles[s]["strokes"][stroke]["mean"],
             opp_profiles[s]["strokes"][stroke]["std"])
            for s in opp_lineup[event_label]["swimmers"]
            if stroke in opp_profiles.get(s, {}).get("strokes", {})
        ]
        # Pass band/stroke so the displayed score is DQ-aware (8U technical strokes)
        # and uses the same lognormal sampling as the sims.
        mc_pts, win_prob = (mc_event(your_mc, opp_mc, age_group=age_group, stroke=stroke)
                            if your_mc else (0.0, 0.0))
        mc_total += mc_pts
        rows.append({
            "event":    event_label,
            "swimmers": data["swimmers"],
            "opt_pts":  data["expected_points"],
            "mc_pts":   mc_pts,
            "win_pct":  win_prob * 100,
        })
    return rows, mc_total


def _prefer_real_over_imputed_fill(your_lineup, your_profiles, events):
    """Operator rule: never field a LEAGUE-AVERAGE fill (a fabricated time for a swimmer
    with no real data, source startswith 'league_avg') in a slot an unused swimmer with
    an ACTUAL time could take. After the optimizer/polish run, replace each league-avg
    entry with the fastest eligible swimmer who has a real or prior-year time and is
    under the per-swimmer event cap. prior_year_z counts as a real time (left alone).
    A genuine fill with no real replacement stays. Returns (lineup, n_swapped)."""
    from Optimizer import is_eligible, parse_event, MAX_EVENTS

    def _is_league_avg(name, stroke):
        sd = (your_profiles.get(name, {}) or {}).get("strokes", {}).get(stroke) or {}
        return str(sd.get("source", "")).startswith("league_avg")

    used = {}
    for d in your_lineup.values():
        for s in d.get("swimmers", []):
            used[s] = used.get(s, 0) + 1

    n_swapped = 0
    for ev in events:
        d = your_lineup.get(ev)
        if not d:
            continue
        age, gender, stroke = parse_event(ev)
        names = d.get("swimmers", [])
        for idx in range(len(names)):
            name = names[idx]
            if not _is_league_avg(name, stroke):
                continue   # real or prior-year time → leave it
            best, best_t = None, None
            for cand, p in your_profiles.items():
                if cand in names:
                    continue
                sd = (p.get("strokes", {}) or {}).get(stroke) or {}
                m = sd.get("mean")
                if m is None or str(sd.get("source", "")).startswith("league_avg"):
                    continue   # replacement must have a real / prior-year time
                if not is_eligible(p, age, gender, stroke):
                    continue
                if used.get(cand, 0) >= MAX_EVENTS:
                    continue
                if best_t is None or m < best_t:
                    best, best_t = cand, m
            if best is not None:
                used[name] = used.get(name, 0) - 1
                names[idx] = best
                used[best] = used.get(best, 0) + 1
                n_swapped += 1
    return your_lineup, n_swapped


def _forfeit_discount_total(your_lineup, opp_lineup, your_profiles, opp_profiles, forfeit_rate):
    """Down-weight league-average FILLS' expected scoring: a thin team forfeits a fill
    event (scores it as if the fabricated swimmer no-showed) `forfeit_rate` of the time.
    Blends each event's points between with-fill and fill-stripped. Applied PER FILL, so
    weak teams (many fills) take proportionally more total discount automatically — that
    is the participation-awareness. Returns adjusted individual-event mc_total (relay is
    handled separately by the caller). forfeit_rate<=0 -> the plain total (no change)."""
    from Optimizer import parse_event
    rows_w, mc_w = _build_lineup_rows(your_lineup, opp_lineup, your_profiles, opp_profiles)
    if forfeit_rate <= 0:
        return mc_w
    stripped = {}
    any_fill = False
    for ev, d in your_lineup.items():
        try:
            _a, _g, stroke = parse_event(ev)
        except Exception:
            stripped[ev] = {"swimmers": list(d.get("swimmers", [])),
                            "expected_points": d.get("expected_points", 0)}
            continue
        kept = [s for s in d.get("swimmers", [])
                if not str((your_profiles.get(s, {}).get("strokes", {}).get(stroke) or {})
                           .get("source", "")).startswith("league_avg")]
        if len(kept) != len(d.get("swimmers", [])):
            any_fill = True
        stripped[ev] = {"swimmers": kept, "expected_points": d.get("expected_points", 0)}
    if not any_fill:
        return mc_w
    rows_o, _mo = _build_lineup_rows(stripped, opp_lineup, your_profiles, opp_profiles)
    o_by = {r["event"]: r["mc_pts"] for r in rows_o}
    total = 0.0
    for r in rows_w:
        ev = r["event"]; w = r["mc_pts"]; o = o_by.get(ev, w)
        total += (1.0 - forfeit_rate) * w + forfeit_rate * o
    return total


def _forfeit_discount_rows(rows, your_lineup, opp_lineup, your_profiles, opp_profiles,
                           forfeit_rate):
    """Per-row version of _forfeit_discount_total: blends each fill-containing event's
    mc_pts toward its fill-stripped score, in place on the DISPLAYED rows, so the
    discount survives the conserving-scoresheet re-sum (which rebuilds the headline
    total from the rows — anything applied only to the total gets discarded there).
    Returns (rows, adjusted_total, n_discounted_events)."""
    from Optimizer import parse_event
    plain = sum(float(r.get("mc_pts", 0.0)) for r in rows)
    if forfeit_rate <= 0:
        return rows, plain, 0
    stripped = {}
    fill_evs = set()
    for ev, d in your_lineup.items():
        try:
            _a, _g, stroke = parse_event(ev)
        except Exception:
            stripped[ev] = {"swimmers": list(d.get("swimmers", [])),
                            "expected_points": d.get("expected_points", 0)}
            continue
        kept = [s for s in d.get("swimmers", [])
                if not str((your_profiles.get(s, {}).get("strokes", {}).get(stroke) or {})
                           .get("source", "")).startswith("league_avg")]
        if len(kept) != len(d.get("swimmers", [])):
            fill_evs.add(ev)
        stripped[ev] = {"swimmers": kept, "expected_points": d.get("expected_points", 0)}
    if not fill_evs:
        return rows, plain, 0
    rows_o, _mo = _build_lineup_rows(stripped, opp_lineup, your_profiles, opp_profiles)
    o_by = {r["event"]: r["mc_pts"] for r in rows_o}
    for r in rows:
        ev = r["event"]
        if ev in fill_evs and ev in o_by:
            r["mc_pts"] = ((1.0 - forfeit_rate) * float(r["mc_pts"])
                           + forfeit_rate * float(o_by[ev]))
    return rows, sum(float(r.get("mc_pts", 0.0)) for r in rows), len(fill_evs)


def _demote_hopeless_swimups(your_lineup, opp_lineup, your_profiles, opp_profiles, events,
                             margin_frac=0.4):
    """Coaching option (OFF by default): a swim-up (a swimmer fielded one age group ABOVE
    their own) is only worth it if they can actually score. If 3+ swimmers in the event
    field (yours + the opponent's) are CLEARLY faster — each beats the swim-up by more
    than margin_frac of the swim-up's own std, i.e. not a near-tie — the swim-up is locked
    out of the top 3, so swap them for the best available NATIVE-age-group swimmer and
    free the young swimmer for their own group. Genuine scoring swim-ups, and swim-ups
    with no native replacement available, are left exactly as the optimizer set them.
    Returns (lineup, n_demoted)."""
    from Optimizer import is_eligible, parse_event, AGE_GROUP_ORDER, MAX_EVENTS

    def _mean_std(profs, nm, stroke):
        sd = (profs.get(nm, {}) or {}).get("strokes", {}).get(stroke) or {}
        return sd.get("mean"), sd.get("std")

    used = {}
    for d in your_lineup.values():
        for s in d.get("swimmers", []):
            used[s] = used.get(s, 0) + 1

    n_demoted = 0
    for ev in events:
        d = your_lineup.get(ev)
        if not d:
            continue
        try:
            age, gender, stroke = parse_event(ev)
            ev_idx = AGE_GROUP_ORDER.index(age)
        except Exception:
            continue
        names = d.get("swimmers", [])
        field = []   # combined field mean times (yours + opp) in this event
        for nm in names:
            m, _s = _mean_std(your_profiles, nm, stroke)
            if m is not None:
                field.append(m)
        for nm in opp_lineup.get(ev, {}).get("swimmers", []):
            m, _s = _mean_std(opp_profiles, nm, stroke)
            if m is not None:
                field.append(m)
        for idx in range(len(names)):
            nm = names[idx]
            home = (your_profiles.get(nm, {}) or {}).get("home_age_group")
            if home not in AGE_GROUP_ORDER or AGE_GROUP_ORDER.index(home) >= ev_idx:
                continue   # native to the event (or not a one-up swim-up) — leave it
            m, s = _mean_std(your_profiles, nm, stroke)
            if m is None:
                continue
            margin = margin_frac * (s or 0.0)
            clearly_faster = sum(1 for t in field if t < m - margin)
            if clearly_faster < 3:
                continue   # realistic top-3 shot — a swim-up worth doing, keep it
            best, best_t = None, None   # best native-age replacement not already entered
            for cand, cp in your_profiles.items():
                if cand in names or cp.get("home_age_group") != age:
                    continue
                cm, _cs = _mean_std(your_profiles, cand, stroke)
                if cm is None or not is_eligible(cp, age, gender, stroke):
                    continue
                if used.get(cand, 0) >= MAX_EVENTS:
                    continue
                if best_t is None or cm < best_t:
                    best, best_t = cand, cm
            if best is not None:
                used[nm] = used.get(nm, 0) - 1
                names[idx] = best
                used[best] = used.get(best, 0) + 1
                n_demoted += 1
    return your_lineup, n_demoted


def _ladder_fill_slots(your_lineup, opp_lineup, your_profiles, opp_profiles, events):
    """Follow the ladder: never leave a faster swimmer on the bench while a slower one
    fills a slot in the same event. For each event, if a benched, eligible swimmer
    with a spare event (< MAX_EVENTS) is faster than a fielded swimmer AND swapping
    them in doesn't lower the event's expected points, make the swap. This fixes a
    hopeless time parked in a throwaway slot (which scores the same as any last-place
    body) while a much-faster swimmer with an open event sat out. Returns
    (lineup, n_swapped)."""
    from Optimizer import is_eligible, parse_event, MAX_EVENTS

    def _mean(profs, nm, stroke):
        # Only REAL times count for the ladder — a league-average guess isn't a real
        # swim (and would be blanked in the display), so never seed or swap on one.
        sd = (profs.get(nm, {}) or {}).get("strokes", {}).get(stroke) or {}
        if str(sd.get("source", "")).startswith("league_avg"):
            return None
        return sd.get("mean")

    def _event_pts(names, ev, age, stroke):
        your_mc = [(your_profiles[s]["strokes"][stroke]["mean"],
                    your_profiles[s]["strokes"][stroke]["std"])
                   for s in names if stroke in your_profiles.get(s, {}).get("strokes", {})]
        if not your_mc:
            return 0.0
        opp_mc = [(opp_profiles[s]["strokes"][stroke]["mean"],
                   opp_profiles[s]["strokes"][stroke]["std"])
                  for s in opp_lineup.get(ev, {}).get("swimmers", [])
                  if stroke in opp_profiles.get(s, {}).get("strokes", {})]
        pts, _ = mc_event(your_mc, opp_mc, age_group=age, stroke=stroke)
        return pts

    used = {}
    for d in your_lineup.values():
        for s in d.get("swimmers", []):
            used[s] = used.get(s, 0) + 1

    n_swapped = 0
    for ev in events:
        d = your_lineup.get(ev)
        if not d:
            continue
        try:
            age, gender, stroke = parse_event(ev)
        except Exception:
            continue
        names = d.get("swimmers", [])
        for _ in range(len(names) + 2):     # bounded; converges to the fastest set
            fielded = [(i, names[i], _mean(your_profiles, names[i], stroke))
                       for i in range(len(names))]
            fielded = [(i, nm, m) for i, nm, m in fielded if m is not None]
            if not fielded:
                break
            i_slow, nm_slow, m_slow = max(fielded, key=lambda x: x[2])
            best_a, best_am = None, None       # fastest benched, eligible, has a spare event
            for cand, cp in your_profiles.items():
                if cand in names or used.get(cand, 0) >= MAX_EVENTS:
                    continue
                if not is_eligible(cp, age, gender, stroke):
                    continue
                cm = _mean(your_profiles, cand, stroke)
                if cm is None:
                    continue
                if best_am is None or cm < best_am:
                    best_a, best_am = cand, cm
            if best_a is None or best_am >= m_slow:
                break     # nobody faster is available
            trial = list(names); trial[i_slow] = best_a
            if _event_pts(trial, ev, age, stroke) < _event_pts(names, ev, age, stroke) - 1e-9:
                break     # swap would cost points (a variance edge case) — leave it
            used[nm_slow] = used.get(nm_slow, 0) - 1
            names[i_slow] = best_a
            used[best_a] = used.get(best_a, 0) + 1
            n_swapped += 1
    return your_lineup, n_swapped


def _prefer_natural_events(your_lineup, opp_lineup, your_profiles, opp_profiles, events,
                           margin=1.0):
    """Coaching preference: field swimmers in their NATURAL (faster) strokes rather than
    parking them in an off-event, when it costs nothing on the scoreboard.

    Only reshuffles swimmers who CAN'T score in an event anyway (3+ of the combined
    field are clearly faster, by `margin` seconds), and only swaps two of them between
    SAME-DISTANCE events (so raw times are comparable and the swap is apples-to-apples).
    Because neither swimmer scores in either arrangement, the team's points are provably
    unchanged — it just moves a freestyler out of breast, etc. A swap is made only when
    it lowers how far each swimmer sits above their own best same-distance time.
    Returns (lineup, n_swaps)."""
    from Optimizer import parse_event, is_eligible

    def _mean(profs, nm, st):
        return (profs.get(nm, {}) or {}).get("strokes", {}).get(st, {}).get("mean")

    def _best_at(nm, dist):
        ms = [v.get("mean") for k, v in (your_profiles.get(nm, {}) or {}).get("strokes", {}).items()
              if v.get("mean") is not None and k.split("-")[0] == dist]
        return min(ms) if ms else None

    def _field(ev, st):
        out = []
        for s in your_lineup.get(ev, {}).get("swimmers", []):
            m = _mean(your_profiles, s, st)
            if m is not None: out.append(m)
        for s in opp_lineup.get(ev, {}).get("swimmers", []):
            m = _mean(opp_profiles, s, st)
            if m is not None: out.append(m)
        return out

    def _cant_score(t, field):
        return sum(1 for f in field if f < t - margin) >= 3   # clearly out of the top 3

    parsed = {}
    for ev in events:
        try:
            parsed[ev] = parse_event(ev)
        except Exception:
            pass

    n_swaps = 0
    for _ in range(8):                     # bounded local search
        improved = False
        evs = [e for e in your_lineup if e in parsed]
        for xi in range(len(evs)):
            for yi in range(xi + 1, len(evs)):
                e1, e2 = evs[xi], evs[yi]
                a1, g1, s1 = parsed[e1]; a2, g2, s2 = parsed[e2]
                if s1 == s2 or s1.split("-")[0] != s2.split("-")[0]:
                    continue                # need distinct strokes at the same distance
                dist = s1.split("-")[0]
                d1 = your_lineup[e1]["swimmers"]; d2 = your_lineup[e2]["swimmers"]
                f1, f2 = _field(e1, s1), _field(e2, s2)
                for ai in range(len(d1)):
                    A = d1[ai]
                    for bi in range(len(d2)):
                        B = d2[bi]
                        if A == B or A in d2 or B in d1:
                            continue
                        mA1 = _mean(your_profiles, A, s1); mA2 = _mean(your_profiles, A, s2)
                        mB1 = _mean(your_profiles, B, s1); mB2 = _mean(your_profiles, B, s2)
                        if None in (mA1, mA2, mB1, mB2):
                            continue
                        if not is_eligible(your_profiles.get(A, {}), a2, g2, s2):
                            continue
                        if not is_eligible(your_profiles.get(B, {}), a1, g1, s1):
                            continue
                        # both swimmers must be non-scoring in BOTH events → point-neutral
                        if not (_cant_score(mA1, f1) and _cant_score(mA2, f2)
                                and _cant_score(mB1, f1) and _cant_score(mB2, f2)):
                            continue
                        bA = _best_at(A, dist); bB = _best_at(B, dist)
                        if not bA or not bB:
                            continue
                        cur = (mA1 - bA) + (mB2 - bB)
                        new = (mA2 - bA) + (mB1 - bB)
                        if new < cur - 1e-6:               # better natural fit
                            d1[ai], d2[bi] = B, A
                            n_swaps += 1
                            improved = True
        if not improved:
            break
    return your_lineup, n_swaps


def _assign_lanes(our_names, our_profiles, opp_names, opp_profiles, stroke, your_is_home,
                  fillers=None):
    """
    Seat both teams' swimmers into NVSL 6-lane positions.
    Home gets odd lanes (1,3,5); away gets even (2,4,6). Fastest in centermost lane,
    outward by speed:  home order = [3, 5, 1],  away order = [4, 2, 6].
    `fillers` (names with no real time for this stroke) seat into any leftover OUR
    lanes, flagged no_real_time and shown without a time.
    Returns 6 lane dicts in lane-number order.
    """
    HOME_LANES = [3, 5, 1]   # fast → slow
    AWAY_LANES = [4, 2, 6]   # fast → slow

    def _seeded(names, profiles):
        out = []
        for n in names:
            t = (profiles.get(n, {}) or {}).get("strokes", {}).get(stroke, {}).get("mean")
            if t is not None:
                out.append((n, float(t)))
        out.sort(key=lambda x: x[1])
        return out

    if your_is_home:
        home_seeds = _seeded(our_names, our_profiles); home_team = "us"
        away_seeds = _seeded(opp_names, opp_profiles); away_team = "opp"
    else:
        home_seeds = _seeded(opp_names, opp_profiles); home_team = "opp"
        away_seeds = _seeded(our_names, our_profiles); away_team = "us"

    lanes = [None] * 6
    for i, (name, t) in enumerate(home_seeds[:3]):
        ln = HOME_LANES[i]
        lanes[ln - 1] = {"lane": ln, "team": home_team, "swimmer": name, "time": t}
    for i, (name, t) in enumerate(away_seeds[:3]):
        ln = AWAY_LANES[i]
        lanes[ln - 1] = {"lane": ln, "team": away_team, "swimmer": name, "time": t}
    for i in range(6):
        if lanes[i] is None:
            ln = i + 1
            team = home_team if ln % 2 == 1 else away_team
            lanes[i] = {"lane": ln, "team": team, "swimmer": None, "time": None}
    # Seat fillers (no real time for this stroke) into leftover OUR lanes so every
    # event shows a full 3. Flagged no_real_time; rendered without a time.
    if fillers:
        fi = 0
        for ln in lanes:
            if fi >= len(fillers):
                break
            if ln["team"] == "us" and ln["swimmer"] is None:
                ln.update(swimmer=fillers[fi], time=None, no_real_time=True)
                fi += 1
    return lanes


def _compute_event_fillers(lineup, profiles):
    """Suggest fillers so every event fields a full 3. A filler is the fastest
    available eligible swimmer who has NO real time for that event's stroke (they
    race a different stroke), under the per-swimmer 2-event cap. Home-band swimmers
    are preferred over swim-ups. Returns {event: [name, ...]}. DISPLAY-ONLY — these
    are flagged 'no real time' in the lanes and are NOT scored (we can't predict a
    swimmer with no time for the stroke), they just tell the coach who to put in
    the empty lane on meet day."""
    from collections import defaultdict
    from Optimizer import MAX_PER_EVENT, MAX_EVENTS, AGE_GROUP_ORDER
    usage = defaultdict(int)
    for ev, d in lineup.items():
        for s in d.get("swimmers", []):
            usage[s] += 1
    out = {}
    for ev in sorted(lineup):
        cur = list(lineup[ev].get("swimmers", []))
        need = MAX_PER_EVENT - len(cur)
        if need <= 0:
            continue
        try:
            age, gender, stroke = parse_event(ev)
            ev_idx = AGE_GROUP_ORDER.index(age)
        except (Exception, ValueError):
            continue
        cands = []
        for nm, p in profiles.items():
            if nm in cur or p.get("gender") != gender:
                continue
            home = p.get("home_age_group")
            if home not in AGE_GROUP_ORDER:
                continue
            home_idx = AGE_GROUP_ORDER.index(home)
            if not (0 <= ev_idx - home_idx <= 1):      # eligible by age (home or swim-up)
                continue
            strokes = p.get("strokes", {})
            if not strokes or stroke in strokes:        # filler = NO real time for this stroke
                continue
            if usage[nm] >= MAX_EVENTS:
                continue
            best = min((sd.get("mean", 9999) for sd in strokes.values()), default=9999)
            cands.append(((ev_idx - home_idx), best, nm))   # home-band first, then fastest
        cands.sort()
        picked = [nm for _su, _b, nm in cands[:need]]
        for nm in picked:
            usage[nm] += 1
        if picked:
            out[ev] = picked
    return out


def _build_event_payload(your_lineup, your_profiles,
                         opp_predicted, opp_lastweek, opp_profiles,
                         rows, your_is_home, fillers=None, rows_lastweek=None):
    """
    Per-event view-data for lineup.html.
    Each entry holds two lane layouts (predicted opp vs last-week opp) + scoring.
    `fillers` ({event: [names]}) seats no-real-time swimmers into empty OUR lanes
    of the PREDICTED layout so every event shows a full 3.
    `rows_lastweek` (optional, from _build_lineup_rows vs the last-week opp) adds
    per-event mc_pts/win% for that matchup so the toggle can swap the numbers.
    """
    fillers = fillers or {}
    rows_by_event = {r["event"]: r for r in (rows or [])}
    rows_lw_by_event = {r["event"]: r for r in (rows_lastweek or [])}

    def _mark_swim_ups(lanes, event_age):
        for lane in lanes:
            if lane["team"] == "us" and lane["swimmer"]:
                prof = your_profiles.get(lane["swimmer"], {}) or {}
                lane["is_swim_up"] = prof.get("home_age_group", event_age) != event_age
            else:
                lane["is_swim_up"] = False
        return lanes

    out = []
    for event_label, data in your_lineup.items():
        _, _, stroke = parse_event(event_label)
        event_age = event_label.split()[0] if event_label else ""
        our_names = data.get("swimmers", []) or []

        # "No real time — pick one": a LEAGUE-AVERAGE fill is a real roster swimmer
        # handed a fabricated baseline time (source startswith 'league_avg'). Among
        # several no-time swimmers these all tie, so which one the optimizer seated
        # is arbitrary. Rather than present that guess as a decision, flag the lane
        # so the Results page shows a "pick one" picker and the coach chooses the
        # swimmer inline. prior_year_z estimates are real prior-season swims — kept.
        def _flag_pick_one(lanes):
            # A league-average fill is a real roster kid with no real time for this
            # stroke. Keep them SHOWN (so the lane has a body) but flag the lane as a
            # no-time pick — the fabricated time is hidden and the coach can swap them.
            for _lane in (lanes or []):
                if _lane.get("team") == "us" and _lane.get("swimmer"):
                    _sd = (your_profiles.get(_lane["swimmer"], {}) or {}).get("strokes", {}).get(stroke) or {}
                    if str(_sd.get("source", "")).startswith("league_avg"):
                        _lane["pick_one"] = True
                        _lane["no_real_time"] = True
                        _lane["time"] = None   # hide the fabricated time; keep the name
            return lanes

        opp_pred_names = (opp_predicted or {}).get(event_label, {}).get("swimmers", []) or []
        # Fill every lane with a body: real swimmers seed by time, and any leftover OUR
        # lane gets the fastest available under-2-event roster kid (flagged no-time).
        # All no-time lanes stay editable dropdowns on the Results page.
        lanes_pred = _flag_pick_one(_mark_swim_ups(
            _assign_lanes(our_names, your_profiles,
                          opp_pred_names, opp_profiles, stroke, your_is_home,
                          fillers=fillers.get(event_label)),
            event_age,
        ))

        lanes_lw = None
        if opp_lastweek and event_label in opp_lastweek:
            opp_lw_names = (opp_lastweek[event_label] or {}).get("swimmers", []) or []
            lanes_lw = _flag_pick_one(_mark_swim_ups(
                _assign_lanes(our_names, your_profiles,
                              opp_lw_names, opp_profiles, stroke, your_is_home),
                event_age,
            ))

        row = rows_by_event.get(event_label, {})
        row_lw = rows_lw_by_event.get(event_label, {})
        out.append({
            "event":   event_label,
            "stroke":  stroke,
            "mc_pts":  row.get("mc_pts", 0),
            "opp_pts": row.get("opp_pts", 0),
            "opt_pts": row.get("opt_pts", 0),
            "win_pct": row.get("win_pct", 0),
            "mc_pts_lastweek":  row_lw.get("mc_pts") if lanes_lw else None,
            "win_pct_lastweek": row_lw.get("win_pct") if lanes_lw else None,
            "lanes_predicted": lanes_pred,
            "lanes_lastweek":  lanes_lw,
        })

    # Order events youngest → oldest by age group. Stable sort preserves the
    # existing sub-order (gender/stroke) within each age group. Unknown ages
    # sort to the end.
    def _age_key(entry):
        age = entry["event"].split()[0] if entry["event"] else ""
        try:
            return AGE_GROUP_ORDER.index(age)
        except ValueError:
            return len(AGE_GROUP_ORDER)
    out.sort(key=_age_key)
    return out


def _build_relay_results(your_profiles, opp_profiles):
    # RELAY_STD_SCALE (staircase fix #1, ~41% of the opp-strength gap): relay leg
    # stds inherit individual-event noise, but four independent draws make the MC
    # chronically unsure — it can never predict the relay sweeps that actually
    # happen in lopsided meets (weak opp: 45.5 actual relay pts vs 35.1 predicted;
    # strong: 17.1 vs 25.3 — staircase_decomp_findings.md). Scaling leg stds
    # sharpens win probs AND the margin sim through one physical knob.
    # Default 1.0 = current behavior; validated values come from the lite sweep.
    _rss = float(os.environ.get("RELAY_STD_SCALE", "0.5"))  # calibrated default (lite+full validated)
    relay_results = {}
    for gender in ["Boys", "Girls"]:
        your_age  = optimize_age_relays(your_profiles, gender)
        opp_age   = optimize_age_relays(opp_profiles,  gender)
        your_mix  = optimize_mixed_relay(your_profiles, gender)
        opp_mix   = optimize_mixed_relay(opp_profiles,  gender)

        age_relays = []
        for relay_name, rdef in AGE_RELAY_DEFS.items():
            your_legs = your_age[relay_name]
            opp_legs  = opp_age[relay_name]
            your_tups = [(l["mean"], l["std"] * _rss) for l in your_legs if l["mean"]]
            opp_tups  = [(l["mean"], l["std"] * _rss) for l in opp_legs  if l["mean"]]
            win_prob, exp_pts = monte_carlo_relay(your_tups, opp_tups)
            age_relays.append({
                "name":       relay_name,
                "legs":       your_legs,
                "your_tups":  your_tups,
                "opp_tups":   opp_tups,
                "total_time": sum(l["mean"] for l in your_legs if l["mean"]),
                "win_pct":    win_prob * 100,
                "exp_pts":    exp_pts,
            })

        your_mix_tups = [(your_mix[s]["mean"], your_mix[s]["std"] * _rss) for s in MIXED_SLOTS if your_mix[s]["mean"]]
        opp_mix_tups  = [(opp_mix[s]["mean"],  opp_mix[s]["std"] * _rss)  for s in MIXED_SLOTS if opp_mix[s]["mean"]]
        win_prob, exp_pts = monte_carlo_relay(your_mix_tups, opp_mix_tups)
        mixed = {
            "total_time": sum(your_mix[s]["mean"] for s in MIXED_SLOTS if your_mix[s]["mean"]),
            "your_tups":  your_mix_tups,
            "opp_tups":   opp_mix_tups,
            "win_pct":    win_prob * 100,
            "exp_pts":    exp_pts,
            "legs": [{"slot": s, "swimmer": your_mix[s]["swimmer"],
                      "mean": your_mix[s]["mean"], "std": your_mix[s]["std"]}
                     for s in MIXED_SLOTS],
        }
        relay_results[gender] = {"age_relays": age_relays, "mixed": mixed}
    return relay_results



if __name__ == "__main__":
    app.run(debug=True, port=5000, threaded=True)
